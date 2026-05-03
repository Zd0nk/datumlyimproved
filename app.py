"""
Datumly — Data-driven FPL Intelligence
=======================================
Data sources:
  1. FPL API (bootstrap-static, fixtures, live GW) — player stats, prices, xG/xA, form, set pieces
  2. football-data.co.uk — betting odds → match probabilities
  3. Club Elo (api.clubelo.com) — dynamic team strength ratings
  4. Custom xPts model — blends all sources + form-weighting + over/underperformance regression

Optimisation:
  - PuLP MILP solver for squad selection (not greedy)
  - Proper constraints: budget, max 3/team, formation, 15-man squad
  - Horizon-aware transfer planner with escalating hit thresholds

UI: Streamlit with 6 tabs
"""

import streamlit as st
import streamlit.components.v1 as components
import requests
import pandas as pd
import numpy as np
import math
import json
import re
import unicodedata
from pulp import (
    LpProblem, LpMaximize, LpVariable, lpSum, LpStatus, value,
    PULP_CBC_CMD,
)
from datetime import datetime, timedelta
from io import StringIO
from concurrent.futures import ThreadPoolExecutor
import warnings
warnings.filterwarnings("ignore")

# ============================================================
# LEAGUE CONFIGURATION SYSTEM
# ============================================================
# All league-specific settings are here. The rest of the app reads from
# the active config, making it easy to add new leagues.

LEAGUE_CONFIGS = {
    "FPL": {
        "name": "Fantasy Premier League",
        "short_name": "FPL",
        "country": "England",
        "base_url": "https://fantasy.premierleague.com/api",
        "football_data_url": "https://www.football-data.co.uk/mmz4281/2526/E0.csv",
        "odds_api_league": "soccer_epl",
        "n_teams": 20,
        "season_gws": 38,
        "half_season_gw": 19,  # chips refresh after this GW
        "budget": 1000,  # 100.0m in 0.1m units
        "currency": "£",
        "currency_unit": "m",
        "max_per_team": 3,
        "squad_size": 15,
        "pts_goal": {1: 10, 2: 6, 3: 5, 4: 4},
        "pts_assist": 3,
        "pts_cs": {1: 4, 2: 4, 3: 1, 4: 0},
        "pts_appearance": 2,
        "pts_bonus_avg": 0.35,
        "transfer_cost": 4,
        "max_banked_ft": 5,
        "league_avg_goals": 1.35,
        "chips": {
            "wildcard": {"name": "Wildcard", "icon": "🃏", "per_half": True, "count": 1},
            "free_hit": {"name": "Free Hit", "icon": "⚡", "per_half": True, "count": 1},
            "triple_captain": {"name": "Triple Captain", "icon": "👑", "per_half": True, "count": 1},
            "bench_boost": {"name": "Bench Boost", "icon": "💪", "per_half": True, "count": 1},
        },
        "chip_api_names": {
            "wildcard": "wildcard", "freehit": "freehit",
            "3xc": "triple_captain", "bboost": "bench_boost",
        },
        "has_defcon": True,
        "page_title": "Datumly - FPL Intelligence",
    },
    "Allsvenskan": {
        "name": "Allsvenskan Fantasy",
        "short_name": "ASV",
        "country": "Sweden",
        "base_url": "https://fantasy.allsvenskan.se/api",
        "football_data_url": None,  # TODO: find Swedish league odds source
        "odds_api_league": "soccer_sweden_allsvenskan",
        "n_teams": 16,
        "season_gws": 30,
        "half_season_gw": 15,
        "budget": 1000,  # 100.0m kr in 0.1m units
        "currency": "kr",
        "currency_unit": "m",
        "max_per_team": 3,
        "squad_size": 15,
        "pts_goal": {1: 6, 2: 6, 3: 5, 4: 5},
        "pts_assist": 3,
        "pts_cs": {1: 5, 2: 4, 3: 1, 4: 0},  # GK gets 5 for CS (higher than FPL)
        "pts_appearance": 2,
        "pts_bonus_avg": 0.55,  # ~16 bonus pts per game avg — much higher than FPL
        "pts_save_divisor": 2,  # 1pt per 2 saves (FPL uses 3)
        "pts_winning_goal": 1,  # +1 for scoring the winning goal (unique to ASV)
        "gc_threshold": 2,  # GK -1 per goal after first 2 conceded
        "transfer_cost": 4,
        "max_banked_ft": 5,
        "league_avg_goals": 1.25,  # Allsvenskan averages slightly fewer goals
        "chips": {
            "wildcard": {"name": "Wildcard", "icon": "🃏", "per_half": True, "count": 1},
            "park_the_bus": {"name": "Park the Bus", "icon": "🚌", "per_half": False, "count": 1},
            "dynamic_duo": {"name": "Dynamic Duo", "icon": "👥", "per_half": False, "count": 1},
            "loan_rangers": {"name": "Loan Rangers", "icon": "🔄", "per_half": False, "count": 1},
        },
        "chip_api_names": {
            "wildcard": "wildcard", "freehit": "loan_rangers",
            "3xc": "dynamic_duo", "bboost": "park_the_bus",
        },
        "has_frikort": True,  # additional free card per half-season
        "has_defcon": False,  # Allsvenskan doesn't have DefCon points
        "page_title": "Datumly - Allsvenskan Intelligence",
    },
    "Eliteserien": {
        "name": "Eliteserien Fantasy",
        "short_name": "ELS",
        "country": "Norway",
        "base_url": "https://fantasy.eliteserien.no/api",
        "football_data_url": None,
        "odds_api_league": "soccer_norway_eliteserien",
        "n_teams": 16,
        "season_gws": 30,
        "half_season_gw": 15,
        "budget": 1000,
        "currency": "kr",
        "currency_unit": "m",
        "max_per_team": 3,
        "squad_size": 15,
        "pts_goal": {1: 6, 2: 6, 3: 5, 4: 5},
        "pts_assist": 3,
        "pts_cs": {1: 4, 2: 4, 3: 1, 4: 0},
        "pts_appearance": 2,
        "pts_bonus_avg": 0.35,
        "transfer_cost": 4,
        "max_banked_ft": 5,
        "league_avg_goals": 1.40,
        "chips": {
            "wildcard": {"name": "Wildcard", "icon": "🃏", "per_half": True, "count": 1},
            "park_the_bus": {"name": "Park the Bus", "icon": "🚌", "per_half": False, "count": 1},
            "dynamic_duo": {"name": "Dynamic Duo", "icon": "👥", "per_half": False, "count": 1},
            "loan_rangers": {"name": "Loan Rangers", "icon": "🔄", "per_half": False, "count": 1},
        },
        "chip_api_names": {
            "wildcard": "wildcard", "freehit": "loan_rangers",
            "3xc": "dynamic_duo", "bboost": "park_the_bus",
        },
        "has_frikort": True,
        "has_defcon": False,
        "page_title": "Datumly - Eliteserien Intelligence",
    },
}

# Default league (will be overridden by UI selection)
ACTIVE_LEAGUE = "FPL"
st.set_page_config(
    page_title="Datumly - FPL Analytics",
    page_icon="favicon.svg",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# League selector (persisted in session state)
if "active_league" not in st.session_state:
    st.session_state["active_league"] = "FPL"

def get_league_config():
    """Get the active league configuration."""
    return LEAGUE_CONFIGS[st.session_state["active_league"]]

def get_chip_labels(lc=None):
    """Build chip display labels from league config."""
    if lc is None:
        lc = get_league_config()
    return {k: f'{v["icon"]} {v["name"]}' for k, v in lc["chips"].items()}

# Set active constants from league config
LC = get_league_config()

FPL_BASE = LC["base_url"]
FOOTBALL_DATA_URL = LC.get("football_data_url")

POS_MAP = {1: "GKP", 2: "DEF", 3: "MID", 4: "FWD"}
POS_FULL = {1: "Goalkeeper", 2: "Defender", 3: "Midfielder", 4: "Forward"}

# Scoring constants (from league config)
PTS_GOAL = LC["pts_goal"]
PTS_ASSIST = LC["pts_assist"]
PTS_CS = LC["pts_cs"]
PTS_APPEARANCE = LC["pts_appearance"]
PTS_BONUS_AVG = LC["pts_bonus_avg"]

# League-specific settings
league_avg_goals = LC["league_avg_goals"]

DATUMLY_LOGO_SVG = '''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 232 80" height="120" role="img" aria-label="Datumly">
<defs>
<linearGradient id="dl_pink" x1="0" y1="0" x2="1" y2="1">
<stop offset="0%" stop-color="#f02d6e"/>
<stop offset="100%" stop-color="#ff6b8a"/>
</linearGradient>
</defs>
<g transform="translate(4 18) scale(0.733)">
<rect x="1" y="1" width="58" height="58" rx="14" fill="url(#dl_pink)"/>
<path d="M19 14 L19 46 L31 46 C40 46 45.5 39.5 45.5 30 C45.5 20.5 40 14 31 14 Z" fill="#0a0e17"/>
<path d="M22 36 L27 31.5 L32 27 L37 22" stroke="#ff6b8a" stroke-width="2.4" fill="none" stroke-linecap="round" stroke-linejoin="round"/>
<circle cx="37" cy="22" r="2.8" fill="#ff6b8a"/>
</g>
<text x="58" y="55" font-family="'Plus Jakarta Sans', 'Inter', sans-serif" font-weight="800" font-size="42" letter-spacing="-1.4" fill="#e7ecf4">datum<tspan fill="url(#dl_pink)">ly</tspan></text>
</svg>'''
POS_FULL = {1: "Goalkeeper", 2: "Defender", 3: "Midfielder", 4: "Forward"}

# Scoring constants already set from LC above

# Team colours: FPL short name -> (primary, secondary, text_colour)
# Separate dicts per league, selected based on active config
TEAM_COLOURS_FPL = {
    "ARS": ("#EF0107", "#FFFFFF", "#FFFFFF"),  # Red, white
    "AVL": ("#670E36", "#95BFE5", "#FFFFFF"),  # Claret, sky blue
    "BOU": ("#DA291C", "#000000", "#FFFFFF"),  # Red, black
    "BRE": ("#FFD700", "#E30613", "#000000"),  # Yellow/red (bee stripes)
    "BHA": ("#0057B8", "#FFFFFF", "#FFFFFF"),  # Blue, white
    "CHE": ("#034694", "#FFFFFF", "#FFFFFF"),  # Blue, white
    "CRY": ("#1B458F", "#C4122E", "#FFFFFF"),  # Blue, red
    "EVE": ("#003399", "#FFFFFF", "#FFFFFF"),  # Blue, white
    "FUL": ("#FFFFFF", "#000000", "#000000"),  # White, black
    "IPS": ("#0044AA", "#FFFFFF", "#FFFFFF"),  # Blue, white
    "LEI": ("#003090", "#FDBE11", "#FFFFFF"),  # Blue, gold
    "LIV": ("#C8102E", "#FFFFFF", "#FFFFFF"),  # Red, white
    "MCI": ("#6CABDD", "#FFFFFF", "#FFFFFF"),  # Sky blue, white
    "MUN": ("#DA291C", "#FBE122", "#FFFFFF"),  # Red, yellow
    "NEW": ("#241F20", "#FFFFFF", "#FFFFFF"),  # Black, white
    "NFO": ("#DD0000", "#FFFFFF", "#FFFFFF"),  # Red, white
    "SOU": ("#D71920", "#FFFFFF", "#FFFFFF"),  # Red, white
    "TOT": ("#FFFFFF", "#132257", "#132257"),  # White, navy
    "WHU": ("#7A263A", "#1BB1E7", "#FFFFFF"),  # Claret, blue
    "WOL": ("#FDB913", "#000000", "#000000"),  # Gold, black
    "LEE": ("#FFFFFF", "#1D428A", "#1D428A"),  # White, blue
    "BUR": ("#6C1D45", "#99D6EA", "#FFFFFF"),  # Claret, blue
    "SUN": ("#EB172B", "#FFFFFF", "#FFFFFF"),  # Red, white
    "SHU": ("#EE2737", "#FFFFFF", "#FFFFFF"),  # Red, white
}

TEAM_COLOURS_ASV = {
    "AIK": ("#1A1A2E", "#FFD700", "#FFD700"),   # Black/gold
    "HÄC": ("#1B3A5C", "#FFD700", "#FFFFFF"),   # Navy/gold — BK Häcken
    "DEG": ("#FF0000", "#FFFFFF", "#FFFFFF"),    # Red/white — Degerfors IF
    "DJU": ("#003DA5", "#E31837", "#FFFFFF"),    # Blue/red — Djurgårdens IF
    "GAI": ("#006B3F", "#FFFFFF", "#FFFFFF"),    # Green/white — GAIS
    "HAM": ("#006B3F", "#FFFFFF", "#FFFFFF"),    # Green/white — Hammarby IF
    "ELF": ("#FFD700", "#000000", "#000000"),    # Yellow/black — IF Elfsborg
    "IFG": ("#1B4E8F", "#FFFFFF", "#FFFFFF"),    # Blue/white — IFK Göteborg
    "KAL": ("#E31837", "#FFFFFF", "#FFFFFF"),    # Red/white — Kalmar FF
    "MFF": ("#87CEEB", "#FFFFFF", "#000000"),    # Sky blue/white — Malmö FF
    "MJÄ": ("#FFD700", "#000000", "#000000"),    # Yellow/black — Mjällby AIF
    "SIR": ("#003DA5", "#FFFFFF", "#FFFFFF"),    # Blue/white — IK Sirius
    "VSK": ("#000000", "#FFFFFF", "#FFFFFF"),    # Black/white — Västerås SK
    "ÖRG": ("#E31837", "#003DA5", "#FFFFFF"),    # Red/blue — Örgryte IS
    "BPK": ("#E31837", "#000000", "#FFFFFF"),    # Red/black — IF Brommapojkarna
    "HBK": ("#003DA5", "#FFFFFF", "#FFFFFF"),    # Blue/white — Halmstads BK
}

TEAM_COLOURS_ELS = {
    "BOD": ("#FFD700", "#000000", "#000000"),    # Yellow/black — Bodø/Glimt
    "BRA": ("#E31837", "#FFFFFF", "#FFFFFF"),    # Red/white — SK Brann
    "FRE": ("#FFFFFF", "#E31837", "#E31837"),    # White/red — Fredrikstad FK
    "HAM": ("#FFFFFF", "#006B3F", "#006B3F"),    # White/green — HamKam
    "KFU": ("#000000", "#FFD700", "#FFD700"),    # Black/gold — KFUM Oslo
    "KRI": ("#1B1B1B", "#FFFFFF", "#FFFFFF"),    # Black/white — Kristiansund BK
    "LIL": ("#FFD700", "#000000", "#000000"),    # Yellow/black — Lillestrøm SK
    "MOL": ("#003DA5", "#FFFFFF", "#FFFFFF"),    # Blue/white — Molde FK
    "ROS": ("#1A1A2E", "#FFFFFF", "#FFFFFF"),    # Black/white — Rosenborg BK
    "SAN": ("#003DA5", "#FFFFFF", "#FFFFFF"),    # Blue/white — Sandefjord
    "SAR": ("#003DA5", "#FFFFFF", "#FFFFFF"),    # Blue/white — Sarpsborg 08
    "STA": ("#E31837", "#FFFFFF", "#FFFFFF"),    # Red/white — IK Start
    "TRO": ("#E31837", "#003DA5", "#FFFFFF"),    # Red/blue — Tromsø IL
    "VIK": ("#003DA5", "#FFD700", "#FFFFFF"),    # Blue/gold — Viking FK
    "VIF": ("#003DA5", "#E31837", "#FFFFFF"),    # Blue/red — Vålerenga IF
    "AAL": ("#FF6600", "#000000", "#FFFFFF"),    # Orange/black — Aalesund FK
}

# Select active colour map based on league
_colour_maps = {"FPL": TEAM_COLOURS_FPL, "ASV": TEAM_COLOURS_ASV, "ELS": TEAM_COLOURS_ELS}
TEAM_COLOURS = _colour_maps.get(LC["short_name"], TEAM_COLOURS_FPL)

# GK kit colours (separate — keepers wear different kits)
GK_COLOURS = {
    "default": ("#2ECC71", "#1A1A2E", "#FFFFFF"),  # Green, dark
}


def make_shirt_svg(team_short, xpts_text, is_gk=False, is_captain=False, width=56, height=56, player_name="", team_code=0):
    """Render the player's kit with an xPts overlay.

    FPL: pulls the official kit PNG from the Premier League CDN (same image FPL itself uses).
    Other leagues / missing code: falls back to a jersey-shaped SVG drawn from TEAM_COLOURS.
    """
    short_name = ""
    if player_name:
        parts = player_name.strip().split()
        short_name = parts[-1][:8] if parts else ""

    cap_html = ""
    if is_captain:
        cap_html = (
            f'<span style="position:absolute;top:-3px;right:-3px;width:15px;height:15px;border-radius:50%;'
            f'background:#FFD700;border:1.5px solid #0a0e17;color:#000;font-size:9px;font-weight:800;'
            f'display:flex;align-items:center;justify-content:center;font-family:Inter,Arial,sans-serif;'
            f'line-height:1;box-shadow:0 1px 3px rgba(0,0,0,0.4);z-index:3;">C</span>'
        )

    # Official FPL kit image when we have an FPL team code
    if team_code and LC.get("short_name") == "FPL":
        suffix = "_1" if is_gk else ""
        url = f"https://fantasy.premierleague.com/dist/img/shirts/standard/shirt_{team_code}{suffix}-66.png"
        xpts_pill = (
            f'<span style="position:absolute;bottom:-1px;left:50%;transform:translateX(-50%);'
            f'background:rgba(10,14,23,0.94);color:#fff;font-size:0.66rem;font-weight:800;'
            f'padding:1px 7px;border-radius:8px;border:1px solid #2a3550;line-height:1.25;'
            f'white-space:nowrap;font-family:Inter,Arial,sans-serif;letter-spacing:-0.2px;z-index:2;">'
            f'{xpts_text}</span>'
        )
        return (
            f'<div style="position:relative;width:{width}px;height:{height}px;display:inline-block;">'
            f'<img src="{url}" width="{width}" height="{height}" alt="{team_short} kit" '
            f'loading="lazy" style="display:block;object-fit:contain;'
            f'filter:drop-shadow(0 2px 3px rgba(0,0,0,0.35));" />'
            f'{cap_html}{xpts_pill}</div>'
        )

    # Fallback: stylized jersey silhouette in team colours
    if is_gk:
        primary, secondary, text_col = GK_COLOURS.get(team_short, GK_COLOURS["default"])
    else:
        primary, secondary, text_col = TEAM_COLOURS.get(team_short, ("#666666", "#FFFFFF", "#FFFFFF"))

    w, h = width, height
    body_w = w * 0.60
    body_x = (w - body_w) / 2
    shoulder_y = h * 0.20
    sleeve_y = h * 0.50
    hem_y = h * 0.94
    neck_w = w * 0.18
    neck_x = (w - neck_w) / 2
    sl_out = w * 0.14
    cap_svg = ""
    if is_captain:
        cap_svg = (
            f'<circle cx="{w-7}" cy="7" r="6.5" fill="#FFD700" stroke="#0a0e17" stroke-width="1"/>'
            f'<text x="{w-7}" y="10.5" text-anchor="middle" font-size="8" font-weight="bold" '
            f'fill="#000" font-family="Arial,sans-serif">C</text>'
        )
    jersey_path = (
        f'M {body_x} {shoulder_y} '
        f'L {neck_x} {shoulder_y} '
        f'Q {w/2} {shoulder_y + h*0.10} {neck_x + neck_w} {shoulder_y} '
        f'L {body_x + body_w} {shoulder_y} '
        f'L {body_x + body_w + sl_out} {shoulder_y + h*0.06} '
        f'L {body_x + body_w + sl_out*0.55} {sleeve_y} '
        f'L {body_x + body_w} {shoulder_y + h*0.30} '
        f'L {body_x + body_w} {hem_y} '
        f'L {body_x} {hem_y} '
        f'L {body_x} {shoulder_y + h*0.30} '
        f'L {body_x - sl_out*0.55} {sleeve_y} '
        f'L {body_x - sl_out} {shoulder_y + h*0.06} Z'
    )
    svg = (
        f'<svg width="{w}" height="{h}" viewBox="0 0 {w} {h}" xmlns="http://www.w3.org/2000/svg">'
        f'<path d="{jersey_path}" fill="{primary}" stroke="{secondary}" stroke-width="1.5" stroke-linejoin="round"/>'
        f'<text x="{w/2}" y="{h*0.56}" text-anchor="middle" font-size="8" font-weight="400" '
        f'fill="{text_col}" opacity="0.7" font-family="Arial,sans-serif">{short_name}</text>'
        f'<text x="{w/2}" y="{h*0.80}" text-anchor="middle" font-size="13" font-weight="700" '
        f'fill="{text_col}" font-family="Arial,sans-serif">{xpts_text}</text>'
        f'{cap_svg}'
        f'</svg>'
    )
    return svg

# ============================================================
# CUSTOM CSS
# ============================================================
st.markdown("""
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&family=Plus+Jakarta+Sans:wght@800&family=JetBrains+Mono:wght@500;700&display=swap" rel="stylesheet">
<style>
    :root {
        --bg: #07090f;
        --bg-grad-1: #0a0e17;
        --bg-grad-2: #0d111c;
        --surface: #111827;
        --surface-2: #161f33;
        --surface-3: #1a2236;
        --border: #1f2940;
        --border-strong: #2a3550;
        --border-focus: #3a4670;
        --text: #e7ecf4;
        --text-muted: #9aa5c0;
        --text-faint: #5a6580;
        --brand: #f02d6e;
        --brand-2: #ff6b8a;
        --brand-grad: linear-gradient(135deg, #f02d6e 0%, #e8456e 50%, #ff6b8a 100%);
        --accent: #38bdf8;
        --good: #34d399;
        --warn: #fbbf24;
        --bad: #f87171;
        --shadow-sm: 0 1px 2px rgba(0,0,0,0.25);
        --shadow-md: 0 4px 12px rgba(0,0,0,0.35);
        --shadow-glow: 0 0 0 1px rgba(240,45,110,0.18), 0 8px 24px rgba(240,45,110,0.12);
    }

    /* ---- Base ---- */
    html, body, [class*="st-"], [class*="css-"], button, input, textarea, select {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif !important;
        -webkit-font-smoothing: antialiased;
        -moz-osx-font-smoothing: grayscale;
    }
    .stApp {
        background:
            radial-gradient(1200px 600px at 12% -10%, rgba(240,45,110,0.07), transparent 60%),
            radial-gradient(900px 500px at 95% 0%, rgba(56,189,248,0.05), transparent 60%),
            linear-gradient(180deg, var(--bg-grad-1) 0%, var(--bg) 100%);
        color: var(--text);
    }
    header[data-testid="stHeader"] {
        background: rgba(7,9,15,0.7);
        backdrop-filter: saturate(160%) blur(14px);
        border-bottom: 1px solid var(--border);
    }
    /* Tabular figures everywhere — stat-heavy app */
    table, code, kbd, .stDataFrame, .metric-value, .metric-sub, .gw-num,
    .pitch-price, .pitch-shirt-xpts, .badge, .fdr-1, .fdr-2, .fdr-3, .fdr-4, .fdr-5 {
        font-variant-numeric: tabular-nums;
        font-feature-settings: "tnum" 1, "ss01" 1;
    }
    /* Custom scrollbar */
    ::-webkit-scrollbar { width: 10px; height: 10px; }
    ::-webkit-scrollbar-track { background: transparent; }
    ::-webkit-scrollbar-thumb { background: #1c2540; border-radius: 8px; border: 2px solid var(--bg); }
    ::-webkit-scrollbar-thumb:hover { background: #2a3658; }
    /* Selection */
    ::selection { background: rgba(240,45,110,0.35); color: #fff; }

    /* ---- Product header ---- */
    .app-header {
        display: flex; align-items: baseline; gap: 0.75rem; flex-wrap: wrap;
        padding: 0.4rem 0 0.9rem 0;
        border-bottom: 1px solid var(--border);
        margin-bottom: 1.1rem;
    }
    .main-title {
        background: var(--brand-grad);
        -webkit-background-clip: text; background-clip: text;
        -webkit-text-fill-color: transparent; color: transparent;
        font-size: 1.85rem; font-weight: 800;
        letter-spacing: -0.03em; line-height: 1; margin: 0;
    }
    .sub-title {
        color: var(--text-muted); font-size: 0.82rem; margin: 0;
        font-weight: 500; letter-spacing: 0.01em;
    }
    .app-tag {
        display: inline-flex; align-items: center; gap: 5px;
        font-size: 0.62rem; font-weight: 700; letter-spacing: 0.12em;
        text-transform: uppercase; color: var(--brand-2);
        background: rgba(240,45,110,0.10);
        border: 1px solid rgba(240,45,110,0.28);
        padding: 3px 8px; border-radius: 999px; line-height: 1;
    }
    .app-tag::before {
        content: ""; width: 5px; height: 5px; border-radius: 50%;
        background: var(--brand-2); box-shadow: 0 0 6px var(--brand);
    }

    /* ---- Metric cards ---- */
    .metric-card {
        background: linear-gradient(180deg, var(--surface) 0%, #0e1522 100%);
        border: 1px solid var(--border-strong);
        border-radius: 14px; padding: 1.05rem 1rem;
        text-align: center; position: relative; overflow: hidden;
        transition: transform 140ms ease, border-color 140ms ease, box-shadow 140ms ease;
    }
    .metric-card::before {
        content:""; position:absolute; left:0; right:0; top:0; height:1px;
        background: linear-gradient(90deg, transparent, rgba(240,45,110,0.45), transparent);
        opacity: 0; transition: opacity 160ms ease;
    }
    .metric-card:hover { transform: translateY(-1px); border-color: var(--border-focus); box-shadow: var(--shadow-md); }
    .metric-card:hover::before { opacity: 1; }
    .metric-label {
        color: var(--text-faint); font-size: 0.66rem; font-weight: 600;
        text-transform: uppercase; letter-spacing: 0.12em; margin-bottom: 4px;
    }
    .metric-value {
        color: var(--text); font-size: 1.7rem; font-weight: 700;
        letter-spacing: -0.02em; line-height: 1.1;
    }
    .metric-sub { color: var(--text-muted); font-size: 0.74rem; margin-top: 3px; font-weight: 500; }

    /* ---- FDR chips ---- */
    .fdr-1, .fdr-2, .fdr-3, .fdr-4, .fdr-5 {
        padding: 2px 8px; border-radius: 6px;
        font-size: 0.72rem; font-weight: 700;
        display: inline-block; margin: 1px;
        border: 1px solid transparent;
    }
    .fdr-1 { background: rgba(6,95,70,0.55);  color: #6ee7b7; border-color: rgba(110,231,183,0.18); }
    .fdr-2 { background: rgba(20,83,45,0.55); color: #86efac; border-color: rgba(134,239,172,0.16); }
    .fdr-3 { background: rgba(120,53,15,0.55); color: #fcd34d; border-color: rgba(252,211,77,0.16); }
    .fdr-4 { background: rgba(124,45,18,0.6); color: #fdba74; border-color: rgba(253,186,116,0.18); }
    .fdr-5 { background: rgba(127,29,29,0.6); color: #fca5a5; border-color: rgba(252,165,165,0.18); }

    /* ---- Cards & panels ---- */
    .transfer-card {
        background: linear-gradient(180deg, var(--surface-3) 0%, #141b2c 100%);
        border: 1px solid var(--border-strong);
        border-radius: 12px; padding: 0.8rem 1rem;
        margin-bottom: 0.45rem;
        transition: border-color 140ms ease, transform 140ms ease;
    }
    .transfer-card:hover { border-color: var(--border-focus); transform: translateY(-1px); }
    .transfer-out { color: var(--bad); font-weight: 600; }
    .transfer-in { color: var(--good); font-weight: 600; }
    .transfer-arrow { color: var(--accent); font-size: 1.1rem; }

    /* ---- GW bar ---- */
    .gw-bar {
        background: linear-gradient(180deg, var(--surface) 0%, #0e1522 100%);
        border: 1px solid var(--border-strong);
        border-radius: 14px;
        padding: 0.7rem 1.15rem;
        display: flex; align-items: center; gap: 1rem;
        margin-bottom: 1.1rem; flex-wrap: wrap;
        box-shadow: var(--shadow-sm);
    }
    .gw-num {
        background: var(--brand-grad);
        -webkit-background-clip: text; background-clip: text;
        -webkit-text-fill-color: transparent; color: transparent;
        font-size: 1.05rem; font-weight: 800; letter-spacing: -0.01em;
    }
    .gw-deadline { color: var(--text-muted); font-size: 0.78rem; font-weight: 500; }

    /* ---- Badges & source tags ---- */
    .badge {
        font-size: 0.66rem; padding: 3px 9px; border-radius: 999px;
        font-weight: 700; letter-spacing: 0.02em;
        border: 1px solid transparent;
    }
    .badge-green  { background: rgba(52,211,153,0.12); color: #34d399; border-color: rgba(52,211,153,0.25); }
    .badge-yellow { background: rgba(251,191,36,0.12); color: #fbbf24; border-color: rgba(251,191,36,0.25); }
    .badge-blue   { background: rgba(56,189,248,0.12); color: #38bdf8; border-color: rgba(56,189,248,0.25); }

    .source-tag {
        display: inline-block; font-size: 0.6rem; padding: 2px 7px; border-radius: 4px;
        font-weight: 700; letter-spacing: 0.04em; margin-left: 6px;
        text-transform: uppercase;
    }
    .src-fpl   { background: rgba(56,189,248,0.13); color: #38bdf8; }
    .src-odds  { background: rgba(251,191,36,0.13); color: #fbbf24; }
    .src-model { background: rgba(167,139,250,0.14); color: #a78bfa; }

    /* ---- Pitch ---- */
    .pitch-row-label {
        color: var(--text-faint); font-size: 0.66rem; font-weight: 600;
        text-transform: uppercase; letter-spacing: 0.14em;
        margin: 14px 0 6px; padding-bottom: 4px;
        border-bottom: 1px dashed var(--border);
    }
    .pitch-shirt-container { display: inline-flex; flex-direction: column; align-items: center; position: relative; margin: 0 auto; }
    .pitch-shirt-xpts { position: absolute; top: 18px; left: 50%; transform: translateX(-50%); font-weight: 700; font-size: 0.72rem; color: #fff; text-shadow: 0 1px 2px rgba(0,0,0,0.5); z-index: 2; }
    .pitch-name { font-size: 0.72rem; font-weight: 600; color: var(--text); margin-top: 5px; line-height: 1.15; }
    .pitch-price { font-size: 0.6rem; color: var(--text-faint); font-weight: 500; }

    /* ---- Section headers ---- */
    .section-header {
        font-size: 1.02rem; font-weight: 700; color: var(--text);
        letter-spacing: -0.01em; line-height: 1.3;
        margin: 1.4rem 0 0.7rem;
        padding-left: 10px;
        border-left: 3px solid var(--brand);
    }

    /* ---- Streamlit overrides ---- */
    /* Tabs */
    .stTabs [data-baseweb="tab-list"] {
        gap: 2px; border-bottom: 1px solid var(--border);
        background: transparent;
    }
    .stTabs [data-baseweb="tab"] {
        background: transparent !important;
        color: var(--text-muted) !important;
        font-weight: 600 !important;
        font-size: 0.86rem !important;
        padding: 10px 14px !important;
        border-radius: 8px 8px 0 0 !important;
        border-bottom: 2px solid transparent !important;
        transition: color 140ms ease, background 140ms ease, border-color 140ms ease;
    }
    .stTabs [data-baseweb="tab"]:hover {
        color: var(--text) !important;
        background: rgba(255,255,255,0.02) !important;
    }
    .stTabs [aria-selected="true"] {
        color: var(--text) !important;
        border-bottom: 2px solid var(--brand) !important;
        background: rgba(240,45,110,0.05) !important;
    }
    .stTabs [data-baseweb="tab-highlight"] { display: none !important; }

    /* Buttons */
    .stButton > button, .stDownloadButton > button {
        font-weight: 600 !important;
        border-radius: 10px !important;
        border: 1px solid var(--border-strong) !important;
        background: var(--surface-2) !important;
        color: var(--text) !important;
        transition: transform 100ms ease, border-color 140ms ease, background 140ms ease, box-shadow 140ms ease;
    }
    .stButton > button:hover, .stDownloadButton > button:hover {
        border-color: var(--border-focus) !important;
        background: #1d2740 !important;
        transform: translateY(-1px);
    }
    .stButton > button:active { transform: translateY(0); }
    .stButton > button[kind="primary"], .stButton > button[data-testid*="primary"] {
        background: var(--brand-grad) !important;
        border: 1px solid transparent !important;
        color: #fff !important;
        box-shadow: var(--shadow-glow);
    }
    .stButton > button[kind="primary"]:hover {
        filter: brightness(1.06);
        box-shadow: 0 0 0 1px rgba(240,45,110,0.35), 0 10px 28px rgba(240,45,110,0.22) !important;
    }

    /* Inputs */
    .stTextInput input, .stNumberInput input, .stSelectbox > div[data-baseweb="select"] > div, .stMultiSelect > div[data-baseweb="select"] > div {
        background: var(--surface-2) !important;
        border-color: var(--border-strong) !important;
        color: var(--text) !important;
        border-radius: 10px !important;
    }
    .stTextInput input:focus, .stNumberInput input:focus {
        border-color: var(--brand) !important;
        box-shadow: 0 0 0 3px rgba(240,45,110,0.15) !important;
    }

    /* Metrics (st.metric) */
    [data-testid="stMetric"] {
        background: linear-gradient(180deg, var(--surface) 0%, #0e1522 100%);
        border: 1px solid var(--border-strong);
        border-radius: 14px;
        padding: 0.85rem 1rem;
    }
    [data-testid="stMetricLabel"] { color: var(--text-faint) !important; font-weight: 600; letter-spacing: 0.08em; text-transform: uppercase; font-size: 0.66rem !important; }
    [data-testid="stMetricValue"] { color: var(--text) !important; font-weight: 700 !important; letter-spacing: -0.02em; }

    /* Expanders */
    .streamlit-expanderHeader, [data-testid="stExpander"] summary {
        background: var(--surface) !important;
        border: 1px solid var(--border-strong) !important;
        border-radius: 10px !important;
        font-weight: 600 !important;
    }

    /* Dataframes */
    .stDataFrame, [data-testid="stDataFrame"] {
        border: 1px solid var(--border-strong);
        border-radius: 10px;
        overflow: hidden;
    }

    /* Alerts / info / warning / success */
    [data-testid="stAlert"] {
        border-radius: 10px;
        border-width: 1px;
        border-style: solid;
    }

    /* Sidebar */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0a0e17 0%, #07090f 100%);
        border-right: 1px solid var(--border);
    }

    /* ---- Landing tiles ("What's new for you") ---- */
    .landing-grid {
        display: grid; gap: 0.85rem;
        grid-template-columns: repeat(3, minmax(0, 1fr));
        margin: 0.4rem 0 1.25rem;
    }
    @media (max-width: 900px) {
        .landing-grid { grid-template-columns: 1fr; }
    }
    .landing-tile {
        background: linear-gradient(180deg, var(--surface) 0%, #0e1522 100%);
        border: 1px solid var(--border-strong);
        border-radius: 14px;
        padding: 0.95rem 1.1rem;
        position: relative; overflow: hidden;
        transition: transform 140ms ease, border-color 140ms ease, box-shadow 140ms ease;
    }
    .landing-tile::before {
        content: ""; position: absolute; left: 0; right: 0; top: 0; height: 1px;
        background: linear-gradient(90deg, transparent, rgba(240,45,110,0.5), transparent);
        opacity: 0; transition: opacity 160ms ease;
    }
    .landing-tile:hover { transform: translateY(-1px); border-color: var(--border-focus); box-shadow: var(--shadow-md); }
    .landing-tile:hover::before { opacity: 1; }
    .lt-icon {
        position: absolute; top: 12px; right: 14px;
        font-size: 1.05rem; opacity: 0.55; line-height: 1;
    }
    .lt-label {
        color: var(--text-faint); font-size: 0.66rem; font-weight: 700;
        text-transform: uppercase; letter-spacing: 0.14em;
        margin-bottom: 6px;
    }
    .lt-value {
        color: var(--text); font-size: 1.55rem; font-weight: 800;
        letter-spacing: -0.02em; line-height: 1.15;
        font-variant-numeric: tabular-nums; font-feature-settings: "tnum" 1;
    }
    .lt-welcome-value { background: var(--brand-grad); -webkit-background-clip: text; background-clip: text; -webkit-text-fill-color: transparent; color: transparent; }
    .lt-captain-name { font-size: 1.4rem; }
    .lt-sub { color: var(--text-muted); font-size: 0.76rem; margin-top: 4px; font-weight: 500; }
    .lt-countdown.lt-urgent { color: var(--brand-2); }
    .lt-countdown.lt-live { color: var(--good); }
    .lt-cta { border-style: dashed; border-color: rgba(240,45,110,0.35); background: linear-gradient(180deg, rgba(240,45,110,0.06) 0%, rgba(240,45,110,0.01) 100%); }

    /* Hide chrome */
    #MainMenu { visibility: hidden; }
    footer { visibility: hidden; }
    [data-testid="stDecoration"] { display: none; }
</style>
""", unsafe_allow_html=True)


# ============================================================
# DATA LAYER
# ============================================================

@st.cache_data(ttl=3600)
def load_fpl_data():
    """Source 1: FPL API — player stats, prices, fixtures, xG/xA."""
    try:
        headers = {"User-Agent": "FPL-Optimizer/2.0"}
        b = requests.get(f"{FPL_BASE}/bootstrap-static/", headers=headers, timeout=30).json()
        f = requests.get(f"{FPL_BASE}/fixtures/", headers=headers, timeout=30).json()
        return b, f, None
    except Exception as e:
        return None, None, str(e)


@st.cache_data(ttl=7200)
def load_betting_odds():
    """Source 2: football-data.co.uk — match betting odds."""
    if FOOTBALL_DATA_URL is None:
        return None, "No odds URL configured for this league"
    try:
        resp = requests.get(FOOTBALL_DATA_URL, timeout=20)
        if resp.status_code != 200:
            return None, "HTTP " + str(resp.status_code)
        df = pd.read_csv(StringIO(resp.text), on_bad_lines="skip")
        if len(df) == 0:
            return None, "Empty CSV"
        return df, None
    except Exception as e:
        return None, str(e)



@st.cache_data(ttl=7200)
def load_club_elo():
    """Source 3: Club Elo ratings — dynamic team strength from clubelo.com."""
    try:
        today = datetime.now().strftime("%Y-%m-%d")
        resp = requests.get(f"http://api.clubelo.com/{today}", timeout=15)
        if resp.status_code != 200:
            return None, f"HTTP {resp.status_code}"
        df = pd.read_csv(StringIO(resp.text), sep=",")
        if len(df) == 0:
            return None, "Empty response"
        country_code = {"FPL": "ENG", "ASV": "SWE", "ELS": "NOR"}.get(LC["short_name"], "ENG")
        league_data = df[(df["Country"] == country_code) & (df["Level"] == 1)].copy()
        elo_map = {}
        for _, row in league_data.iterrows():
            elo_map[row["Club"]] = float(row["Elo"])
        return elo_map, None
    except Exception as e:
        return None, str(e)


# Elo and Odds API team maps — league-specific
ELO_NAME_MAP_FPL = {
    "Arsenal": "ARS", "Aston Villa": "AVL", "Bournemouth": "BOU",
    "Brentford": "BRE", "Brighton": "BHA", "Chelsea": "CHE",
    "Crystal Palace": "CRY", "Everton": "EVE", "Fulham": "FUL",
    "Ipswich": "IPS", "Leicester": "LEI", "Liverpool": "LIV",
    "Man City": "MCI", "Man United": "MUN", "Newcastle": "NEW",
    "Nottingham Forest": "NFO", "Southampton": "SOU", "Tottenham": "TOT",
    "West Ham": "WHU", "Wolves": "WOL",
    "Leeds": "LEE", "Burnley": "BUR", "Sunderland": "SUN",
    "Sheffield United": "SHU", "Norwich": "NOR",
    "Middlesbrough": "MID", "Luton": "LUT",
}

ELO_NAME_MAP_ASV = {
    "AIK": "AIK", "Hacken": "HÄC", "Degerfors": "DEG",
    "Djurgarden": "DJU", "GAIS": "GAI", "Hammarby": "HAM",
    "Elfsborg": "ELF", "Goteborg": "IFG", "Kalmar": "KAL",
    "Malmo": "MFF", "Mjallby": "MJÄ", "Sirius": "SIR",
    "Vasteras": "VSK", "Orgryte": "ÖRG",
    "Brommapojkarna": "BPK", "Halmstad": "HBK",
}

ELO_NAME_MAP_ELS = {
    "Bodo/Glimt": "BOD", "Brann": "BRA", "Fredrikstad": "FRE",
    "HamKam": "HAM", "KFUM Oslo": "KFU", "Kristiansund": "KRI",
    "Lillestrom": "LIL", "Molde": "MOL", "Rosenborg": "ROS",
    "Sandefjord": "SAN", "Sarpsborg 08": "SAR", "Start": "STA",
    "Tromso": "TRO", "Viking": "VIK", "Valerenga": "VIF",
    "Aalesund": "AAL",
}

_elo_maps = {"FPL": ELO_NAME_MAP_FPL, "ASV": ELO_NAME_MAP_ASV, "ELS": ELO_NAME_MAP_ELS}
ELO_NAME_MAP = _elo_maps.get(LC["short_name"], ELO_NAME_MAP_FPL)

ODDS_API_KEY = "e6df27ee56e4f85f1b20b194e4ffd080"

ODDS_API_TEAM_MAP_FPL = {
    "Arsenal": "ARS", "Aston Villa": "AVL", "AFC Bournemouth": "BOU",
    "Brentford": "BRE", "Brighton and Hove Albion": "BHA", "Chelsea": "CHE",
    "Crystal Palace": "CRY", "Everton": "EVE", "Fulham": "FUL",
    "Ipswich Town": "IPS", "Leicester City": "LEI", "Liverpool": "LIV",
    "Manchester City": "MCI", "Manchester United": "MUN", "Newcastle United": "NEW",
    "Nottingham Forest": "NFO", "Southampton": "SOU", "Tottenham Hotspur": "TOT",
    "West Ham United": "WHU", "Wolverhampton Wanderers": "WOL",
    "Leeds United": "LEE", "Burnley": "BUR", "Sunderland": "SUN",
    "Sheffield United": "SHU",
}

ODDS_API_TEAM_MAP_ASV = {
    "AIK": "AIK", "BK Hacken": "HÄC", "BK Häcken": "HÄC",
    "Degerfors IF": "DEG", "Djurgardens IF": "DJU", "Djurgårdens IF": "DJU",
    "GAIS": "GAI", "Hammarby": "HAM", "Hammarby IF": "HAM",
    "IF Elfsborg": "ELF", "Elfsborg": "ELF",
    "IFK Goteborg": "IFG", "IFK Göteborg": "IFG",
    "Kalmar FF": "KAL", "Malmö FF": "MFF", "Malmo FF": "MFF",
    "Mjallby AIF": "MJÄ", "Mjällby AIF": "MJÄ",
    "IK Sirius": "SIR", "Sirius": "SIR",
    "Vasteras SK": "VSK", "Västerås SK": "VSK",
    "Orgryte IS": "ÖRG", "Örgryte IS": "ÖRG",
    "IF Brommapojkarna": "BPK", "Brommapojkarna": "BPK",
    "Halmstads BK": "HBK", "Halmstad": "HBK",
}

ODDS_API_TEAM_MAP_ELS = {
    "Bodø/Glimt": "BOD", "Bodo/Glimt": "BOD", "FK Bodø/Glimt": "BOD",
    "SK Brann": "BRA", "Brann": "BRA",
    "Fredrikstad FK": "FRE", "Fredrikstad": "FRE",
    "HamKam": "HAM", "Hamarkameratene": "HAM",
    "KFUM Oslo": "KFU", "KFUM": "KFU",
    "Kristiansund BK": "KRI", "Kristiansund": "KRI",
    "Lillestrøm SK": "LIL", "Lillestrom": "LIL", "Lillestrøm": "LIL",
    "Molde FK": "MOL", "Molde": "MOL",
    "Rosenborg BK": "ROS", "Rosenborg": "ROS",
    "Sandefjord Fotball": "SAN", "Sandefjord": "SAN",
    "Sarpsborg 08": "SAR", "Sarpsborg 08 FF": "SAR",
    "IK Start": "STA", "Start": "STA",
    "Tromsø IL": "TRO", "Tromso": "TRO", "Tromsø": "TRO",
    "Viking FK": "VIK", "Viking": "VIK",
    "Vålerenga IF": "VIF", "Valerenga": "VIF", "Vålerenga": "VIF",
    "Aalesunds FK": "AAL", "Aalesund": "AAL",
}

_odds_maps = {"FPL": ODDS_API_TEAM_MAP_FPL, "ASV": ODDS_API_TEAM_MAP_ASV, "ELS": ODDS_API_TEAM_MAP_ELS}
ODDS_API_TEAM_MAP = _odds_maps.get(LC["short_name"], ODDS_API_TEAM_MAP_FPL)


@st.cache_data(ttl=21600)
def load_live_odds():
    """Source 4: The Odds API — live fixture-specific match odds."""
    if not ODDS_API_KEY:
        return None, "No API key"
    try:
        odds_league = LC.get("odds_api_league", "soccer_epl")
        url = f"https://api.the-odds-api.com/v4/sports/{odds_league}/odds?apiKey={ODDS_API_KEY}&regions=uk&markets=h2h,totals&oddsFormat=decimal"
        resp = requests.get(url, timeout=20)
        if resp.status_code == 401: return None, "Invalid API key"
        if resp.status_code == 429: return None, "Rate limit exceeded"
        if resp.status_code != 200: return None, f"HTTP {resp.status_code}"
        remaining = resp.headers.get("x-requests-remaining", "?")
        data = resp.json()
        if not data: return None, "No upcoming fixtures"
        fixture_odds = {}
        for event in data:
            home_team, away_team = event.get("home_team", ""), event.get("away_team", "")
            home_fpl, away_fpl = ODDS_API_TEAM_MAP.get(home_team), ODDS_API_TEAM_MAP.get(away_team)
            if not home_fpl or not away_fpl: continue
            h_list, d_list, a_list = [], [], []
            over_lines = []  # over/under 2.5 goals implied probabilities

            for bookie in event.get("bookmakers", []):
                for market in bookie.get("markets", []):
                    if market.get("key") == "h2h":
                        oc = {o["name"]: o["price"] for o in market.get("outcomes", [])}
                        if home_team in oc and away_team in oc and "Draw" in oc:
                            h_list.append(oc[home_team]); d_list.append(oc["Draw"]); a_list.append(oc[away_team])
                    elif market.get("key") == "totals":
                        # Extract over/under 2.5 goals line
                        for outcome in market.get("outcomes", []):
                            if outcome.get("point") == 2.5 and outcome.get("name") == "Over":
                                over_lines.append(outcome["price"])

            if h_list:
                ah, ad, aa = np.mean(h_list), np.mean(d_list), np.mean(a_list)
                ornd = (1/ah + 1/ad + 1/aa)
                hp, dp, ap = (1/ah)/ornd, (1/ad)/ornd, (1/aa)/ornd

                # Over/under 2.5 goals → implied total goals for the match
                # Better empirical mapping from over 2.5 probability:
                # 40% → ~2.15, 50% → ~2.45, 60% → ~2.65, 70% → ~2.85
                if over_lines:
                    avg_over_price = np.mean(over_lines)
                    over_prob = min(1.0 / avg_over_price, 0.95)
                    expected_total_goals = 1.75 + over_prob * 1.6
                    expected_total_goals = max(expected_total_goals, 1.5)
                else:
                    expected_total_goals = 2.6  # PL average ~2.6 goals per match

                # Split total goals between home and away using win probabilities
                # Higher win prob → more goals for that team
                home_goal_share = (hp * 0.6 + dp * 0.35 + ap * 0.05)
                away_goal_share = (ap * 0.6 + dp * 0.35 + hp * 0.05)
                share_sum = home_goal_share + away_goal_share
                home_expected_goals = expected_total_goals * (home_goal_share / share_sum)
                away_expected_goals = expected_total_goals * (away_goal_share / share_sum)

                # CS probability: use Poisson with bookmaker-implied goals
                # P(0 goals) = e^(-expected_goals)
                home_cs_prob_poisson = math.exp(-away_expected_goals)
                away_cs_prob_poisson = math.exp(-home_expected_goals)

                fixture_odds[(home_fpl, away_fpl)] = {
                    "home_win_prob": round(hp, 3), "draw_prob": round(dp, 3), "away_win_prob": round(ap, 3),
                    "home_cs_prob": round(home_cs_prob_poisson, 3),
                    "away_cs_prob": round(away_cs_prob_poisson, 3),
                    "expected_total_goals": round(expected_total_goals, 2),
                    "home_expected_goals": round(home_expected_goals, 2),
                    "away_expected_goals": round(away_expected_goals, 2),
                    "home_attack_str": round(home_expected_goals / 1.35, 3),  # normalise to league avg
                    "away_attack_str": round(away_expected_goals / 1.35, 3),
                    "home_defence_str": round(away_expected_goals / 1.35, 3),  # conceding = opponent scoring
                    "away_defence_str": round(home_expected_goals / 1.35, 3),
                    "n_bookmakers": len(h_list),
                    "has_totals": len(over_lines) > 0,
                }
        return {"fixtures": fixture_odds, "remaining": remaining}, None
    except Exception as e:
        return None, str(e)

@st.cache_data(ttl=3600)
def load_recent_gw_live_data(current_gw_id, n_recent=7):
    """
    Fetch per-GW live stats for the last N completed gameweeks.
    Uses event/{gw}/live/ endpoint — one call per GW, returns all players.
    Returns: dict {player_id: [{gw, minutes, xG, xA, xGC, goals, assists, ...}, ...]}
    Now also tracks 0-minute GWs for rotation analysis.
    """
    headers = {"User-Agent": "FPL-Optimizer/2.0"}
    player_gw_data = {}

    start_gw = max(1, current_gw_id - n_recent)
    gws_to_fetch = list(range(start_gw, current_gw_id))
    if not gws_to_fetch:
        return player_gw_data

    def _fetch(gw):
        try:
            resp = requests.get(
                f"{FPL_BASE}/event/{gw}/live/",
                headers=headers, timeout=15,
            )
            if resp.status_code != 200:
                return gw, None
            return gw, resp.json()
        except Exception:
            return gw, None

    # Fetch all GWs in parallel — FPL endpoints are independent so we get
    # ~Nx speedup on cold cache instead of serial round-trips.
    with ThreadPoolExecutor(max_workers=min(len(gws_to_fetch), 8)) as ex:
        results = list(ex.map(_fetch, gws_to_fetch))

    # Process in chronological order so per-player lists stay time-ordered
    for gw, data in sorted(results, key=lambda r: r[0]):
        if data is None:
            continue
        elements = data.get("elements", [])
        for el in elements:
            pid = el["id"]
            stats = el.get("stats", {})
            minutes = stats.get("minutes", 0)

            if pid not in player_gw_data:
                player_gw_data[pid] = []
            player_gw_data[pid].append({
                "gw": gw,
                "minutes": minutes,
                "started": 1 if minutes >= 60 else 0,
                "appeared": 1 if minutes > 0 else 0,
                "goals": stats.get("goals_scored", 0),
                "assists": stats.get("assists", 0),
                "xG": float(stats.get("expected_goals", 0) or 0),
                "xA": float(stats.get("expected_assists", 0) or 0),
                "xGC": float(stats.get("expected_goals_conceded", 0) or 0),
                "xGI": float(stats.get("expected_goal_involvements", 0) or 0),
                "clean_sheets": stats.get("clean_sheets", 0),
                "bonus": stats.get("bonus", 0),
                "total_points": stats.get("total_points", 0),
            })

    return player_gw_data


def compute_rotation_risk(player_gw_data, current_gw_id, n_recent=7):
    """
    Analyse recent start patterns to predict rotation risk.

    Returns: dict {player_id: {
        start_rate: 0-1 (% of recent GWs started),
        appear_rate: 0-1 (% of recent GWs appeared),
        avg_recent_mins: float,
        rotation_risk: 'low'|'medium'|'high',
        consistency: float (0-1, how consistent their minutes are),
        projected_start_prob: float (probability of starting next GW),
    }}

    Key patterns detected:
    - Nailed starters (start 90%+ of games) → low risk
    - Rotation players (start 50-80%) → medium risk (Pep roulette, etc.)
    - Bench warmers / fringe (start <50%) → high risk
    - Injury returnees (recent 0-min GWs followed by starts) → medium risk
    """
    rotation_data = {}

    for pid, gw_list in player_gw_data.items():
        # Only use recent N GWs
        recent = sorted(gw_list, key=lambda x: x["gw"], reverse=True)[:n_recent]

        if len(recent) < 3:
            continue

        n_gws = len(recent)
        n_started = sum(1 for g in recent if g["started"])
        n_appeared = sum(1 for g in recent if g["appeared"])
        mins_list = [g["minutes"] for g in recent]
        avg_mins = np.mean(mins_list)

        # Per-appearance minutes: only count GWs they actually played in.
        # This is the right input for `expected_mins = avg_when_played * P(start)`,
        # because avg_mins (above) already encodes start rate via zero-min benched
        # GWs and would double-count if multiplied by P(start) again.
        appeared_mins = [g["minutes"] for g in recent if g["appeared"]]
        avg_mins_when_played = float(np.mean(appeared_mins)) if appeared_mins else 0.0

        start_rate = n_started / n_gws
        appear_rate = n_appeared / n_gws

        # Consistency: how stable are their minutes? (low std = consistent)
        mins_std = np.std(mins_list)
        # Normalise: 0 std = perfectly consistent (1.0), high std = inconsistent (0.0)
        consistency = max(0, 1.0 - (mins_std / 45.0))  # 45 mins std = 0 consistency

        # Detect recent trend: are they being phased in/out?
        if len(recent) >= 4:
            recent_half = recent[:len(recent)//2]
            older_half = recent[len(recent)//2:]
            recent_start_rate = sum(1 for g in recent_half if g["started"]) / len(recent_half)
            older_start_rate = sum(1 for g in older_half if g["started"]) / len(older_half)
            trend = recent_start_rate - older_start_rate  # positive = trending up
        else:
            trend = 0

        # Projected start probability
        # Weight recent starts more heavily (exponential decay)
        weighted_starts = 0
        weight_sum = 0
        for i, g in enumerate(recent):
            w = 0.85 ** i  # most recent = 1.0, then 0.85, 0.72...
            weighted_starts += g["started"] * w
            weight_sum += w

        weighted_start_rate = weighted_starts / weight_sum if weight_sum > 0 else start_rate

        # Blend weighted rate with overall rate
        projected_start_prob = weighted_start_rate * 0.7 + start_rate * 0.3

        # Classify risk
        if projected_start_prob >= 0.85:
            risk = "low"
        elif projected_start_prob >= 0.55:
            risk = "medium"
        else:
            risk = "high"

        rotation_data[pid] = {
            "start_rate": round(start_rate, 2),
            "appear_rate": round(appear_rate, 2),
            "avg_recent_mins": round(avg_mins, 1),
            "avg_mins_when_played": round(avg_mins_when_played, 1),
            "consistency": round(consistency, 2),
            "trend": round(trend, 2),
            "rotation_risk": risk,
            "projected_start_prob": round(projected_start_prob, 2),
            "n_recent_gws": n_gws,
        }

    return rotation_data


def compute_form_weighted_xg(player_gw_data, n_recent=7):
    """
    Compute form-weighted xG/90 and xA/90 from recent GW data.
    Uses exponential decay: most recent GW has highest weight.
    Returns: dict {player_id: {xg_form_per90, xa_form_per90, xgc_form_per90, form_minutes}}
    """
    result = {}
    decay_factor = 0.85  # each older GW is worth 85% of the next

    for pid, gw_list in player_gw_data.items():
        # Sort by GW descending (most recent first)
        sorted_gws = sorted(gw_list, key=lambda x: x["gw"], reverse=True)[:n_recent]

        if not sorted_gws:
            continue

        weighted_xg = 0
        weighted_xa = 0
        weighted_xgc = 0
        weighted_mins = 0
        total_weight = 0

        for i, gw_data in enumerate(sorted_gws):
            weight = decay_factor ** i  # most recent = 1.0, then 0.85, 0.72, 0.61...
            mins = gw_data["minutes"]
            if mins > 0:
                weighted_xg += gw_data["xG"] * weight
                weighted_xa += gw_data["xA"] * weight
                weighted_xgc += gw_data["xGC"] * weight
                weighted_mins += mins * weight
                total_weight += weight

        if weighted_mins > 0 and total_weight > 0:
            nineties = weighted_mins / 90.0
            result[pid] = {
                "xg_form_per90": weighted_xg / nineties,
                "xa_form_per90": weighted_xa / nineties,
                "xgc_form_per90": weighted_xgc / nineties,
                "form_minutes": weighted_mins / total_weight,  # avg weighted mins
                "form_gws": len(sorted_gws),
            }

    return result


def detect_blank_double_gws(fixtures, planning_gw_id, n_gws=6, teams=None):
    """
    Detect blank and double gameweeks from fixture data.
    Returns: dict {team_id: {gw: fixture_count}} where 0 = blank, 2+ = double
    """
    if teams is None:
        teams = {}

    team_fixture_counts = {}
    for t_id in teams:
        team_fixture_counts[t_id] = {}
        for gw in range(planning_gw_id, planning_gw_id + n_gws):
            team_fixture_counts[t_id][gw] = 0

    for f in fixtures:
        ev = f.get("event")
        if ev and planning_gw_id <= ev < planning_gw_id + n_gws:
            if f["team_h"] in team_fixture_counts:
                team_fixture_counts[f["team_h"]][ev] = team_fixture_counts[f["team_h"]].get(ev, 0) + 1
            if f["team_a"] in team_fixture_counts:
                team_fixture_counts[f["team_a"]][ev] = team_fixture_counts[f["team_a"]].get(ev, 0) + 1

    return team_fixture_counts


def odds_to_probabilities(odds_df, teams_map):
    """
    Convert betting odds to match probabilities per team.
    Returns dict: team_name -> {attack_strength, defence_strength, cs_prob, goal_expectation}
    """
    if odds_df is None or len(odds_df) == 0:
        return {}

    # Standardise column names
    cols = odds_df.columns.tolist()
    # We need: HomeTeam, AwayTeam, B365H, B365D, B365A, FTHG, FTAG
    required = ["HomeTeam", "AwayTeam"]
    if not all(c in cols for c in required):
        return {}

    # Use B365 odds (most common), fallback to average odds
    h_col = "B365H" if "B365H" in cols else ("AvgH" if "AvgH" in cols else None)
    d_col = "B365D" if "B365D" in cols else ("AvgD" if "AvgD" in cols else None)
    a_col = "B365A" if "B365A" in cols else ("AvgA" if "AvgA" in cols else None)

    if not all([h_col, d_col, a_col]):
        return {}

    team_stats = {}

    # Process each team's matches
    all_teams = set(odds_df["HomeTeam"].unique()) | set(odds_df["AwayTeam"].unique())

    for team in all_teams:
        home_matches = odds_df[odds_df["HomeTeam"] == team].copy()
        away_matches = odds_df[odds_df["AwayTeam"] == team].copy()

        # Calculate implied probabilities from odds (with overround removal)
        win_probs, cs_probs, goals_for, goals_against = [], [], [], []

        for _, m in home_matches.iterrows():
            try:
                h, d, a = float(m[h_col]), float(m[d_col]), float(m[a_col])
                overround = (1/h + 1/d + 1/a)
                win_probs.append((1/h) / overround)
                # Clean sheet proxy: P(win)*0.35 + P(draw)*0.55 (approx CS given result)
                p_win = (1/h) / overround
                p_draw = (1/d) / overround
                cs_probs.append(p_win * 0.35 + p_draw * 0.55)
            except (ValueError, ZeroDivisionError):
                continue
            # Actual goals if available
            if "FTHG" in m and pd.notna(m.get("FTHG")):
                try:
                    goals_for.append(float(m["FTHG"]))
                    goals_against.append(float(m["FTAG"]))
                except (ValueError, TypeError):
                    pass

        for _, m in away_matches.iterrows():
            try:
                h, d, a = float(m[h_col]), float(m[d_col]), float(m[a_col])
                overround = (1/h + 1/d + 1/a)
                win_probs.append((1/a) / overround)
                p_win = (1/a) / overround
                p_draw = (1/d) / overround
                cs_probs.append(p_win * 0.30 + p_draw * 0.55)
            except (ValueError, ZeroDivisionError):
                continue
            if "FTAG" in m and pd.notna(m.get("FTAG")):
                try:
                    goals_for.append(float(m["FTAG"]))
                    goals_against.append(float(m["FTHG"]))
                except (ValueError, TypeError):
                    pass

        if win_probs:
            avg_gf = np.mean(goals_for) if goals_for else 1.3
            avg_ga = np.mean(goals_against) if goals_against else 1.3
            team_stats[team] = {
                "win_prob": np.mean(win_probs),
                "cs_prob": np.mean(cs_probs),
                "avg_goals_for": avg_gf,
                "avg_goals_against": avg_ga,
                "attack_strength": avg_gf / 1.3,  # relative to league average
                "defence_strength": avg_ga / 1.3,
            }

    return team_stats


# Team name mapping: football-data names → FPL short names
# For Allsvenskan, this is empty since we don't use football-data.co.uk
TEAM_NAME_MAP_FPL = {
    "Arsenal": "ARS", "Aston Villa": "AVL", "Bournemouth": "BOU",
    "Brentford": "BRE", "Brighton": "BHA", "Chelsea": "CHE",
    "Crystal Palace": "CRY", "Everton": "EVE", "Fulham": "FUL",
    "Ipswich": "IPS", "Ipswich Town": "IPS", "Leicester": "LEI", "Leicester City": "LEI",
    "Liverpool": "LIV",
    "Man City": "MCI", "Manchester City": "MCI",
    "Man United": "MUN", "Manchester Utd": "MUN", "Manchester United": "MUN",
    "Newcastle": "NEW", "Newcastle Utd": "NEW",
    "Nott'm Forest": "NFO", "Nottingham Forest": "NFO", "Nott'ham Forest": "NFO",
    "Southampton": "SOU",
    "Spurs": "TOT", "Tottenham": "TOT", "Tottenham Hotspur": "TOT",
    "West Ham": "WHU", "West Ham United": "WHU",
    "Wolves": "WOL", "Wolverhampton": "WOL",
    # 2025-26 promoted teams (adjust as needed)
    "Leeds": "LEE", "Leeds United": "LEE",
    "Burnley": "BUR", "Sunderland": "SUN",
    "Sheffield Utd": "SHU", "Sheffield United": "SHU",
    "Norwich": "NOR", "Middlesbrough": "MID",
    "Luton": "LUT", "Luton Town": "LUT",
}

TEAM_NAME_MAP_ASV = {}  # No football-data.co.uk source for Allsvenskan

TEAM_NAME_MAP = TEAM_NAME_MAP_FPL if LC["short_name"] == "FPL" else TEAM_NAME_MAP_ASV


def build_xpts_model(players_df, team_odds, teams_map, fixtures, current_gw_id,
                     form_xg_data=None, team_fixture_counts=None, elo_ratings=None,
                     live_odds=None, rotation_data=None):
    """
    Source 3: Custom expected points model.

    For each player, for the next N gameweeks, estimate xPts by combining:
      - FPL API xG/xA per 90
      - Betting odds (team attack/defence strength, CS probability)
      - Playing time probability
      - FPL scoring system

    Returns: dict player_id -> {gw: xpts, ...} for next 6 GWs
    """
    # Build FPL team short_name -> odds stats mapping
    odds_by_fpl = {}
    for odds_name, fpl_short in TEAM_NAME_MAP.items():
        if odds_name in team_odds:
            odds_by_fpl[fpl_short] = team_odds[odds_name]

    # Build FPL short_name -> Elo rating mapping
    elo_by_fpl = {}
    if elo_ratings:
        for elo_name, fpl_short in ELO_NAME_MAP.items():
            if elo_name in elo_ratings:
                elo_by_fpl[fpl_short] = elo_ratings[elo_name]

    # League average Elo (for normalisation)
    if elo_by_fpl:
        avg_elo = np.mean(list(elo_by_fpl.values()))
    else:
        avg_elo = 1600  # default PL average

    # Build opponent map per team per GW
    upcoming = {}  # team_id -> [{gw, opp_id, home, difficulty}]
    for t_id in teams_map:
        upcoming[t_id] = []
    for f in fixtures:
        ev = f.get("event")
        if ev and current_gw_id <= ev < current_gw_id + 6:
            if f["team_h"] in upcoming:
                upcoming[f["team_h"]].append({
                    "gw": ev, "opp_id": f["team_a"], "home": True,
                    "difficulty": f.get("team_h_difficulty", 3)
                })
            if f["team_a"] in upcoming:
                upcoming[f["team_a"]].append({
                    "gw": ev, "opp_id": f["team_h"], "home": False,
                    "difficulty": f.get("team_a_difficulty", 3)
                })

    # League average goals per game (from config)
    league_avg_goals = LC["league_avg_goals"]

    xpts_all = {}
    xpts_breakdown = {}  # {pid: {gw: {component: value}}}

    for _, p in players_df.iterrows():
        pid = p["id"]
        pos = p["pos_id"]
        team_short = p["team"]
        mins = p["minutes"]
        starts = max(p.get("starts", 0), 0)
        total_gws_played = max(current_gw_id - 1, 1)  # GWs elapsed so far

        # ============================================================
        # EXPECTED MINUTES MODEL (with rotation prediction)
        # ============================================================
        # Uses three signals:
        # 1. Season average minutes (baseline)
        # 2. FPL's chance_of_playing (injury/availability)
        # 3. Rotation risk from recent start patterns (Pep roulette etc.)

        # Average minutes per GW this season
        avg_mins_per_gw = mins / total_gws_played

        # Get rotation data if available
        rot = (rotation_data or {}).get(pid)

        # FPL's chance_of_playing (None = no news = likely available)
        chance = p.get("chance_playing", None)
        if chance is not None and not pd.isna(chance):
            availability = float(chance) / 100.0
        elif rot:
            # Use rotation model — projected start probability from recent pattern
            # This is the key improvement: a player who starts 60% of games
            # (e.g. Pep roulette victim) gets 0.60 availability, not 0.95
            availability = rot["projected_start_prob"]
        else:
            # Fallback: basic pattern from season average
            if avg_mins_per_gw >= 60:
                availability = 0.95
            elif avg_mins_per_gw >= 30:
                availability = 0.75
            elif avg_mins_per_gw >= 10:
                availability = 0.40
            elif mins > 0:
                availability = 0.15
            else:
                availability = 0.0

        # Expected minutes: E[mins] = E[mins | played] * P(plays). Three branches:
        #   1. rot exists with recent appearances → use per-appearance mins × P(start).
        #   2. rot exists but the player hasn't appeared in the last 7 GWs →
        #      they're effectively dropped. expected_mins = 0 (was the regression bug:
        #      previously fell back to avg_mins_per_gw, leaving benched players with
        #      inflated projections from early-season minutes).
        #   3. No rot data (early season / thin history) → heuristic fallback using
        #      season average × availability category.
        appeared_mins = (rot or {}).get("avg_mins_when_played", 0)
        if rot and appeared_mins > 0:
            # Blend recent per-appearance mins with season per-appearance baseline
            season_appeared = mins / max(p.get("starts", 0) or 1, 1)
            mins_per_start = appeared_mins * 0.7 + season_appeared * 0.3
            expected_mins = min(mins_per_start * availability, 90)
        elif rot:
            # Hasn't appeared in the recent window — out of the picture
            expected_mins = 0.0
        else:
            # No rotation data — heuristic fallback
            expected_mins = min(avg_mins_per_gw * availability, 90)

        # Fringe-player cliff: if the recent record shows almost no starts AND mostly
        # absent, hard-cap expected minutes regardless of season-history inertia.
        # Catches benched fringe forwards (Awoniyi-type) and backup keepers whose
        # season averages still look respectable from earlier starts.
        if rot:
            _start_rate = rot.get("start_rate", 0)
            _appear_rate = rot.get("appear_rate", 0)
            if _start_rate < 0.15 and _appear_rate < 0.5:
                expected_mins = min(expected_mins, 5.0)

        # Convert to "expected 90s" for scaling xG/xA
        expected_90s = expected_mins / 90.0

        # Playing probability (for appearance points — did they get on the pitch?)
        play_prob = min(availability, 0.98)

        # Full 60+ min probability (for clean sheet, appearance pts)
        full_game_prob = expected_mins / 90.0 if expected_mins >= 45 else expected_mins / 180.0

        # ============================================================
        # PER-90 STATS — blend season average with form-weighted recent
        # ============================================================
        mins_played = max(mins, 1)
        nineties = mins_played / 90.0

        # Season-average xG/xA from FPL API
        season_xg = float(p.get("xg_per90", 0) or 0)
        season_xa = float(p.get("xa_per90", 0) or 0)

        # Fallback: use actual goals/assists per 90 ONLY if enough minutes
        if season_xg == 0 and p["goals"] > 0 and mins >= 270:
            season_xg = p["goals"] / nineties
        if season_xa == 0 and p["assists"] > 0 and mins >= 270:
            season_xa = p["assists"] / nineties

        # Form-weighted xG/xA from recent 7 GWs (if available)
        form_data = (form_xg_data or {}).get(pid)
        if form_data and form_data.get("form_gws", 0) >= 3:
            form_xg = form_data["xg_form_per90"]
            form_xa = form_data["xa_form_per90"]
            # Blend: 55% recent form, 45% season average.
            # Recent form is more predictive but noisier — slightly tilted towards
            # form, but conservative enough that hot streaks don't dominate the
            # projection (over/underperformance regression below catches the rest).
            xg_per90 = form_xg * 0.55 + season_xg * 0.45
            xa_per90 = form_xa * 0.55 + season_xa * 0.45
        else:
            xg_per90 = season_xg
            xa_per90 = season_xa

        # Also get form-weighted xGC for GKs/DEFs
        form_xgc = None
        if form_data and form_data.get("form_gws", 0) >= 3:
            form_xgc = form_data.get("xgc_form_per90")

        # Apply regression to the mean for low-sample players
        sample_weight = min(nineties / 10.0, 1.0)
        pos_avg_xg = {1: 0.0, 2: 0.02, 3: 0.12, 4: 0.35}
        pos_avg_xa = {1: 0.01, 2: 0.05, 3: 0.10, 4: 0.12}
        xg_per90 = xg_per90 * sample_weight + pos_avg_xg.get(pos, 0.1) * (1 - sample_weight)
        xa_per90 = xa_per90 * sample_weight + pos_avg_xa.get(pos, 0.08) * (1 - sample_weight)

        # ============================================================
        # OVER/UNDERPERFORMANCE REGRESSION
        # ============================================================
        # If a player has scored significantly more/fewer goals than their xG,
        # they're likely to regress. Adjust xG towards actual performance mean.
        # E.g., player with 12 goals from 8.0 xG is overperforming — reduce projected xG
        xg_total = float(p.get("xg_total", 0) or 0)
        actual_goals = p["goals"]
        if xg_total > 0 and nineties >= 5:
            overperformance = (actual_goals - xg_total) / max(nineties, 1)  # per 90
            # Apply 30% regression towards xG (don't fully regress — some players are genuinely clinical)
            regression_factor = 0.30
            xg_per90 -= overperformance * regression_factor

        xa_total = float(p.get("xa_total", 0) or 0)
        actual_assists = p["assists"]
        if xa_total > 0 and nineties >= 5:
            xa_overperf = (actual_assists - xa_total) / max(nineties, 1)
            xa_per90 -= xa_overperf * 0.25  # assists regress less aggressively

        # Floor at 0
        xg_per90 = max(xg_per90, 0)
        xa_per90 = max(xa_per90, 0)

        # ============================================================
        # SET PIECE TAKER BONUS
        # ============================================================
        # NOTE on penalties: FPL's expected_goals_per_90 ALREADY includes
        # penalties taken this season. So a pen taker's xG/90 naturally
        # reflects their penalty duty. We do NOT add a separate pen xG boost
        # to avoid double-counting.
        #
        # However, we keep a small boost (+0.015) as a forward-looking signal
        # for penalty ORDER — the FPL API confirms who is on pens even if
        # they haven't taken many yet this season.
        #
        # Corner/FK takers get an xA boost (more delivery opportunities)
        # and direct FK takers get a small xG boost.
        pen_order = int(p.get("penalties_order", 0) or 0)
        corner_order = int(p.get("corners_order", 0) or 0)
        fk_order = int(p.get("freekicks_order", 0) or 0)

        pen_xg_boost = 0.0
        set_piece_xa_boost = 0.0

        if pen_order == 1:
            # Small forward-looking boost only (FPL xG already includes pen xG)
            pen_xg_boost = 0.015
        elif pen_order == 2:
            pen_xg_boost = 0.005

        if corner_order == 1:
            set_piece_xa_boost += 0.03
        if fk_order == 1:
            set_piece_xa_boost += 0.02
            xg_per90 += 0.01  # direct FK goal threat

        xg_per90 += pen_xg_boost
        xa_per90 += set_piece_xa_boost

        player_gw_xpts = {}
        fix_list = upcoming.get(p["team_id"], [])

        for fix in fix_list:
            gw = fix["gw"]
            opp_team = teams_map.get(fix["opp_id"], {})
            opp_short = opp_team.get("short_name", "???")

            # === LIVE FIXTURE-SPECIFIC ODDS (The Odds API) ===
            # Check if we have real bookmaker odds for THIS specific match
            live_fixture = None
            if live_odds:
                fixture_odds = live_odds.get("fixtures", {})
                if fix["home"]:
                    live_fixture = fixture_odds.get((team_short, opp_short))
                else:
                    live_fixture = fixture_odds.get((opp_short, team_short))

            if live_fixture:
                # Use fixture-specific odds — much more accurate than season averages
                if fix["home"]:
                    team_atk_str = live_fixture["home_attack_str"]
                    opp_def_str = live_fixture["away_defence_str"]
                    opp_atk_str_fix = live_fixture["away_attack_str"]
                    cs_prob_from_live = live_fixture["home_cs_prob"]
                else:
                    team_atk_str = live_fixture["away_attack_str"]
                    opp_def_str = live_fixture["home_defence_str"]
                    opp_atk_str_fix = live_fixture["home_attack_str"]
                    cs_prob_from_live = live_fixture["away_cs_prob"]
            else:
                # Fallback: season-average odds + Elo blend
                opp_odds = odds_by_fpl.get(opp_short, {})
                team_attack_odds = odds_by_fpl.get(team_short, {})

                # Get Elo ratings for both teams
                team_elo = elo_by_fpl.get(team_short, avg_elo)
                opp_elo = elo_by_fpl.get(opp_short, avg_elo)
                team_elo_str = team_elo / avg_elo if avg_elo > 0 else 1.0
                opp_elo_str = opp_elo / avg_elo if avg_elo > 0 else 1.0

                opp_def_str_odds = opp_odds.get("defence_strength", 1.0)
                team_atk_str_odds = team_attack_odds.get("attack_strength", 1.0)

                if elo_by_fpl:
                    # Elo adds a dynamic signal but shouldn't dominate
                    # 70% odds-based (fixture-specific), 30% Elo (form-adjusted)
                    opp_def_str = (opp_def_str_odds * 0.7) + ((1.0 / max(opp_elo_str, 0.5)) * 0.3)
                    team_atk_str = (team_atk_str_odds * 0.7) + (team_elo_str * 0.3)
                else:
                    opp_def_str = opp_def_str_odds
                    team_atk_str = team_atk_str_odds

                opp_atk_str_fix = None  # will use season average below
                cs_prob_from_live = None

            # Scale factor: blend opponent weakness with team strength
            # Target: ~0.80 for hard fixtures (vs top 4 away), ~1.20 for easy (vs bottom 3 home)
            # This gives a ~40% swing which matches real FPL points variance by fixture
            raw_scale = (opp_def_str * 0.55 + team_atk_str * 0.25 + 0.20)
            scale = raw_scale  # no dampening — let fixtures matter
            home_boost = 1.08 if fix["home"] else 0.96

            # Separate penalty xG from open-play xG before scaling
            # Penalties are fixture-independent (~0.76 xG regardless of opponent)
            # so they shouldn't be scaled by fixture difficulty
            pen_xg_component = 0.0
            if pen_order == 1 and nineties >= 3:
                # Estimate pen xG/90: ~1 pen per 5-8 games × 0.76 xG per pen
                # Use actual data: penalties_missed + goals from penalties ≈ total pens taken
                pens_scored = float(p.get("penalties_scored", 0) or 0)  
                pens_missed = float(p.get("penalties_missed", 0) or 0)
                total_pens = pens_scored + pens_missed
                if total_pens > 0 and nineties > 0:
                    pen_xg_component = (total_pens * 0.76) / nineties  # per 90
                else:
                    pen_xg_component = 0.10  # fallback: ~1 pen per 7.5 games
            elif pen_order == 2:
                pen_xg_component = 0.03  # backup pen taker, rarely takes

            open_play_xg = max(xg_per90 - pen_xg_component, 0)

            adj_xg = open_play_xg * scale * home_boost + pen_xg_component  # pens unscaled
            adj_xa = xa_per90 * scale * home_boost

            # Clean sheet probability
            # Use live fixture odds if available, otherwise season averages
            if opp_atk_str_fix is not None:
                opp_atk_str = opp_atk_str_fix
            else:
                opp_odds_season = odds_by_fpl.get(opp_short, {})
                opp_atk_str = opp_odds_season.get("attack_strength", 1.0)

            team_odds_season = odds_by_fpl.get(team_short, {})
            team_def_str = team_odds_season.get("defence_strength", 1.0)

            # Base CS from odds — use form-weighted xGC if available, else season
            actual_xgc_per90 = float(p.get("xgc_per90", 0) or 0)
            if form_xgc is not None and form_xgc > 0:
                # Blend form xGC with season xGC (form is more recent)
                actual_xgc_per90 = form_xgc * 0.65 + actual_xgc_per90 * 0.35
            if pos in [1, 2] and actual_xgc_per90 > 0:
                # Use actual xGC as a strong signal for defensive quality
                # Poisson approximation: P(0 goals) ≈ e^(-xGC)
                base_cs = math.exp(-actual_xgc_per90 * opp_atk_str)
            else:
                # Fallback: odds-based estimate, penalised by opponent attack
                base_cs = 0.30 * (1.0 / max(opp_atk_str, 0.5))
                # Penalise teams that concede a lot (team_def_str > 1 = concede more)
                base_cs *= (1.0 / max(team_def_str, 0.5))
                base_cs = min(base_cs, 0.50)

            # Blend with odds-derived CS if available
            if cs_prob_from_live is not None:
                # Live fixture odds CS is the most accurate — weight heavily
                cs_prob = (base_cs * 0.3 + cs_prob_from_live * 0.7)
            else:
                team_cs_from_odds = team_odds_season.get("cs_prob")
                if team_cs_from_odds and actual_xgc_per90 == 0:
                    cs_prob = (base_cs * 0.4 + team_cs_from_odds * 0.6)
                else:
                    cs_prob = base_cs

            # Hard cap — no team keeps a CS more than 50% of the time
            cs_prob = min(cs_prob, 0.50)

            # Apply recent form adjustment: if team has lost 3+ of last 5, reduce CS further
            try:
                team_form_list = p["team_form"] if "team_form" in p.index else []
                if isinstance(team_form_list, list) and len(team_form_list) >= 3:
                    recent_losses = sum(1 for r in team_form_list[:5] if r == "L")
                    if recent_losses >= 3:
                        cs_prob *= 0.6  # 40% penalty for bad recent form
                    elif recent_losses >= 2:
                        cs_prob *= 0.8  # 20% penalty
            except (KeyError, TypeError):
                pass

            # Calculate expected FPL points for this fixture
            xpts = 0.0

            # Appearance points (2 pts if 60+ mins, 1 pt if <60 mins)
            xpts += 2.0 * full_game_prob + 1.0 * max(play_prob - full_game_prob, 0)

            # Goal points (scale by expected 90s played, not just binary)
            xpts += adj_xg * expected_90s * PTS_GOAL.get(pos, 4)

            # Assist points
            xpts += adj_xa * expected_90s * PTS_ASSIST

            # Clean sheet points (GK and DEF mainly — need 60+ mins)
            xpts += cs_prob * PTS_CS.get(pos, 0) * full_game_prob

            # Bonus points estimate (proportional to involvement)
            xpts += PTS_BONUS_AVG * play_prob

            # Goals conceded penalty for GK/DEF
            # FPL rule: -1 point per 2 goals conceded (i.e. -0.5 per goal) from goal 1
            if pos in [1, 2]:
                if live_fixture:
                    # Use bookmaker-implied expected goals directly
                    if fix["home"]:
                        expected_conceded = live_fixture.get("away_expected_goals", league_avg_goals)
                    else:
                        expected_conceded = live_fixture.get("home_expected_goals", league_avg_goals)
                else:
                    expected_conceded = league_avg_goals * opp_atk_str
                    if not fix["home"]:
                        expected_conceded *= 1.05  # slight away penalty
                xpts -= expected_conceded * 0.5 * full_game_prob

            # Save points for GKs: ~1pt per 3 saves
            if pos == 1:
                expected_saves = 3.0 * opp_atk_str * 0.7
                save_points = (expected_saves / 3.0)
                xpts += save_points * full_game_prob

            # ============================================================
            # DEFENSIVE CONTRIBUTION (DefCon) POINTS
            # ============================================================
            # Only applies to leagues with DefCon (e.g. FPL 2025/26)
            defcon_xpts = 0
            defcon_per90 = float(p.get("defcon_per90", 0) or 0)
            if LC.get("has_defcon", False) and defcon_per90 > 0 and pos in [2, 3, 4] and nineties >= 3:
                # defcon_per90 from FPL API = DC points earned per 90 (0-2 scale)
                #
                # Position-specific scaling:
                # DEFs need 10 CBIT — CBs hit this regularly (30-70% of games)
                # MIDs need 12 CBIRT — only elite CDMs hit this (15-35% of games)
                # FWDs need 12 CBIRT — almost never hit this (<10% of games)
                #
                # The API's defcon_per90 already reflects actual DC points earned,
                # but we still need position-specific dampening because:
                # 1. MIDs/FWDs have a higher threshold (12 vs 10)
                # 2. Their rates are less stable/predictable
                # 3. We want to avoid over-projecting for non-defensive players

                raw_prob = defcon_per90 / 2.0  # 0-1 scale

                if pos == 2:  # DEF
                    # Conservative but fair — CBs are the primary DefCon earners
                    defcon_prob = min((raw_prob ** 0.5) * 0.6, 0.70)
                elif pos == 3:  # MID
                    # Much more conservative — only elite CDMs earn DC regularly
                    # Most MIDs (attackers, wingers, AMs) almost never hit 12 CBIRT
                    defcon_prob = min((raw_prob ** 0.5) * 0.35, 0.40)
                else:  # FWD (pos == 4)
                    # Extremely rare — almost no forwards hit 12 CBIRT
                    defcon_prob = min((raw_prob ** 0.5) * 0.15, 0.20)

                # Mild fixture adjustment
                defcon_prob *= (0.9 + 0.1 * opp_atk_str)
                defcon_prob = min(defcon_prob, 0.70 if pos == 2 else 0.40 if pos == 3 else 0.20)

                defcon_xpts = 2.0 * defcon_prob * full_game_prob
                xpts += defcon_xpts

            # Accumulate xPts — important for DGWs where a player has 2 fixtures
            # in the same GW. Apply a 10% fatigue/rotation penalty to second-and-later
            # fixtures: in practice DGW players are subbed off earlier, sometimes
            # rotated for fitness, and rarely sustain peak intensity across both
            # matches. Without this the model treats DGWs as fully independent
            # fixture rolls and over-projects premium attackers.
            gw_xpts_so_far = player_gw_xpts.get(gw, 0)
            fixture_xpts = max(xpts, 0)
            if gw_xpts_so_far > 0:
                fixture_xpts *= 0.90
            player_gw_xpts[gw] = round(gw_xpts_so_far + fixture_xpts, 2)

            # Store breakdown for this fixture
            if pid not in xpts_breakdown:
                xpts_breakdown[pid] = {}
            if gw not in xpts_breakdown[pid]:
                xpts_breakdown[pid][gw] = {
                    "opponent": opp_short,
                    "home": fix["home"],
                    "xg_per90": round(xg_per90, 3),
                    "xa_per90": round(xa_per90, 3),
                    "adj_xg": round(adj_xg, 3),
                    "adj_xa": round(adj_xa, 3),
                    "play_prob": round(play_prob, 2),
                    "full_game_prob": round(full_game_prob, 2),
                    "expected_90s": round(expected_90s, 2),
                    "cs_prob": round(cs_prob, 3),
                    "opp_def_str": round(opp_def_str, 3),
                    "team_atk_str": round(team_atk_str, 3),
                    "opp_atk_str": round(opp_atk_str, 3),
                    "appearance_pts": round(2.0 * full_game_prob + 1.0 * max(play_prob - full_game_prob, 0), 2),
                    "goal_pts": round(adj_xg * expected_90s * PTS_GOAL.get(pos, 4), 2),
                    "assist_pts": round(adj_xa * expected_90s * PTS_ASSIST, 2),
                    "cs_pts": round(cs_prob * PTS_CS.get(pos, 0) * full_game_prob, 2),
                    "bonus_pts": round(PTS_BONUS_AVG * play_prob, 2),
                    "conceded_pts": round(-(expected_conceded * 0.5 * full_game_prob), 2) if pos in [1, 2] else 0,
                    "defcon_pts": round(defcon_xpts, 2) if defcon_per90 > 0 and pos in [2, 3, 4] and nineties >= 3 else 0,
                    "total": round(max(xpts, 0), 2),
                }

        # Apply blank GW override
        # Blank GW (0 fixtures) = 0 xPts — fixture loop won't have added anything,
        # but we set explicitly to be safe.
        # DGWs are already handled: the fixture loop processes both fixtures and
        # accumulates xPts, so a DGW player naturally gets ~2x a single-GW player.
        # This means DGW players only rank higher if their per-fixture xPts justify it.
        if team_fixture_counts:
            team_counts = team_fixture_counts.get(p["team_id"], {})
            for gw in list(player_gw_xpts.keys()):
                fixture_count = team_counts.get(gw, 1)
                if fixture_count == 0:
                    player_gw_xpts[gw] = 0.0

        # Total xPts over next 6 GWs
        xpts_all[pid] = player_gw_xpts

    return xpts_all, xpts_breakdown


# ============================================================
# MILP SOLVER
# ============================================================

def solve_optimal_squad(players_df, xpts_col="xpts_total", budget=1000,
                        locked_ids=None, banned_ids=None):
    """
    XI-focused MILP squad optimisation with bench cost penalty.

    locked_ids: set of player IDs that MUST be in the squad
    banned_ids: set of player IDs that MUST NOT be in the squad

    Returns: DataFrame of selected 15 players, or None
    """
    BENCH_COST_PENALTY = 0.10

    if locked_ids is None:
        locked_ids = set()
    if banned_ids is None:
        banned_ids = set()

    eligible = players_df[
        (players_df["minutes"] > 45) &
        (players_df["status"].isin(["a", "d", ""])) &
        (players_df[xpts_col] > 0)
    ].copy()

    # Also include locked players even if they fail the minutes/status filter
    if locked_ids:
        locked_players = players_df[players_df["id"].isin(locked_ids)]
        eligible = pd.concat([eligible, locked_players]).drop_duplicates(subset="id")

    # Remove banned players
    if banned_ids:
        eligible = eligible[~eligible["id"].isin(banned_ids)]

    if len(eligible) < 15:
        return None, "Not enough eligible players"

    eligible[xpts_col] = eligible[xpts_col].fillna(0).astype(float)
    eligible["now_cost"] = eligible["now_cost"].fillna(0).astype(int)
    eligible = eligible[eligible[xpts_col].notna() & eligible["now_cost"].notna()]

    if len(eligible) < 15:
        return None, "Not enough eligible players after NaN removal"

    prob = LpProblem("FPL_Squad_XI_Focused", LpMaximize)

    eligible = eligible.reset_index(drop=True)
    pid_to_idx = {row["id"]: i for i, row in eligible.iterrows()}
    xpts_vals = eligible[xpts_col].tolist()
    cost_vals = eligible["now_cost"].tolist()
    pos_vals = eligible["pos_id"].tolist()
    team_vals = eligible["team_id"].tolist()
    player_ids = eligible["id"].tolist()

    # Decision variables
    s = {pid: LpVariable(f"s_{pid}", cat="Binary") for pid in player_ids}
    xi = {pid: LpVariable(f"xi_{pid}", cat="Binary") for pid in player_ids}

    # Objective: maximise XI xPts - penalise bench cost
    # bench[pid] = s[pid] - xi[pid] (1 if on bench, 0 otherwise)
    # bench_cost_penalty = sum(bench[pid] * cost[pid] * PENALTY)
    #
    # Expanded: XI_xPts - bench_cost_penalty
    # = sum(xi * xpts) - sum((s - xi) * cost * PENALTY)
    # = sum(xi * xpts) - sum(s * cost * PENALTY) + sum(xi * cost * PENALTY)
    # = sum(xi * (xpts + cost * PENALTY)) - sum(s * cost * PENALTY)
    prob += lpSum(
        xi[pid] * (xpts_vals[pid_to_idx[pid]] + cost_vals[pid_to_idx[pid]] * BENCH_COST_PENALTY)
        - s[pid] * cost_vals[pid_to_idx[pid]] * BENCH_COST_PENALTY
        for pid in player_ids
    )

    # --- SQUAD constraints (15 players) ---
    prob += lpSum(s[pid] for pid in player_ids) == 15
    prob += lpSum(s[pid] * cost_vals[pid_to_idx[pid]] for pid in player_ids) <= budget

    for pos_id, count in [(1, 2), (2, 5), (3, 5), (4, 3)]:
        pos_pids = [pid for pid in player_ids if pos_vals[pid_to_idx[pid]] == pos_id]
        prob += lpSum(s[pid] for pid in pos_pids) == count

    for team_id in set(team_vals):
        team_pids = [pid for pid in player_ids if team_vals[pid_to_idx[pid]] == team_id]
        prob += lpSum(s[pid] for pid in team_pids) <= 3

    # --- LOCKED players: must be in squad ---
    for pid in player_ids:
        if pid in locked_ids:
            prob += s[pid] == 1

    # --- XI constraints (11 from the 15) ---
    prob += lpSum(xi[pid] for pid in player_ids) == 11
    for pid in player_ids:
        prob += xi[pid] <= s[pid]

    gk_pids = [pid for pid in player_ids if pos_vals[pid_to_idx[pid]] == 1]
    def_pids = [pid for pid in player_ids if pos_vals[pid_to_idx[pid]] == 2]
    mid_pids = [pid for pid in player_ids if pos_vals[pid_to_idx[pid]] == 3]
    fwd_pids = [pid for pid in player_ids if pos_vals[pid_to_idx[pid]] == 4]

    prob += lpSum(xi[pid] for pid in gk_pids) == 1
    prob += lpSum(xi[pid] for pid in def_pids) >= 3
    prob += lpSum(xi[pid] for pid in def_pids) <= 5
    prob += lpSum(xi[pid] for pid in mid_pids) >= 2
    prob += lpSum(xi[pid] for pid in mid_pids) <= 5
    prob += lpSum(xi[pid] for pid in fwd_pids) >= 1
    prob += lpSum(xi[pid] for pid in fwd_pids) <= 3

    try:
        solver = PULP_CBC_CMD(msg=0, timeLimit=30)
        prob.solve(solver)
    except Exception as e:
        return None, f"Solver error: {e}"

    if LpStatus[prob.status] != "Optimal":
        return None, f"Solver status: {LpStatus[prob.status]}"

    selected_ids = [pid for pid in player_ids if value(s[pid]) is not None and value(s[pid]) > 0.5]
    xi_ids = [pid for pid in player_ids if value(xi[pid]) is not None and value(xi[pid]) > 0.5]
    squad = eligible[eligible["id"].isin(selected_ids)].copy()
    squad["is_xi"] = squad["id"].isin(xi_ids)

    return squad, None


def solve_best_xi(squad_df, xpts_col="xpts_next_gw"):
    """
    From a 15-man squad, pick the best starting XI using MILP.
    Must have exactly 1 GK, and valid formation (3-5-2, 4-4-2, etc.)
    """
    if squad_df is None or len(squad_df) < 11:
        return None, None

    prob = LpProblem("FPL_XI", LpMaximize)
    squad_df = squad_df.reset_index(drop=True)
    pids = squad_df["id"].tolist()
    pid_to_idx = {row["id"]: i for i, row in squad_df.iterrows()}
    xpts_vals = squad_df[xpts_col].fillna(0).tolist()
    pos_list = squad_df["pos_id"].tolist()
    x = {pid: LpVariable(f"xi_{pid}", cat="Binary") for pid in pids}

    # Objective
    prob += lpSum(x[pid] * xpts_vals[pid_to_idx[pid]] for pid in pids)

    # Exactly 11 starters
    prob += lpSum(x[pid] for pid in pids) == 11

    # Exactly 1 GK
    gk_pids = [pid for pid in pids if pos_list[pid_to_idx[pid]] == 1]
    prob += lpSum(x[pid] for pid in gk_pids) == 1

    # DEF: 3-5
    def_pids = [pid for pid in pids if pos_list[pid_to_idx[pid]] == 2]
    prob += lpSum(x[pid] for pid in def_pids) >= 3
    prob += lpSum(x[pid] for pid in def_pids) <= 5

    # MID: 2-5
    mid_pids = [pid for pid in pids if pos_list[pid_to_idx[pid]] == 3]
    prob += lpSum(x[pid] for pid in mid_pids) >= 2
    prob += lpSum(x[pid] for pid in mid_pids) <= 5

    # FWD: 1-3
    fwd_pids = [pid for pid in pids if pos_list[pid_to_idx[pid]] == 4]
    prob += lpSum(x[pid] for pid in fwd_pids) >= 1
    prob += lpSum(x[pid] for pid in fwd_pids) <= 3

    try:
        solver = PULP_CBC_CMD(msg=0, timeLimit=15)
        prob.solve(solver)
    except Exception:
        return None, None

    if LpStatus[prob.status] != "Optimal":
        return None, None

    xi_ids = [pid for pid in pids if value(x[pid]) is not None and value(x[pid]) > 0.5]
    xi = squad_df[squad_df["id"].isin(xi_ids)].copy()
    bench = squad_df[~squad_df["id"].isin(xi_ids)].copy()
    return xi, bench


# ============================================================
# DATA ENRICHMENT
# ============================================================

@st.cache_data(ttl=3600, show_spinner=False)
def enrich_data(bootstrap, fixtures, team_odds):
    """Combine all data sources into a single enriched DataFrame.

    Cached for 1 hour because this is the heavyweight pipeline (xPts model
    over ~700 players × 6 GWs). Inputs are already-cached outputs of
    load_fpl_data / load_betting_odds, so the cache key is stable across
    Streamlit reruns until those inner caches refresh.
    """
    players_raw = bootstrap["elements"]
    teams = {t["id"]: t for t in bootstrap["teams"]}
    events = bootstrap["events"]

    # Determine the NEXT gameweek to plan for (transfers are for the future)
    # Priority: is_next (upcoming), or if is_current is not finished yet use that,
    # otherwise find the first unfinished GW
    next_gw = next((e for e in events if e.get("is_next")), None)
    current_gw_obj = next((e for e in events if e.get("is_current")), None)

    if next_gw:
        planning_gw = next_gw
    elif current_gw_obj and not current_gw_obj.get("finished"):
        planning_gw = current_gw_obj
    else:
        # All current GWs finished, find first unfinished
        unfinished = [e for e in events if not e.get("finished")]
        planning_gw = unfinished[0] if unfinished else (events[-1] if events else None)

    # For display purposes, current_gw is the latest active/completed
    current_gw = current_gw_obj or planning_gw

    # Planning GW ID — this is what we use for fixture windows and xPts
    planning_gw_id = planning_gw["id"] if planning_gw else 1
    gw_id = planning_gw_id  # all fixture lookups use this

    # Build upcoming fixtures
    upcoming = {t_id: [] for t_id in teams}
    for f in sorted(fixtures, key=lambda x: x.get("event", 0) or 0):
        ev = f.get("event")
        if ev and gw_id <= ev < gw_id + 6:
            if f["team_h"] in upcoming:
                upcoming[f["team_h"]].append({
                    "gw": ev, "opp": f["team_a"], "home": True,
                    "difficulty": f.get("team_h_difficulty", 3)
                })
            if f["team_a"] in upcoming:
                upcoming[f["team_a"]].append({
                    "gw": ev, "opp": f["team_h"], "home": False,
                    "difficulty": f.get("team_a_difficulty", 3)
                })

    # Recent results
    recent = {t_id: [] for t_id in teams}
    for f in sorted(fixtures, key=lambda x: x.get("event", 0) or 0, reverse=True):
        if f.get("finished") and f.get("team_h_score") is not None:
            h = "W" if f["team_h_score"] > f["team_a_score"] else ("D" if f["team_h_score"] == f["team_a_score"] else "L")
            a = "W" if f["team_a_score"] > f["team_h_score"] else ("D" if f["team_a_score"] == f["team_h_score"] else "L")
            if len(recent.get(f["team_h"], [])) < 5:
                recent[f["team_h"]].append(h)
            if len(recent.get(f["team_a"], [])) < 5:
                recent[f["team_a"]].append(a)

    # Build player rows
    rows = []
    for p in players_raw:
        td = teams.get(p["team"], {})
        price = p["now_cost"] / 10
        mins = p.get("minutes", 0) or 0
        pts = p.get("total_points", 0) or 0

        rows.append({
            "id": p["id"],
            "name": p.get("web_name", ""),
            "first_name": p.get("first_name", ""),
            "second_name": p.get("second_name", ""),
            "team_id": p["team"],
            "team": td.get("short_name", "???"),
            "team_name": td.get("name", "???"),
            "team_code": td.get("code", 0),
            "pos_id": p["element_type"],
            "pos": POS_MAP.get(p["element_type"], "?"),
            "price": price,
            "now_cost": p["now_cost"],
            "total_points": pts,
            "form": float(p.get("form", 0) or 0),
            "form_str": str(p.get("form", "0.0")),
            "ict_index": round(float(p.get("ict_index", 0) or 0), 1),
            "minutes": mins,
            "starts": p.get("starts", 0) or 0,
            "goals": p.get("goals_scored", 0) or 0,
            "assists": p.get("assists", 0) or 0,
            "clean_sheets": p.get("clean_sheets", 0) or 0,
            "xg_per90": float(p.get("expected_goals_per_90", 0) or 0),
            "xa_per90": float(p.get("expected_assists_per_90", 0) or 0),
            "xgi_per90": float(p.get("expected_goal_involvements_per_90", 0) or 0),
            "xgc_per90": float(p.get("expected_goals_conceded_per_90", 0) or 0),
            # Season totals for over/underperformance regression
            "xg_total": float(p.get("expected_goals", 0) or 0),
            "xa_total": float(p.get("expected_assists", 0) or 0),
            # Set piece taker status (1 = first choice, None/0 = not)
            "penalties_order": p.get("penalties_order") or 0,
            "corners_order": p.get("corners_and_indirect_freekicks_order") or 0,
            "freekicks_order": p.get("direct_freekicks_order") or 0,
            # Defensive contributions (DefCon) — new for 2025/26
            "defcon_total": int(p.get("defensive_contributions", 0) or 0),
            "defcon_per90": float(p.get("defensive_contribution_per_90", 0) or 0),
            "selected_pct": float(p.get("selected_by_percent", 0) or 0),
            "transfers_in": p.get("transfers_in_event", 0) or 0,
            "transfers_out": p.get("transfers_out_event", 0) or 0,
            "status": p.get("status", "a"),
            "chance_playing": p.get("chance_of_playing_next_round"),
            "news": p.get("news", ""),
            "ppg": float(p.get("points_per_game", 0) or 0),
            "upcoming": upcoming.get(p["team"], []),
            "team_form": recent.get(p["team"], []),
        })

    df = pd.DataFrame(rows)

    # Load form-weighted xG data from recent GWs
    player_gw_data = load_recent_gw_live_data(gw_id, n_recent=7)
    form_xg_data = compute_form_weighted_xg(player_gw_data, n_recent=7)

    # Compute rotation risk from recent start patterns
    rotation_data = compute_rotation_risk(player_gw_data, gw_id, n_recent=7)

    # Detect blank/double gameweeks
    team_fixture_counts = detect_blank_double_gws(fixtures, gw_id, n_gws=6, teams=teams)

    # Load Club Elo ratings
    elo_ratings, elo_err = load_club_elo()

    # Load live fixture-specific odds (The Odds API)
    live_odds_data, live_odds_err = load_live_odds()

    # Build xPts model (uses planning_gw_id, so only future fixtures)
    xpts_map, xpts_breakdown = build_xpts_model(df, team_odds, teams, fixtures, gw_id,
                                 form_xg_data=form_xg_data,
                                 team_fixture_counts=team_fixture_counts,
                                 elo_ratings=elo_ratings,
                                 live_odds=live_odds_data,
                                 rotation_data=rotation_data)

    # Add xPts columns
    # xpts_next_gw = xPts for the specific next gameweek (planning_gw_id)
    df["xpts_next_gw"] = df["id"].map(
        lambda pid: xpts_map.get(pid, {}).get(planning_gw_id, 0)
    )
    # xpts_total = sum of xPts over next 6 GWs (all from planning_gw_id onwards)
    df["xpts_total"] = df["id"].map(
        lambda pid: sum(xpts_map.get(pid, {}).values())
    )

    # Avg fixture difficulty
    df["avg_difficulty"] = df["upcoming"].apply(
        lambda u: round(np.mean([f["difficulty"] for f in u[:4]]), 2) if u else 3.0
    )

    # Value metric
    df["value"] = df.apply(
        lambda r: round(r["xpts_total"] / max(r["price"], 1), 2), axis=1
    )

    # Add rotation risk data
    df["rotation_risk"] = df["id"].map(
        lambda pid: rotation_data.get(pid, {}).get("rotation_risk", "unknown")
    )
    df["start_prob"] = df["id"].map(
        lambda pid: rotation_data.get(pid, {}).get("projected_start_prob", 0)
    )

    return df, teams, current_gw, planning_gw_id, upcoming, fixtures, xpts_map, team_fixture_counts, xpts_breakdown


# ============================================================
# MANAGER TEAM FETCHER
# ============================================================

@st.cache_data(ttl=1800)
def fetch_manager_team(manager_id, current_gw_id):
    """
    Fetch a manager's current squad, bank, free transfers,
    and purchase prices (for correct selling price calculation).
    """
    try:
        headers = {"User-Agent": "FPL-Optimizer/2.0"}

        # 1. Basic manager info
        entry = requests.get(
            f"{FPL_BASE}/entry/{manager_id}/",
            headers=headers, timeout=15,
        ).json()

        manager_name = f"{entry.get('player_first_name', '')} {entry.get('player_last_name', '')}"
        team_name = entry.get("name", "Unknown")
        overall_rank = entry.get("summary_overall_rank") or "-"
        total_points = entry.get("summary_overall_points", 0)

        # 2. History endpoint — gives bank and transfer count per GW
        history = requests.get(
            f"{FPL_BASE}/entry/{manager_id}/history/",
            headers=headers, timeout=15,
        ).json()

        current_hist = history.get("current", [])
        bank = 0
        free_transfers = 1  # default

        if current_hist:
            latest_gw = current_hist[-1]
            bank = latest_gw.get("bank", 0)

            # Calculate free transfers available for NEXT gameweek
            # Logic: start with 1 FT per GW, can bank up to max 5
            # We look at all GWs to count how many were banked
            #
            # Important: the team's first GW (started_event) doesn't bank a FT
            # even if event_transfers = 0, because that's the initial squad setup
            started_event = entry.get("started_event", 1)

            ft = 1  # everyone gets 1 at start

            # Build set of GWs where chips were played
            chips_played_gws = {}  # {gw: chip_name}
            for c in history.get("chips", []):
                c_event = c.get("event")
                c_name = c.get("name", "")
                if c_event:
                    chips_played_gws[c_event] = c_name

            for gw_data in current_hist:
                gw_number = gw_data.get("event", 0)
                transfers_made = gw_data.get("event_transfers", 0)
                transfers_cost = gw_data.get("event_transfers_cost", 0)

                # Stop at the in-progress (or later) GW. We want `ft` to represent
                # the FT count entering current_gw_id, which is the value AFTER
                # processing every completed GW strictly before it. If we process
                # the in-progress GW too, `ft` advances one GW too far and the user
                # sees the count for the GW after the one they're planning for.
                if gw_number >= current_gw_id:
                    break

                # Skip the GW the team was created — no FT banking on first GW
                if gw_number <= started_event:
                    ft = 1
                    continue

                chip_this_gw = chips_played_gws.get(gw_number, "")

                # Free Hit: FTs are frozen — you don't gain or lose any
                # You come out of FH with the same FTs you had going in
                if chip_this_gw == "freehit":
                    continue  # skip entirely, FTs unchanged

                # Wildcard: unlimited transfers, FTs reset to 1 afterwards
                if chip_this_gw == "wildcard":
                    ft = 1
                    continue

                if transfers_cost > 0:
                    # They took hits: used all FTs + some extra, reset to 1
                    ft = 1
                elif transfers_made == 0:
                    # Banked a FT
                    ft = min(ft + 1, 5)
                elif transfers_made <= ft:
                    # Used some/all FTs without a hit
                    remaining = ft - transfers_made
                    ft = min(remaining + 1, 5)  # +1 for the new GW's FT
                else:
                    ft = 1

            free_transfers = ft

        # Most recent completed GW summary (for landing block)
        last_gw_points = None
        last_gw_rank = None
        last_gw_id = None
        if current_hist:
            _last = current_hist[-1]
            last_gw_id = _last.get("event")
            last_gw_points = _last.get("points")
            last_gw_rank = _last.get("rank")

        # Extract chips already played
        chips_played = history.get("chips", [])
        # chips_played is a list of {"name": "wildcard", "time": "...", "event": 12}

        # 2025/26 chip rules:
        # Each chip refreshes at GW20. You get 1 of each per half-season.
        # First half: GW1-19, Second half: GW20-38
        # So at any point, you have 0 or 1 of each chip remaining for the
        # CURRENT half-season.
        half_gw = LC.get("half_season_gw", 19)
        current_half = 1 if current_gw_id <= half_gw else 2

        chips_remaining = {}
        for chip_name in ["wildcard", "freehit", "3xc", "bboost"]:
            # Check if this chip was played in the current half
            if current_half == 1:
                used_this_half = any(
                    c.get("name") == chip_name and c.get("event", 0) <= half_gw
                    for c in chips_played
                )
            else:
                used_this_half = any(
                    c.get("name") == chip_name and c.get("event", 0) >= half_gw + 1
                    for c in chips_played
                )
            chips_remaining[chip_name] = 0 if used_this_half else 1

        # 3. Transfer history — gives purchase prices (element_in_cost)
        transfers = requests.get(
            f"{FPL_BASE}/entry/{manager_id}/transfers/",
            headers=headers, timeout=15,
        ).json()

        # Build purchase price map: player_id -> purchase_price (in 0.1m units)
        # Latest transfer for each player is their purchase price
        purchase_prices = {}
        for t in transfers:
            purchase_prices[t["element_in"]] = t["element_in_cost"]

        # 4. Current picks
        # Important: if the current/most recent GW was a Free Hit, the picks endpoint
        # returns the FH squad, not the real squad. We need to detect this and
        # fall back to the GW before the FH to get the actual squad.
        picks_data = None
        active_chip = None

        for gw in [current_gw_id, current_gw_id - 1]:
            if gw < 1:
                continue
            try:
                resp = requests.get(
                    f"{FPL_BASE}/entry/{manager_id}/event/{gw}/picks/",
                    headers=headers, timeout=15,
                )
                if resp.status_code == 200:
                    pd_temp = resp.json()
                    chip_active = pd_temp.get("active_chip")

                    if chip_active == "freehit":
                        # This GW was a Free Hit — skip it and use the previous GW's squad
                        # (which is the real squad that reverts after FH)
                        active_chip = "freehit"
                        continue
                    else:
                        picks_data = pd_temp
                        if active_chip != "freehit":
                            active_chip = chip_active
                        break
            except Exception:
                continue

        # If both GWs were FH (shouldn't happen) or no data, try going further back
        if picks_data is None:
            for gw in range(current_gw_id - 2, max(0, current_gw_id - 5), -1):
                if gw < 1:
                    continue
                try:
                    resp = requests.get(
                        f"{FPL_BASE}/entry/{manager_id}/event/{gw}/picks/",
                        headers=headers, timeout=15,
                    )
                    if resp.status_code == 200:
                        pd_temp = resp.json()
                        if pd_temp.get("active_chip") != "freehit":
                            picks_data = pd_temp
                            break
                except Exception:
                    continue

        if picks_data is None:
            return None, "Could not fetch team picks"

        picks = picks_data.get("picks", [])
        active_chip = picks_data.get("active_chip")

        # entry_history within picks gives exact FT info in some API versions
        entry_hist = picks_data.get("entry_history", {})
        if entry_hist:
            bank = entry_hist.get("bank", bank)

        squad_ids = [p["element"] for p in picks]
        captains = {p["element"]: p.get("is_captain", False) for p in picks}
        vice_captains = {p["element"]: p.get("is_vice_captain", False) for p in picks}
        positions_in_team = {p["element"]: p.get("position", 0) for p in picks}

        # picks may contain selling_price in newer API
        selling_prices_api = {}
        for p in picks:
            if "selling_price" in p:
                selling_prices_api[p["element"]] = p["selling_price"]

        return {
            "manager_name": manager_name.strip(),
            "team_name": team_name,
            "overall_rank": overall_rank,
            "total_points": total_points,
            "bank": bank,
            "free_transfers": free_transfers,
            "squad_ids": squad_ids,
            "captains": captains,
            "vice_captains": vice_captains,
            "positions": positions_in_team,
            "active_chip": active_chip,
            "purchase_prices": purchase_prices,
            "selling_prices_api": selling_prices_api,
            "chips_remaining": chips_remaining,
            "last_gw_id": last_gw_id,
            "last_gw_points": last_gw_points,
            "last_gw_rank": last_gw_rank,
        }, None

    except requests.exceptions.HTTPError:
        return None, "Manager ID not found"
    except Exception as e:
        return None, str(e)


def calculate_selling_price(player_id, current_price, purchase_prices, selling_prices_api):
    """
    Calculate the correct FPL selling price.
    Rule: you get 50% of profit (rounded down).
    selling_price = purchase_price + floor((current_price - purchase_price) / 2)
    If current_price < purchase_price, selling_price = current_price (full loss).
    """
    # If the API gave us the selling price directly, use it
    if player_id in selling_prices_api:
        return selling_prices_api[player_id]

    purchase = purchase_prices.get(player_id, current_price)
    if current_price <= purchase:
        return current_price  # no profit or a loss — sell at current
    profit = current_price - purchase
    return purchase + (profit // 2)  # 50% of profit, rounded down


def find_optimal_transfers(squad_df, all_players_df, bank, free_transfers,
                           purchase_prices, selling_prices_api,
                           n_transfers=1, xpts_col="xpts_total",
                           hit_cost=4):
    """
    Find the best transfers using correct sale/buy prices and hit-aware logic.

    For each candidate transfer:
    - OUT player sold at their SELLING price (50% profit rule)
    - IN player bought at current market price (now_cost)
    - If n_transfers > free_transfers, apply -4 per extra transfer
    - Only suggest if net xPts gain > hit penalty
    """
    if squad_df is None or len(squad_df) == 0:
        return []

    squad_ids = set(squad_df["id"].tolist())

    # Available players not in squad
    available = all_players_df[
        (~all_players_df["id"].isin(squad_ids)) &
        (all_players_df["minutes"] > 45) &
        (all_players_df["status"].isin(["a", "d", ""])) &
        (all_players_df[xpts_col] > 0)
    ].copy()

    # Calculate selling price for each squad player
    squad_df = squad_df.copy()
    squad_df["sell_price"] = squad_df.apply(
        lambda r: calculate_selling_price(
            r["id"], r["now_cost"], purchase_prices, selling_prices_api
        ), axis=1
    )

    # Hit penalty: transfers beyond free ones cost 4 pts each
    extra_transfers = max(0, n_transfers - free_transfers)
    total_hit = extra_transfers * hit_cost

    suggestions = []

    if n_transfers == 1:
        for _, out_p in squad_df.iterrows():
            # Budget available = bank + selling price of outgoing player
            budget_available = bank + out_p["sell_price"]
            remaining_squad = squad_df[squad_df["id"] != out_p["id"]]
            team_counts = remaining_squad["team_id"].value_counts().to_dict()

            cands = available[
                (available["pos_id"] == out_p["pos_id"]) &
                (available["now_cost"] <= budget_available)
            ].copy()
            cands = cands[cands["team_id"].map(lambda tid: team_counts.get(tid, 0) < 3)]

            if len(cands) == 0:
                continue

            best_in = cands.loc[cands[xpts_col].idxmax()]
            xpts_gain = best_in[xpts_col] - out_p[xpts_col]
            net_gain = xpts_gain - total_hit

            if net_gain > 0.5:  # only suggest if meaningfully better
                suggestions.append({
                    "out": [out_p.to_dict()],
                    "in": [best_in.to_dict()],
                    "xpts_gain": round(xpts_gain, 1),
                    "net_gain": round(net_gain, 1),
                    "hit": total_hit,
                    "cost_change": round((best_in["now_cost"] - out_p["sell_price"]) / 10, 1),
                    "budget_after": round((budget_available - best_in["now_cost"]) / 10, 1),
                })

    elif n_transfers >= 2:
        squad_list = squad_df.to_dict("records")
        seen = set()
        for i, out1 in enumerate(squad_list):
            for j, out2 in enumerate(squad_list):
                if j <= i:
                    continue
                pair_key = (min(out1["id"], out2["id"]), max(out1["id"], out2["id"]))
                if pair_key in seen:
                    continue
                seen.add(pair_key)

                freed = bank + out1["sell_price"] + out2["sell_price"]
                remaining = squad_df[~squad_df["id"].isin([out1["id"], out2["id"]])]
                tc = remaining["team_id"].value_counts().to_dict()

                best_pair_xpts = 0
                best_pair = None

                cands1 = available[
                    (available["pos_id"] == out1["pos_id"]) &
                    (available["now_cost"] <= freed)
                ]
                cands1 = cands1[cands1["team_id"].map(lambda t: tc.get(t, 0) < 3)]

                for _, in1 in cands1.nlargest(5, xpts_col).iterrows():
                    remaining_budget = freed - in1["now_cost"]
                    tc2 = tc.copy()
                    tc2[in1["team_id"]] = tc2.get(in1["team_id"], 0) + 1

                    cands2 = available[
                        (available["pos_id"] == out2["pos_id"]) &
                        (available["now_cost"] <= remaining_budget) &
                        (available["id"] != in1["id"])
                    ]
                    cands2 = cands2[cands2["team_id"].map(lambda t: tc2.get(t, 0) < 3)]

                    if len(cands2) == 0:
                        continue

                    in2 = cands2.loc[cands2[xpts_col].idxmax()]
                    pair_xpts = in1[xpts_col] + in2[xpts_col]

                    if pair_xpts > best_pair_xpts:
                        best_pair_xpts = pair_xpts
                        best_pair = (in1, in2)

                if best_pair:
                    old_xpts = out1[xpts_col] + out2[xpts_col]
                    xpts_gain = best_pair_xpts - old_xpts
                    net_gain = xpts_gain - total_hit

                    if net_gain > 0.5:
                        total_in_cost = best_pair[0]["now_cost"] + best_pair[1]["now_cost"]
                        suggestions.append({
                            "out": [out1, out2],
                            "in": [best_pair[0].to_dict(), best_pair[1].to_dict()],
                            "xpts_gain": round(xpts_gain, 1),
                            "net_gain": round(net_gain, 1),
                            "hit": total_hit,
                            "cost_change": round((total_in_cost - out1["sell_price"] - out2["sell_price"]) / 10, 1),
                            "budget_after": round((freed - total_in_cost) / 10, 1),
                        })

    suggestions.sort(key=lambda x: x["net_gain"], reverse=True)
    return suggestions[:10]




def solve_best_xi_for_gw(squad_df, xpts_map, gw_id):
    """Pick best starting XI from 15-man squad for a specific gameweek.
    Players with 0 xPts (blanking) are heavily penalised to avoid starting them."""
    if squad_df is None or len(squad_df) < 11:
        return None, None

    # Build per-GW xPts column
    sq = squad_df.copy()
    sq["xpts_gw"] = sq["id"].map(lambda pid: xpts_map.get(pid, {}).get(gw_id, 0))
    sq["xpts_gw"] = sq["xpts_gw"].fillna(0)

    # For the solver objective, penalise blanking players (xPts=0) heavily
    # so they only start if there's literally no valid formation without them
    sq["xpts_solver"] = sq["xpts_gw"].apply(lambda v: v if v > 0 else -5.0)

    prob = LpProblem(f"FPL_XI_GW{gw_id}", LpMaximize)
    sq = sq.reset_index(drop=True)
    pids = sq["id"].tolist()
    pid_to_idx = {row["id"]: i for i, row in sq.iterrows()}
    xpts_vals = sq["xpts_solver"].tolist()
    pos_list = sq["pos_id"].tolist()
    x = {pid: LpVariable(f"xi_{gw_id}_{pid}", cat="Binary") for pid in pids}

    prob += lpSum(x[pid] * xpts_vals[pid_to_idx[pid]] for pid in pids)
    prob += lpSum(x[pid] for pid in pids) == 11

    gk_pids = [pid for pid in pids if pos_list[pid_to_idx[pid]] == 1]
    prob += lpSum(x[pid] for pid in gk_pids) == 1
    def_pids = [pid for pid in pids if pos_list[pid_to_idx[pid]] == 2]
    prob += lpSum(x[pid] for pid in def_pids) >= 3
    prob += lpSum(x[pid] for pid in def_pids) <= 5
    mid_pids = [pid for pid in pids if pos_list[pid_to_idx[pid]] == 3]
    prob += lpSum(x[pid] for pid in mid_pids) >= 2
    prob += lpSum(x[pid] for pid in mid_pids) <= 5
    fwd_pids = [pid for pid in pids if pos_list[pid_to_idx[pid]] == 4]
    prob += lpSum(x[pid] for pid in fwd_pids) >= 1
    prob += lpSum(x[pid] for pid in fwd_pids) <= 3

    try:
        prob.solve(PULP_CBC_CMD(msg=0, timeLimit=10))
    except Exception:
        return None, None

    if LpStatus[prob.status] != "Optimal":
        return None, None

    xi_ids = [pid for pid in pids if value(x[pid]) is not None and value(x[pid]) > 0.5]
    xi = sq[sq["id"].isin(xi_ids)].copy()
    bench = sq[~sq["id"].isin(xi_ids)].copy()
    return xi, bench


def find_best_single_transfer_for_gw(squad_df, all_players_df, bank,
                                      purchase_prices, selling_prices_api,
                                      xpts_map, gw_id, exclude_ids=None,
                                      horizon_end=None):
    """
    Find the single best transfer considering the REMAINING HORIZON.
    
    For a GW31 transfer with horizon ending at GW35:
    - Compare players on sum(xPts from GW31 to GW35), not just GW31
    - This ensures transfers are forward-looking, not myopic
    
    Also returns the single-GW gain for display purposes.
    """
    if squad_df is None or len(squad_df) == 0:
        return None

    if exclude_ids is None:
        exclude_ids = set()
    
    # Default horizon: 6 GWs from this GW
    if horizon_end is None:
        horizon_end = gw_id + 6
    
    horizon_gws = list(range(gw_id, horizon_end))

    squad_ids = set(squad_df["id"].tolist()) | exclude_ids
    available = all_players_df[
        (~all_players_df["id"].isin(squad_ids)) &
        (all_players_df["minutes"] > 45) &
        (all_players_df["status"].isin(["a", "d", ""]))
    ].copy()

    # Remaining horizon xPts (what matters for the transfer decision)
    squad_df = squad_df.copy()
    squad_df["xpts_horizon"] = squad_df["id"].map(
        lambda pid: sum(xpts_map.get(pid, {}).get(gw, 0) for gw in horizon_gws)
    )
    available["xpts_horizon"] = available["id"].map(
        lambda pid: sum(xpts_map.get(pid, {}).get(gw, 0) for gw in horizon_gws)
    )
    
    # Also get single-GW xPts for display
    squad_df["xpts_gw"] = squad_df["id"].map(lambda pid: xpts_map.get(pid, {}).get(gw_id, 0))
    available["xpts_gw"] = available["id"].map(lambda pid: xpts_map.get(pid, {}).get(gw_id, 0))

    squad_df["sell_price"] = squad_df.apply(
        lambda r: calculate_selling_price(r["id"], r["now_cost"], purchase_prices, selling_prices_api),
        axis=1,
    )

    best = None
    best_gain = -999

    for _, out_p in squad_df.iterrows():
        budget_avail = bank + out_p["sell_price"]
        remaining = squad_df[squad_df["id"] != out_p["id"]]
        tc = remaining["team_id"].value_counts().to_dict()

        cands = available[
            (available["pos_id"] == out_p["pos_id"]) &
            (available["now_cost"] <= budget_avail)
        ]
        cands = cands[cands["team_id"].map(lambda tid: tc.get(tid, 0) < 3)]

        if len(cands) == 0:
            continue

        # Pick best by HORIZON xPts, not single GW
        top = cands.loc[cands["xpts_horizon"].idxmax()]
        horizon_gain = top["xpts_horizon"] - out_p["xpts_horizon"]
        gw_gain = top["xpts_gw"] - out_p["xpts_gw"]
        
        if horizon_gain > best_gain and horizon_gain > 0.05:
            best_gain = horizon_gain
            best = {
                "out": out_p.to_dict(),
                "in": top.to_dict(),
                "xpts_gain": round(horizon_gain, 2),  # horizon gain for decision-making
                "xpts_gw_gain": round(gw_gain, 2),     # single GW gain for display
                "new_bank": int(budget_avail - top["now_cost"]),
            }

    return best


def solve_free_hit_squad(all_players_df, xpts_map, gw_id, budget=1000, locked_ids=None,
                         max_per_team=3):
    """Free Hit / Loan Rangers: pick best possible 15-man squad for a single GW.
    
    max_per_team: 3 for normal Free Hit, None/999 for Loan Rangers (no team limit).
    """
    if locked_ids is None:
        locked_ids = set()
    if max_per_team is None:
        max_per_team = 999  # effectively unlimited
    eligible = all_players_df[
        (all_players_df["minutes"] > 45) &
        (all_players_df["status"].isin(["a", "d", ""]))
    ].copy()
    # Include locked players even if they fail filters
    if locked_ids:
        locked_players = all_players_df[all_players_df["id"].isin(locked_ids)]
        eligible = pd.concat([eligible, locked_players]).drop_duplicates(subset="id")
    eligible["xpts_gw"] = eligible["id"].map(lambda pid: xpts_map.get(pid, {}).get(gw_id, 0))
    has_fixture = eligible["xpts_gw"] > 0
    is_locked = eligible["id"].isin(locked_ids)
    eligible = eligible[has_fixture | is_locked].copy()
    eligible["xpts_gw"] = eligible["xpts_gw"].fillna(0)
    eligible["now_cost"] = eligible["now_cost"].fillna(0).astype(int)
    if len(eligible) < 15:
        return None

    eligible = eligible.reset_index(drop=True)
    pid_map = {row["id"]: i for i, row in eligible.iterrows()}
    xv = eligible["xpts_gw"].tolist()
    cv = eligible["now_cost"].tolist()
    pv = eligible["pos_id"].tolist()
    tv = eligible["team_id"].tolist()
    pids = eligible["id"].tolist()

    prob = LpProblem(f"FH_GW{gw_id}", LpMaximize)
    x = {pid: LpVariable(f"fh_{gw_id}_{pid}", cat="Binary") for pid in pids}
    prob += lpSum(x[pid] * xv[pid_map[pid]] for pid in pids)
    prob += lpSum(x[pid] * cv[pid_map[pid]] for pid in pids) <= budget
    prob += lpSum(x[pid] for pid in pids) == 15
    for pos_id, cnt in [(1, 2), (2, 5), (3, 5), (4, 3)]:
        prob += lpSum(x[pid] for pid in pids if pv[pid_map[pid]] == pos_id) == cnt
    for tid in set(tv):
        prob += lpSum(x[pid] for pid in pids if tv[pid_map[pid]] == tid) <= max_per_team
    for pid in pids:
        if pid in locked_ids:
            prob += x[pid] == 1
    try:
        prob.solve(PULP_CBC_CMD(msg=0, timeLimit=30))
    except Exception:
        return None
    if LpStatus[prob.status] != "Optimal":
        return None
    sel = [pid for pid in pids if value(x[pid]) is not None and value(x[pid]) > 0.5]
    return eligible[eligible["id"].isin(sel)].copy()


def solve_wildcard_squad(all_players_df, xpts_map, planning_gw, n_future, budget=1000,
                         team_fixture_counts=None, locked_ids=None):
    """
    Wildcard: best 15-man squad optimised for total xPts over remaining GWs,
    with per-GW XI awareness and locked player support.
    """
    if locked_ids is None:
        locked_ids = set()
    gw_range = list(range(planning_gw, planning_gw + n_future))

    eligible = all_players_df[
        (all_players_df["minutes"] > 45) &
        (all_players_df["status"].isin(["a", "d", ""]))
    ].copy()

    # Include locked players even if they fail filters
    if locked_ids:
        locked_players = all_players_df[all_players_df["id"].isin(locked_ids)]
        eligible = pd.concat([eligible, locked_players]).drop_duplicates(subset="id")

    # Calculate total xPts across horizon
    eligible["xpts_rem"] = eligible["id"].map(
        lambda pid: sum(xpts_map.get(pid, {}).get(gw, 0) for gw in gw_range)
    )
    eligible = eligible[eligible["xpts_rem"] > 0].copy()

    # Re-include locked players even if xpts_rem is 0 (they might blank but user wants them)
    if locked_ids:
        locked_missing = all_players_df[
            (all_players_df["id"].isin(locked_ids)) &
            (~all_players_df["id"].isin(eligible["id"]))
        ].copy()
        if len(locked_missing) > 0:
            locked_missing["xpts_rem"] = locked_missing["id"].map(
                lambda pid: sum(xpts_map.get(pid, {}).get(gw, 0) for gw in gw_range)
            )
            eligible = pd.concat([eligible, locked_missing]).drop_duplicates(subset="id")
    eligible["xpts_rem"] = eligible["xpts_rem"].fillna(0)
    eligible["now_cost"] = eligible["now_cost"].fillna(0).astype(int)
    if len(eligible) < 15:
        return None

    # Pre-compute which players have a fixture in each GW
    # (a player has a fixture if their team has >= 1 fixture that GW)
    player_has_fixture = {}  # {pid: {gw: bool}}
    for _, p in eligible.iterrows():
        pid = p["id"]
        tid = p["team_id"]
        player_has_fixture[pid] = {}
        for gw in gw_range:
            if team_fixture_counts:
                fc = team_fixture_counts.get(tid, {}).get(gw, 1)
                player_has_fixture[pid][gw] = (fc > 0)
            else:
                # No fixture count data — check if xPts > 0 as proxy
                player_has_fixture[pid][gw] = (xpts_map.get(pid, {}).get(gw, 0) > 0)

    eligible = eligible.reset_index(drop=True)
    pid_map = {row["id"]: i for i, row in eligible.iterrows()}
    xv = eligible["xpts_rem"].tolist()
    cv = eligible["now_cost"].tolist()
    pv = eligible["pos_id"].tolist()
    tv = eligible["team_id"].tolist()
    pids = eligible["id"].tolist()

    prob = LpProblem("Wildcard", LpMaximize)
    x = {pid: LpVariable(f"wc_{pid}", cat="Binary") for pid in pids}

    # Per-GW XI variables: xi[gw][pid] = 1 if player starts in that GW
    # This properly handles blanks — a blanking player is benched that GW
    BENCH_WEIGHT = 0.05
    xi_gw = {}
    for gw in gw_range:
        xi_gw[gw] = {pid: LpVariable(f"wcxi_{gw}_{pid}", cat="Binary") for pid in pids}

    # Pre-compute per-player per-GW xPts
    player_gw_xpts = {}
    for pid in pids:
        player_gw_xpts[pid] = {}
        for gw in gw_range:
            player_gw_xpts[pid][gw] = xpts_map.get(pid, {}).get(gw, 0)

    # Objective: sum over all GWs of (XI players at full value + bench at discount)
    obj_terms = []
    for gw in gw_range:
        for pid in pids:
            gw_xpts = player_gw_xpts[pid][gw]
            obj_terms.append(xi_gw[gw][pid] * gw_xpts * (1.0 - BENCH_WEIGHT))
            obj_terms.append(x[pid] * gw_xpts * BENCH_WEIGHT)
    prob += lpSum(obj_terms)

    # Budget
    prob += lpSum(x[pid] * cv[pid_map[pid]] for pid in pids) <= budget

    # Squad = 15
    prob += lpSum(x[pid] for pid in pids) == 15

    # Per-GW XI constraints
    for gw in gw_range:
        # XI = 11 per GW, subset of squad
        prob += lpSum(xi_gw[gw][pid] for pid in pids) == 11
        for pid in pids:
            prob += xi_gw[gw][pid] <= x[pid]
            # Cannot start if blanking (0 xPts)
            if player_gw_xpts[pid][gw] == 0:
                prob += xi_gw[gw][pid] == 0

        # XI formation per GW
        prob += lpSum(xi_gw[gw][pid] for pid in pids if pv[pid_map[pid]] == 1) == 1
        prob += lpSum(xi_gw[gw][pid] for pid in pids if pv[pid_map[pid]] == 2) >= 3
        prob += lpSum(xi_gw[gw][pid] for pid in pids if pv[pid_map[pid]] == 2) <= 5
        prob += lpSum(xi_gw[gw][pid] for pid in pids if pv[pid_map[pid]] == 3) >= 2
        prob += lpSum(xi_gw[gw][pid] for pid in pids if pv[pid_map[pid]] == 3) <= 5
        prob += lpSum(xi_gw[gw][pid] for pid in pids if pv[pid_map[pid]] == 4) >= 1
        prob += lpSum(xi_gw[gw][pid] for pid in pids if pv[pid_map[pid]] == 4) <= 3

    # Position constraints for squad
    for pos_id, cnt in [(1, 2), (2, 5), (3, 5), (4, 3)]:
        prob += lpSum(x[pid] for pid in pids if pv[pid_map[pid]] == pos_id) == cnt

    # Max 3 per team
    for tid in set(tv):
        prob += lpSum(x[pid] for pid in pids if tv[pid_map[pid]] == tid) <= 3

    # Locked players must be in squad
    for pid in pids:
        if pid in locked_ids:
            prob += x[pid] == 1

    try:
        prob.solve(PULP_CBC_CMD(msg=0, timeLimit=60))
    except Exception:
        return None
    if LpStatus[prob.status] != "Optimal":
        # If infeasible (too many blanks), fall back to simpler model without per-GW XI
        prob2 = LpProblem("Wildcard_relaxed", LpMaximize)
        x2 = {pid: LpVariable(f"wc2_{pid}", cat="Binary") for pid in pids}
        prob2 += lpSum(x2[pid] * xv[pid_map[pid]] for pid in pids)
        prob2 += lpSum(x2[pid] * cv[pid_map[pid]] for pid in pids) <= budget
        prob2 += lpSum(x2[pid] for pid in pids) == 15
        for pos_id, cnt in [(1, 2), (2, 5), (3, 5), (4, 3)]:
            prob2 += lpSum(x2[pid] for pid in pids if pv[pid_map[pid]] == pos_id) == cnt
        for tid in set(tv):
            prob2 += lpSum(x2[pid] for pid in pids if tv[pid_map[pid]] == tid) <= 3
        try:
            prob2.solve(PULP_CBC_CMD(msg=0, timeLimit=45))
        except Exception:
            return None
        if LpStatus[prob2.status] != "Optimal":
            return None
        sel = [pid for pid in pids if value(x2[pid]) is not None and value(x2[pid]) > 0.5]
        return eligible[eligible["id"].isin(sel)].copy()

    sel = [pid for pid in pids if value(x[pid]) is not None and value(x[pid]) > 0.5]
    return eligible[eligible["id"].isin(sel)].copy()


def find_best_captain(squad_df, xpts_map, gw_id):
    """Find best captain (highest xPts) for a specific GW."""
    if squad_df is None or len(squad_df) == 0:
        return None
    sq = squad_df.copy()
    sq["xpts_gw"] = sq["id"].map(lambda pid: xpts_map.get(pid, {}).get(gw_id, 0))
    return sq.loc[sq["xpts_gw"].idxmax()]


def find_best_vice_captain(squad_df, xpts_map, gw_id, captain_id=None):
    """Find best vice captain (2nd highest xPts) for Dynamic Duo chip."""
    if squad_df is None or len(squad_df) == 0:
        return None
    sq = squad_df.copy()
    sq["xpts_gw"] = sq["id"].map(lambda pid: xpts_map.get(pid, {}).get(gw_id, 0))
    if captain_id is not None:
        sq = sq[sq["id"] != captain_id]
    if len(sq) == 0:
        return None
    return sq.loc[sq["xpts_gw"].idxmax()]


def build_rolling_plan(my_squad_df, all_players_df, bank, free_transfers,
                       purchase_prices, selling_prices_api, xpts_map,
                       planning_gw_id, n_gws=6, chip_schedule=None,
                       team_fixture_counts=None, locked_ids=None, banned_ids=None):
    """
    Chip-aware rolling planner.
    chip_schedule: {gw_id: chip_name} e.g. {31: "wildcard", 33: "bench_boost"}
    locked_ids: players that must NOT be sold (transfers won't suggest selling them)
    banned_ids: players that must NOT be bought (excluded from transfer candidates)
    """
    if chip_schedule is None:
        chip_schedule = {}
    if locked_ids is None:
        locked_ids = set()
    if banned_ids is None:
        banned_ids = set()

    plan = []
    current_squad = my_squad_df.copy()
    current_bank = bank
    current_ft = free_transfers
    current_purchase = purchase_prices.copy()
    current_selling = selling_prices_api.copy()
    pre_fh_squad = None

    for i in range(n_gws):
        gw = planning_gw_id + i
        chip = chip_schedule.get(gw, None)

        # Restore squad after free hit (must happen before any other logic)
        if pre_fh_squad is not None:
            current_squad = pre_fh_squad
            pre_fh_squad = None

        has_fixtures = any(xpts_map.get(pid, {}).get(gw, 0) > 0 for pid in current_squad["id"])
        if not has_fixtures and chip != "free_hit":
            # No fixtures for current squad this GW — but don't break the whole plan
            # Just record an empty GW and continue (the next GW might have fixtures)
            gw_entry = {
                "gw": gw, "chip": chip, "transfer": None, "hit": 0,
                "squad": current_squad.copy(), "xi": None, "bench": None,
                "captain": None, "captain_multiplier": 2, "bench_boost": False,
                "total_xpts": 0, "transfers": [], "ft_used": 0,
            }
            current_ft = min(current_ft + 1, 5)  # still gain an FT
            plan.append(gw_entry)
            continue

        gw_entry = {
            "gw": gw, "chip": chip, "transfer": None, "hit": 0,
            "squad": current_squad.copy(), "xi": None, "bench": None,
            "captain": None, "captain_multiplier": 2, "bench_boost": False,
        }

        # === FREE HIT ===
        if chip == "free_hit":
            pre_fh_squad = current_squad.copy()
            total_val = int(current_bank + current_squad["now_cost"].sum())
            fh_pool = all_players_df[~all_players_df["id"].isin(banned_ids)] if banned_ids else all_players_df
            fh_squad = solve_free_hit_squad(fh_pool, xpts_map, gw, total_val, locked_ids=locked_ids)
            if fh_squad is not None:
                gw_entry["squad"] = fh_squad
                xi, bench = solve_best_xi_for_gw(fh_squad, xpts_map, gw)
            else:
                xi, bench = solve_best_xi_for_gw(current_squad, xpts_map, gw)
            gw_entry["xi"] = xi
            gw_entry["bench"] = bench
            gw_entry["captain"] = find_best_captain(xi, xpts_map, gw) if xi is not None else None
            fh_total = 0
            if xi is not None and len(xi) > 0:
                fh_total = xi["id"].map(lambda pid: xpts_map.get(pid, {}).get(gw, 0)).sum()
                cap = gw_entry.get("captain")
                cap_mult = gw_entry.get("captain_multiplier", 2)
                if cap is not None:
                    cap_id = cap.get("id", 0) if isinstance(cap, dict) else getattr(cap, "id", 0)
                    cap_pts = xpts_map.get(cap_id, {}).get(gw, 0)
                    fh_total += cap_pts * (cap_mult - 1)
            gw_entry["total_xpts"] = round(fh_total, 1)
            current_ft = min(current_ft + 1, 5)
            plan.append(gw_entry)
            continue

        # === LOAN RANGERS (Allsvenskan — Free Hit with no team limit) ===
        if chip == "loan_rangers":
            pre_fh_squad = current_squad.copy()
            total_val = int(current_bank + current_squad["now_cost"].sum())
            lr_pool = all_players_df[~all_players_df["id"].isin(banned_ids)] if banned_ids else all_players_df
            lr_squad = solve_free_hit_squad(lr_pool, xpts_map, gw, total_val,
                                           locked_ids=locked_ids, max_per_team=None)
            if lr_squad is not None:
                gw_entry["squad"] = lr_squad
                xi, bench = solve_best_xi_for_gw(lr_squad, xpts_map, gw)
            else:
                xi, bench = solve_best_xi_for_gw(current_squad, xpts_map, gw)
            gw_entry["xi"] = xi
            gw_entry["bench"] = bench
            gw_entry["captain"] = find_best_captain(xi, xpts_map, gw) if xi is not None else None
            lr_total = 0
            if xi is not None and len(xi) > 0:
                lr_total = xi["id"].map(lambda pid: xpts_map.get(pid, {}).get(gw, 0)).sum()
                cap = gw_entry.get("captain")
                if cap is not None:
                    cap_id = cap.get("id", 0) if isinstance(cap, dict) else getattr(cap, "id", 0)
                    lr_total += xpts_map.get(cap_id, {}).get(gw, 0)  # captain ×2 → +1 extra
            gw_entry["total_xpts"] = round(lr_total, 1)
            current_ft = min(current_ft + 1, 5)
            plan.append(gw_entry)
            continue

        # === PARK THE BUS (Allsvenskan — double DEF points, no captain) ===
        if chip == "park_the_bus":
            xi, bench = solve_best_xi_for_gw(current_squad, xpts_map, gw)
            gw_entry["xi"] = xi
            gw_entry["bench"] = bench
            gw_entry["captain"] = None  # no captain with Park the Bus
            gw_entry["park_the_bus"] = True
            ptb_total = 0
            if xi is not None and len(xi) > 0:
                for _, p in xi.iterrows():
                    p_xpts = xpts_map.get(p["id"], {}).get(gw, 0)
                    if p["pos_id"] == 2:  # DEF
                        p_xpts *= 2  # double defender points
                    ptb_total += p_xpts
                # No captain bonus with Park the Bus
            gw_entry["total_xpts"] = round(ptb_total, 1)
            gw_entry["squad"] = current_squad.copy()
            plan.append(gw_entry)
            continue

        # === DYNAMIC DUO (Allsvenskan — captain ×3, vice captain ×2) ===
        if chip == "dynamic_duo":
            xi, bench = solve_best_xi_for_gw(current_squad, xpts_map, gw)
            gw_entry["xi"] = xi
            gw_entry["bench"] = bench
            cap = find_best_captain(xi, xpts_map, gw) if xi is not None else None
            cap_id = None
            if cap is not None:
                cap_id = cap.get("id", 0) if isinstance(cap, dict) else getattr(cap, "id", 0)
            vice = find_best_vice_captain(xi, xpts_map, gw, captain_id=cap_id) if xi is not None else None
            gw_entry["captain"] = cap
            gw_entry["vice_captain"] = vice
            gw_entry["captain_multiplier"] = 3  # captain ×3
            gw_entry["dynamic_duo"] = True
            dd_total = 0
            if xi is not None and len(xi) > 0:
                dd_total = xi["id"].map(lambda pid: xpts_map.get(pid, {}).get(gw, 0)).sum()
                # Captain gets ×3 → +2 extra (already counted once in XI sum)
                if cap is not None:
                    cap_pts = xpts_map.get(cap_id, {}).get(gw, 0)
                    dd_total += cap_pts * 2  # ×3 total = base + 2 extra
                # Vice captain gets ×2 → +1 extra
                if vice is not None:
                    vice_id = vice.get("id", 0) if isinstance(vice, dict) else getattr(vice, "id", 0)
                    vice_pts = xpts_map.get(vice_id, {}).get(gw, 0)
                    dd_total += vice_pts * 1  # ×2 total = base + 1 extra
            gw_entry["total_xpts"] = round(dd_total, 1)
            gw_entry["squad"] = current_squad.copy()
            plan.append(gw_entry)
            continue

        # === FRIKORT (Allsvenskan — permanent unlimited transfers, like Wildcard) ===
        if chip == "frikort":
            total_val = int(current_bank + current_squad["now_cost"].sum())
            fk_pool = all_players_df[~all_players_df["id"].isin(banned_ids)] if banned_ids else all_players_df
            fk_squad = solve_wildcard_squad(fk_pool, xpts_map, gw, n_gws - i, total_val,
                                            team_fixture_counts=team_fixture_counts,
                                            locked_ids=locked_ids)
            if fk_squad is not None:
                current_squad = fk_squad
                gw_entry["squad"] = fk_squad
                for _, p in fk_squad.iterrows():
                    current_purchase[p["id"]] = p["now_cost"]
                current_bank = total_val - fk_squad["now_cost"].sum()
            current_ft = 1
            xi, bench = solve_best_xi_for_gw(current_squad, xpts_map, gw)
            gw_entry["xi"] = xi
            gw_entry["bench"] = bench
            gw_entry["captain"] = find_best_captain(xi, xpts_map, gw) if xi is not None else None
            fk_total = 0
            if xi is not None and len(xi) > 0:
                fk_total = xi["id"].map(lambda pid: xpts_map.get(pid, {}).get(gw, 0)).sum()
                cap = gw_entry.get("captain")
                if cap is not None:
                    cap_id = cap.get("id", 0) if isinstance(cap, dict) else getattr(cap, "id", 0)
                    fk_total += xpts_map.get(cap_id, {}).get(gw, 0)
            gw_entry["total_xpts"] = round(fk_total, 1)
            plan.append(gw_entry)
            continue

        # === WILDCARD ===
        if chip == "wildcard":
            total_val = int(current_bank + current_squad["now_cost"].sum())
            wc_pool = all_players_df[~all_players_df["id"].isin(banned_ids)] if banned_ids else all_players_df
            wc_squad = solve_wildcard_squad(wc_pool, xpts_map, gw, n_gws - i, total_val,
                                               team_fixture_counts=team_fixture_counts,
                                               locked_ids=locked_ids)
            if wc_squad is not None:
                current_squad = wc_squad
                gw_entry["squad"] = wc_squad
                for _, p in wc_squad.iterrows():
                    current_purchase[p["id"]] = p["now_cost"]
                current_bank = total_val - wc_squad["now_cost"].sum()
            current_ft = 1
            xi, bench = solve_best_xi_for_gw(current_squad, xpts_map, gw)
            gw_entry["xi"] = xi
            gw_entry["bench"] = bench
            gw_entry["captain"] = find_best_captain(xi, xpts_map, gw) if xi is not None else None
            # Calculate total xPts
            wc_total = 0
            if xi is not None and len(xi) > 0:
                wc_total = xi["id"].map(lambda pid: xpts_map.get(pid, {}).get(gw, 0)).sum()
                cap = gw_entry.get("captain")
                cap_mult = gw_entry.get("captain_multiplier", 2)
                if cap is not None:
                    cap_id = cap.get("id", 0) if isinstance(cap, dict) else getattr(cap, "id", 0)
                    cap_pts = xpts_map.get(cap_id, {}).get(gw, 0)
                    wc_total += cap_pts * (cap_mult - 1)  # only the EXTRA bonus
            gw_entry["total_xpts"] = round(wc_total, 1)
            plan.append(gw_entry)
            continue

        # === NORMAL / TC / BB ===
        transfers_made = []
        transfers_ft_used = 0
        total_hit = 0
        recently_sold = set()

        # The horizon shrinks as we move through GWs
        horizon_end = planning_gw_id + n_gws

        # Filter out banned players from the candidate pool
        transfer_pool = all_players_df[~all_players_df["id"].isin(banned_ids)] if banned_ids else all_players_df

        # Locked players can't be sold — add them to exclude set
        transfer_exclude = set(locked_ids)

        # === FORCE LOCKED PLAYERS IN ===
        # If locked players aren't in the current squad, buy them first
        # by selling the worst player in the same position
        squad_ids = set(current_squad["id"].tolist())
        locked_to_buy = locked_ids - squad_ids
        if locked_to_buy and i == 0:  # only force buys in the first GW of the plan
            for lock_pid in locked_to_buy:
                lock_player = all_players_df[all_players_df["id"] == lock_pid]
                if len(lock_player) == 0:
                    continue
                lock_p = lock_player.iloc[0]
                lock_pos = lock_p["pos_id"]
                lock_cost = lock_p["now_cost"]

                # Find the worst player in the same position to sell (not locked)
                squad_same_pos = current_squad[
                    (current_squad["pos_id"] == lock_pos) &
                    (~current_squad["id"].isin(locked_ids))
                ].copy()
                if len(squad_same_pos) == 0:
                    continue

                # Add horizon xPts for comparison
                squad_same_pos["xpts_h"] = squad_same_pos["id"].map(
                    lambda pid: sum(xpts_map.get(pid, {}).get(g, 0) for g in range(gw, horizon_end))
                )
                worst = squad_same_pos.loc[squad_same_pos["xpts_h"].idxmin()]
                sell_price = calculate_selling_price(
                    worst["id"], worst["now_cost"],
                    current_purchase, current_selling
                )

                # Check if we can afford it
                if current_bank + sell_price >= lock_cost:
                    # Make the transfer
                    transfers_made.append({
                        "out": worst.to_dict(),
                        "in": lock_p.to_dict(),
                        "xpts_gain": 0,  # forced transfer
                        "xpts_gw_gain": 0,
                        "new_bank": int(current_bank + sell_price - lock_cost),
                    })
                    recently_sold.add(worst["id"])
                    current_squad = current_squad[current_squad["id"] != worst["id"]]
                    current_squad = pd.concat([current_squad, lock_player.iloc[:1]], ignore_index=True)
                    current_bank = int(current_bank + sell_price - lock_cost)
                    current_purchase[lock_pid] = lock_cost
                    if worst["id"] in current_selling:
                        del current_selling[worst["id"]]
                    transfers_ft_used += 1

        # Keep finding improving transfers until no more gains
        # Cap at FTs + 2 hits max (so worst case is -8, never -12 or more)
        remaining_ft = max(current_ft - transfers_ft_used, 0)
        max_transfers = min(remaining_ft + 2, 7)

        # === VALUE OF ROLLING A FREE TRANSFER ===
        # Banking a FT gives you flexibility next week.
        # Estimated value of an extra FT = ~1.5-2.5 xPts (the average gain from
        # the best available transfer next week). This means we should only
        # USE a free transfer if the gain exceeds this threshold.
        # The value diminishes as we approach max FTs (5) since we'd waste the roll.
        if current_ft >= 5:
            roll_value = 0.0  # already at max, must use or lose
        elif current_ft >= 3:
            roll_value = 1.0  # already have good flexibility, lower bar
        else:
            roll_value = 1.5  # banking from 1→2 or 2→3 is very valuable

        for t_num in range(max_transfers):
            transfer = find_best_single_transfer_for_gw(
                current_squad, transfer_pool, current_bank,
                current_purchase, current_selling, xpts_map, gw,
                exclude_ids=recently_sold | transfer_exclude,
                horizon_end=horizon_end,
            )

            if transfer is None:
                break  # no improving transfer found

            # Is this a free transfer or a hit?
            is_free = (t_num < current_ft)
            hit_number = t_num - current_ft + 1 if not is_free else 0

            if is_free:
                # Only use the FT if the gain exceeds the value of rolling
                # First FT: compare against roll_value
                # Second+ FT: lower threshold (we've already decided to use at least one)
                if t_num == 0:
                    if transfer["xpts_gain"] < roll_value:
                        break  # better to roll the FT
                else:
                    if transfer["xpts_gain"] < 0.3:
                        break  # marginal gain, stop
            else:
                # Each hit costs exactly 4 points and the xpts_gain is over the
                # full horizon, so a flat 4-pt break-even is the correct threshold.
                # Escalating thresholds artificially refused profitable double hits.
                if transfer["xpts_gain"] < 4.0:
                    break
                total_hit += 4

            # Apply transfer
            transfers_made.append(transfer)
            out_id = transfer["out"]["id"]
            in_id = transfer["in"]["id"]
            recently_sold.add(out_id)  # don't suggest buying back
            current_squad = current_squad[current_squad["id"] != out_id]
            in_player = all_players_df[all_players_df["id"] == in_id]
            if len(in_player) > 0:
                current_squad = pd.concat([current_squad, in_player.iloc[:1]], ignore_index=True)
            current_bank = transfer["new_bank"]
            current_purchase[in_id] = transfer["in"]["now_cost"]
            if out_id in current_selling:
                del current_selling[out_id]
            if is_free:
                transfers_ft_used += 1

        # Store all transfers for this GW
        gw_entry["transfers"] = transfers_made
        gw_entry["hit"] = total_hit
        gw_entry["ft_used"] = transfers_ft_used

        # FT accounting: spent some FTs, then gain 1 for next GW
        remaining_ft = current_ft - transfers_ft_used
        current_ft = min(remaining_ft + 1, 5)

        xi, bench = solve_best_xi_for_gw(current_squad, xpts_map, gw)
        gw_entry["xi"] = xi
        gw_entry["bench"] = bench
        gw_entry["squad"] = current_squad.copy()
        gw_entry["captain"] = find_best_captain(xi, xpts_map, gw) if xi is not None else None

        if chip == "triple_captain":
            gw_entry["captain_multiplier"] = 3
        if chip == "bench_boost":
            gw_entry["bench_boost"] = True

        # Calculate total xPts for this GW (used by chip strategy optimiser)
        gw_total = 0
        if xi is not None and len(xi) > 0:
            xi_xpts = xi["id"].map(lambda pid: xpts_map.get(pid, {}).get(gw, 0))
            gw_total = xi_xpts.sum()
            # Captain bonus
            cap = gw_entry.get("captain")
            cap_mult = gw_entry.get("captain_multiplier", 2)
            if cap is not None:
                cap_id = cap.get("id", 0) if isinstance(cap, dict) else getattr(cap, "id", 0)
                cap_pts = xpts_map.get(cap_id, {}).get(gw, 0)
                gw_total += cap_pts * (cap_mult - 1)
            # Bench boost
            if gw_entry.get("bench_boost") and bench is not None and len(bench) > 0:
                bench_xpts = bench["id"].map(lambda pid: xpts_map.get(pid, {}).get(gw, 0))
                gw_total += bench_xpts.sum()
        # Subtract hits
        gw_total -= gw_entry.get("hit", 0)
        gw_entry["total_xpts"] = round(gw_total, 1)

        plan.append(gw_entry)

    return plan


def get_formation_str(xi_df):
    """Get formation string like '4-4-2' from starting XI."""
    if xi_df is None:
        return "-"
    d = len(xi_df[xi_df["pos_id"] == 2])
    m = len(xi_df[xi_df["pos_id"] == 3])
    f = len(xi_df[xi_df["pos_id"] == 4])
    return f"{d}-{m}-{f}"


def render_fdr(upcoming, teams):
    """Render fixture difficulty badges."""
    badges = []
    for f in upcoming[:5]:
        opp = teams.get(f["opp"], {}).get("short_name", "???")
        pre = "" if f["home"] else "@"
        d = f.get("difficulty", 3)
        badges.append(f'<span class="fdr-{d}">{pre}{opp}</span>')
    return " ".join(badges) if badges else "-"


def render_landing_block(current_gw, df, team_data):
    """Render the 'What's new for you' block above the tabs.

    Three tiles: deadline countdown (server-rendered, refreshes each rerun),
    personalized welcome / last-GW score, and a recommended captain. When no
    team is loaded, the second tile becomes a CTA and the third falls back
    to the global top pick.

    Note: HTML is built as flat single-line strings — Streamlit's markdown
    renderer treats any line indented 4+ spaces as a code block.
    """
    # --- Tile 1: deadline countdown (server-computed) ---
    countdown_text = "—"
    countdown_class = ""
    deadline_str = "—"
    gw_label = "GW —"
    if current_gw and current_gw.get("deadline_time"):
        try:
            deadline = datetime.fromisoformat(current_gw["deadline_time"].replace("Z", "+00:00"))
            deadline_str = deadline.strftime("%a %d %b · %H:%M")
            gw_label = f"GW{current_gw.get('id', '?')}"
            now = datetime.now(deadline.tzinfo)
            diff = (deadline - now).total_seconds()
            if current_gw.get("finished") or diff <= 0:
                countdown_text = "Live"
                countdown_class = "lt-live"
            else:
                d = int(diff // 86400)
                h = int((diff % 86400) // 3600)
                m = int((diff % 3600) // 60)
                if d > 0:
                    countdown_text = f"{d}d {h:02d}h {m:02d}m"
                else:
                    s = int(diff % 60)
                    countdown_text = f"{h:02d}h {m:02d}m {s:02d}s"
                if d == 0 and h < 6:
                    countdown_class = "lt-urgent"
        except Exception:
            pass

    tile_deadline = (
        f'<div class="landing-tile lt-deadline">'
        f'<div class="lt-icon">⏱</div>'
        f'<div class="lt-label">{gw_label} Deadline</div>'
        f'<div class="lt-value lt-countdown {countdown_class}">{countdown_text}</div>'
        f'<div class="lt-sub">{deadline_str}</div>'
        f'</div>'
    )

    # --- Tile 2: welcome / last-GW score ---
    if team_data:
        team_name = str(team_data.get("team_name", "Your team"))
        last_pts = team_data.get("last_gw_points")
        last_rank = team_data.get("last_gw_rank")
        overall_rank = team_data.get("overall_rank")
        if last_pts is not None:
            big = f"{last_pts} pts"
            sub = (
                f"Last GW · GW rank {last_rank:,}"
                if isinstance(last_rank, int)
                else "Last GW result"
            )
        else:
            big = "Welcome back"
            sub = (
                f"Overall rank {overall_rank:,}"
                if isinstance(overall_rank, int)
                else "Squad loaded"
            )
        tile_welcome = (
            f'<div class="landing-tile lt-welcome">'
            f'<div class="lt-icon">👋</div>'
            f'<div class="lt-label">{team_name}</div>'
            f'<div class="lt-value lt-welcome-value">{big}</div>'
            f'<div class="lt-sub">{sub}</div>'
            f'</div>'
        )
    else:
        tile_welcome = (
            '<div class="landing-tile lt-welcome lt-cta">'
            '<div class="lt-icon">👋</div>'
            '<div class="lt-label">Make it personal</div>'
            '<div class="lt-value lt-welcome-value">Add your FPL ID</div>'
            '<div class="lt-sub">See your team, captain pick, and transfer ideas</div>'
            '</div>'
        )

    # --- Tile 3: recommended captain ---
    captain_pool = None
    captain_label = "Top Captain Pick"
    if team_data and df is not None and "squad_ids" in team_data:
        squad = df[df["id"].isin(team_data["squad_ids"])]
        if len(squad) > 0:
            captain_pool = squad
            captain_label = "Recommended Captain"
    if captain_pool is None and df is not None:
        captain_pool = df[df["minutes"] > 0] if "minutes" in df.columns else df

    cap_name, cap_team, cap_xpts = "—", "", None
    if captain_pool is not None and len(captain_pool) > 0 and "xpts_next_gw" in captain_pool.columns:
        top = captain_pool.sort_values("xpts_next_gw", ascending=False).iloc[0]
        cap_name = str(top.get("name", "—"))
        cap_team = str(top.get("team", ""))
        cap_xpts = float(top.get("xpts_next_gw", 0) or 0)

    cap_sub = f"{cap_xpts:.1f} xPts · {cap_team}" if cap_xpts is not None else "—"
    tile_captain = (
        f'<div class="landing-tile lt-captain">'
        f'<div class="lt-icon">©</div>'
        f'<div class="lt-label">{captain_label}</div>'
        f'<div class="lt-value lt-captain-name">{cap_name}</div>'
        f'<div class="lt-sub">{cap_sub}</div>'
        f'</div>'
    )

    block = f'<div class="landing-grid">{tile_deadline}{tile_welcome}{tile_captain}</div>'
    st.markdown(block, unsafe_allow_html=True)


# ============================================================
# MAIN APP
# ============================================================

def main():
    # Header with logo — centered
    st.markdown(
        f'<div style="text-align:center; margin-bottom:0.5rem;">{DATUMLY_LOGO_SVG}</div>',
        unsafe_allow_html=True,
    )

    # === League Switcher ===
    col_league, col_refresh, col_spacer = st.columns([1.5, 1, 4.5])
    with col_league:
        league_options = list(LEAGUE_CONFIGS.keys())
        league_labels = {k: v["name"] for k, v in LEAGUE_CONFIGS.items()}
        current_league = st.session_state.get("active_league", "FPL")
        selected_league = st.selectbox(
            "🏆 League",
            options=league_options,
            format_func=lambda x: league_labels[x],
            index=league_options.index(current_league),
            key="league_selector",
            label_visibility="collapsed",
        )
        if selected_league != current_league:
            st.session_state["active_league"] = selected_league
            st.cache_data.clear()
            st.rerun()

    with col_refresh:
        if st.button("🔄 Refresh Data"):
            st.cache_data.clear()
            st.rerun()

    # Get active league config
    LC = get_league_config()
    league_name = LC["name"]

    # === Load data ===
    fetch_time = datetime.now()

    with st.spinner(f"Loading {league_name} data..."):
        bootstrap, fixtures_raw, fpl_err = load_fpl_data()

    if fpl_err or bootstrap is None:
        st.error(f"Failed to load {league_name} data: {fpl_err}")
        if st.button("🔄 Retry"):
            st.cache_data.clear()
            st.rerun()
        return

    with st.spinner("Loading betting odds from football-data.co.uk..."):
        odds_df, odds_err = load_betting_odds()

    odds_status = "✅ Loaded" if odds_df is not None else f"⚠️ {odds_err or 'Unavailable'}"
    team_odds = odds_to_probabilities(odds_df, TEAM_NAME_MAP) if odds_df is not None else {}

    # Check Elo status (loaded inside enrich_data but we check here for display)
    elo_check, elo_check_err = load_club_elo()
    elo_status = f"✅ {len(elo_check)} teams" if elo_check else f"⚠️ {elo_check_err or 'Unavailable'}"

    # Check live odds status
    live_check, live_check_err = load_live_odds()
    if live_check:
        n_fixtures = len(live_check.get("fixtures", {}))
        remaining = live_check.get("remaining", "?")
        live_status = f"✅ {n_fixtures} fixtures ({remaining} req left)"
    else:
        live_status = f"⚠️ {live_check_err or 'Unavailable'}"

    with st.spinner("Building xPts model & enriching data..."):
        df, teams, current_gw, planning_gw_id, upcoming_map, fixtures_list, xpts_map, team_fixture_counts, xpts_breakdown = enrich_data(
            bootstrap, fixtures_raw, team_odds
        )

    # === GW Info ===
    if current_gw:
        deadline = datetime.fromisoformat(current_gw["deadline_time"].replace("Z", "+00:00"))
        status = "Completed" if current_gw.get("finished") else ("In Progress" if current_gw.get("is_current") else "Upcoming")
        bc = "badge-green" if status == "Completed" else ("badge-yellow" if status == "In Progress" else "badge-blue")
        planning_str = f"Planning for GW{planning_gw_id}" if planning_gw_id != current_gw.get("id") else ""
        st.markdown(f"""<div class="gw-bar">
            <span class="gw-num">Gameweek {current_gw['id']}</span>
            <span class="gw-deadline">Deadline: {deadline.strftime('%a %d %b, %H:%M')}</span>
            <span class="badge {bc}">{status}</span>
            {f'<span class="badge badge-blue">{planning_str}</span>' if planning_str else ''}
            <span style="color:#5a6580; font-size:0.68rem;">
                Odds: {odds_status} · Elo: {elo_status} · Live: {live_status} ·
                {fetch_time.strftime('%d %b %H:%M')} (cached)
            </span>
        </div>""", unsafe_allow_html=True)

    # === Landing block (above tabs): personalized snapshot for returning visitors ===
    # Auto-load team data if fpl_id is already known (from session or URL query param)
    _saved_fpl_id = st.session_state.get("fpl_id", "")
    if not _saved_fpl_id:
        _qp_id = st.query_params.get("fpl_id", "") if hasattr(st, "query_params") else ""
        if _qp_id and str(_qp_id).strip().isdigit():
            st.session_state["fpl_id"] = str(_qp_id).strip()
            _saved_fpl_id = st.session_state["fpl_id"]
    _team_data_for_landing = st.session_state.get("team_data")
    if _saved_fpl_id and not _team_data_for_landing and current_gw:
        try:
            _td, _ = fetch_manager_team(int(_saved_fpl_id), current_gw["id"])
            if _td:
                st.session_state["team_data"] = _td
                _team_data_for_landing = _td
        except Exception:
            _team_data_for_landing = None
    render_landing_block(current_gw, df, _team_data_for_landing)

    # === Tabs ===
    tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
        "🏠 My Team", "📊 Dashboard", "👥 Player Projections",
        "⭐ Optimal Squad (MILP)", "🔄 Transfer Planner", "📅 Fixtures",
        "🔬 Backtest", "🎯 Chip Strategy"
    ])

    if st.session_state.pop("navigate_to_my_team", False):
        components.html("""<script>
        setTimeout(function(){
            var t=window.parent.document.querySelectorAll('.stTabs button[data-baseweb="tab"]');
            if(t&&t.length>0)t[0].click();
        },300);
        </script>""", height=0)

    active = df[df["minutes"] > 0].copy()
    qualified = df[df["minutes"] > 45].copy()

    # ==================== MY TEAM ====================
    with tab1:
        st.markdown(
            '<div class="section-header">🏠 My Team — Enter Your FPL ID</div>',
            unsafe_allow_html=True,
        )

        # FPL ID input
        col_id, col_btn = st.columns([3, 1])
        with col_id:
            fpl_id = st.text_input(
                "FPL Team ID",
                value=st.session_state.get("fpl_id", ""),
                placeholder="e.g. 123456",
                help="Find your ID in the URL when you view your team on the FPL website",
                key="fpl_id_input",
            )
        with col_btn:
            st.markdown("<br>", unsafe_allow_html=True)
            load_team = st.button("Load My Team", use_container_width=True)

        if load_team and fpl_id:
            st.session_state["fpl_id"] = fpl_id
            # Persist in the URL so the returning visit auto-loads
            try:
                st.query_params["fpl_id"] = fpl_id.strip()
            except Exception:
                pass

        if fpl_id and fpl_id.strip().isdigit():
            manager_id = int(fpl_id.strip())
            gw_id = current_gw["id"] if current_gw else 1

            with st.spinner("Fetching your team..."):
                team_data, team_err = fetch_manager_team(manager_id, gw_id)
                if team_data:
                    st.session_state["team_data"] = team_data

            if team_err:
                st.error(f"Could not load team: {team_err}")
                st.info("Make sure your FPL ID is correct. You can find it in the URL when viewing your team on the FPL website.")
            elif team_data:
                # Manager info header
                c1, c2, c3, c4 = st.columns(4)
                with c1:
                    st.markdown(f"""<div class="metric-card">
                        <div class="metric-label">Manager</div>
                        <div class="metric-value" style="font-size:1.1rem;">{team_data['team_name']}</div>
                        <div class="metric-sub">{team_data['manager_name']}</div>
                    </div>""", unsafe_allow_html=True)
                with c2:
                    st.markdown(f"""<div class="metric-card">
                        <div class="metric-label">Overall Rank</div>
                        <div class="metric-value">{team_data['overall_rank']:,}</div>
                        <div class="metric-sub">{team_data['total_points']} pts</div>
                    </div>""", unsafe_allow_html=True)
                with c3:
                    bank_display = team_data['bank'] / 10
                    ft_display = team_data.get('free_transfers', 1)
                    st.markdown(f"""<div class="metric-card">
                        <div class="metric-label">Bank / Free Transfers</div>
                        <div class="metric-value">£{bank_display:.1f}m</div>
                        <div class="metric-sub">~{ft_display} FT available</div>
                    </div>""", unsafe_allow_html=True)
                with c4:
                    squad_xpts = 0
                    my_squad = df[df["id"].isin(team_data["squad_ids"])].copy()
                    if len(my_squad) > 0:
                        squad_xpts = my_squad["xpts_total"].sum()
                    st.markdown(f"""<div class="metric-card">
                        <div class="metric-label">Squad xPts (6GW)</div>
                        <div class="metric-value">{squad_xpts:.1f}</div>
                        <div class="metric-sub">{len(my_squad)} players loaded</div>
                    </div>""", unsafe_allow_html=True)

                # === CHIPS REMAINING (lightweight display) ===
                chips_rem = team_data.get("chips_remaining", {})
                has_any_chips = any(v > 0 for v in chips_rem.values())
                ft_available = team_data.get("free_transfers", 1)

                if has_any_chips:
                    chip_labels_brief = {
                        "wildcard": "🃏 Wildcard", "freehit": "⚡ Free Hit",
                        "3xc": "👑 Triple Captain", "bboost": "💪 Bench Boost"
                    }
                    remaining_str = " · ".join([
                        chip_labels_brief.get(k, k) for k, v in chips_rem.items() if v > 0
                    ])
                    st.markdown(
                        f'<div style="background:#1a1e2e;border-radius:8px;padding:0.8rem;margin-bottom:0.8rem;">'
                        f'<span style="color:#f02d6e;font-weight:700;">🎯 Chips Remaining:</span> '
                        f'<span style="color:#8892a8;">{remaining_str}</span> · '
                        f'<span style="color:#5a6580;font-size:0.75rem;">See 🎯 Chip Strategy tab for optimal timing</span></div>',
                        unsafe_allow_html=True,
                    )

                st.markdown("")

                if len(my_squad) > 0:
                    # Mark starters vs bench
                    my_squad["is_starter"] = my_squad["id"].map(
                        lambda pid: team_data["positions"].get(pid, 99) <= 11
                    )
                    my_squad["is_captain"] = my_squad["id"].map(
                        lambda pid: team_data["captains"].get(pid, False)
                    )
                    my_squad["is_vice"] = my_squad["id"].map(
                        lambda pid: team_data["vice_captains"].get(pid, False)
                    )

                    # Show current squad
                    st.subheader("Your Current Squad")
                    starters = my_squad[my_squad["is_starter"]].sort_values(
                        ["pos_id", "xpts_total"], ascending=[True, False]
                    )
                    bench = my_squad[~my_squad["is_starter"]].sort_values("pos_id")

                    # Pitch view of starters
                    for pid_val, plabel in [(4, "Forwards"), (3, "Midfielders"), (2, "Defenders"), (1, "Goalkeeper")]:
                        pp = starters[starters["pos_id"] == pid_val]
                        if len(pp) > 0:
                            st.markdown(f"<div class='pitch-row-label'>{plabel}</div>", unsafe_allow_html=True)
                            cols = st.columns(max(len(pp), 1))
                            for i, (_, p) in enumerate(pp.iterrows()):
                                is_gk = (p["pos_id"] == 1)
                                is_cap = p.get("is_captain", False)
                                cap_badge = " (C)" if is_cap else (" (V)" if p.get("is_vice", False) else "")
                                shirt_svg = make_shirt_svg(p["team"], f"{p['xpts_next_gw']:.1f}", is_gk=is_gk, is_captain=is_cap, player_name=p.get("name", ""), team_code=int(p.get("team_code", 0) or 0))
                                with cols[i]:
                                    html = f'<div style="text-align:center;">{shirt_svg}<div class="pitch-name">{p["name"]}{cap_badge}</div><div class="pitch-price">£{p["price"]:.1f}m · {p["xpts_total"]:.1f} xPts</div></div>'
                                    st.markdown(html, unsafe_allow_html=True)

                    if len(bench) > 0:
                        st.markdown("**Bench**")
                        bcols = st.columns(max(len(bench), 1))
                        for i, (_, p) in enumerate(bench.iterrows()):
                            is_gk = (p["pos_id"] == 1)
                            shirt_svg = make_shirt_svg(p["team"], f"{p['xpts_next_gw']:.1f}", is_gk=is_gk, width=44, height=44, player_name=p.get("name", ""), team_code=int(p.get("team_code", 0) or 0))
                            with bcols[i]:
                                html = f'<div style="text-align:center;opacity:0.6;">{shirt_svg}<div class="pitch-name">{p["name"]}</div><div class="pitch-price">{p["pos"]} · £{p["price"]:.1f}m</div></div>'
                                st.markdown(html, unsafe_allow_html=True)

                    st.markdown("")

                    # === CHIP & TRANSFER PLANNER ===
                    st.markdown(
                        '<div class="section-header">🗓️ Gameweek-by-Gameweek Planner '
                        '<span class="source-tag src-model">Rolling Planner</span></div>',
                        unsafe_allow_html=True,
                    )

                    ft_available = team_data.get("free_transfers", 1)
                    st.caption(f"You have **{ft_available} free transfer(s)** available.")

                    # Step 1: Chip selection
                    st.markdown("**Step 1: Set your chip schedule**")
                    gw_options = [planning_gw_id + i for i in range(6)]

                    # Check if a chip strategy was applied from the Chip Strategy tab
                    applied = st.session_state.pop("applied_chip_schedule", None)
                    if applied:
                        # Build reverse map: planner_chip_name → gw
                        planner_to_key = {
                            "wildcard": "wc_gw", "free_hit": "fh_gw",
                            "triple_captain": "tc_gw", "bench_boost": "bb_gw",
                        }
                        # Reset all chip keys first
                        for sk in ["wc_gw", "fh_gw", "tc_gw", "bb_gw"]:
                            if sk in st.session_state:
                                del st.session_state[sk]
                        # Set the applied chips
                        for gw_a, cname in applied.items():
                            key = planner_to_key.get(cname)
                            if key:
                                st.session_state[key] = gw_a
                        st.success("✅ Chip strategy applied from the 🎯 Chip Strategy tab!")

                    chip_cols = st.columns(4)
                    with chip_cols[0]:
                        wc_gw = st.selectbox("🃏 Wildcard", ["None"] + gw_options, key="wc_gw")
                    with chip_cols[1]:
                        fh_gw = st.selectbox("⚡ Free Hit", ["None"] + gw_options, key="fh_gw")
                    with chip_cols[2]:
                        tc_gw = st.selectbox("👑 Triple Captain", ["None"] + gw_options, key="tc_gw")
                    with chip_cols[3]:
                        bb_gw = st.selectbox("💪 Bench Boost", ["None"] + gw_options, key="bb_gw")

                    chip_schedule = {}
                    if wc_gw != "None":
                        chip_schedule[int(wc_gw)] = "wildcard"
                    if fh_gw != "None":
                        chip_schedule[int(fh_gw)] = "free_hit"
                    if tc_gw != "None":
                        chip_schedule[int(tc_gw)] = "triple_captain"
                    if bb_gw != "None":
                        chip_schedule[int(bb_gw)] = "bench_boost"

                    # Validate
                    chip_gws_used = [g for g in [wc_gw, fh_gw, tc_gw, bb_gw] if g != "None"]
                    if len(chip_gws_used) != len(set(chip_gws_used)):
                        st.error("You can only play one chip per gameweek.")

                    # Show chip summary
                    if chip_schedule:
                        chip_labels = {"wildcard": "🃏 Wildcard", "free_hit": "⚡ Free Hit",
                                       "triple_captain": "👑 Triple Captain", "bench_boost": "💪 Bench Boost"}
                        chip_summary = " · ".join([
                            f"GW{gw}: {chip_labels.get(c, c)}" for gw, c in sorted(chip_schedule.items())
                        ])
                        st.info(f"Chip plan: {chip_summary}")
                    else:
                        st.markdown(
                            "<span style='color:#5a6580;font-size:0.8rem;'>No chips selected — planning with normal transfers only</span>",
                            unsafe_allow_html=True,
                        )

                    st.markdown("")

                    # Step 2: Lock & Ban players
                    st.markdown("**Step 2: Lock & ban players**")
                    # Build player options from the full player pool
                    all_opts = df[df["minutes"] > 0].sort_values("xpts_total", ascending=False)
                    planner_labels = {
                        row["id"]: f"{row['name']} ({row['team']}, {row['pos']}, £{row['price']:.1f}m)"
                        for _, row in all_opts.iterrows()
                    }
                    # Separate current squad players for the lock dropdown
                    squad_ids = set(my_squad["id"].tolist())
                    squad_labels = {pid: planner_labels[pid] for pid in squad_ids if pid in planner_labels}
                    non_squad_labels = {pid: planner_labels[pid] for pid in planner_labels if pid not in squad_ids}

                    lock_col, ban_col = st.columns(2)
                    with lock_col:
                        planner_locked = st.multiselect(
                            "🔒 Lock (must be in squad)",
                            options=list(planner_labels.keys()),
                            format_func=lambda pid: planner_labels.get(pid, str(pid)),
                            placeholder="Players to always include...",
                            key="planner_lock",
                        )
                    with ban_col:
                        planner_banned = st.multiselect(
                            "🚫 Ban (don't buy these)",
                            options=list(non_squad_labels.keys()),
                            format_func=lambda pid: non_squad_labels.get(pid, str(pid)),
                            placeholder="Players to avoid...",
                            key="planner_ban",
                        )

                    planner_locked_ids = set(planner_locked)
                    planner_banned_ids = set(planner_banned)

                    _settings_fp = json.dumps({
                        "chip_schedule": chip_schedule,
                        "locked": sorted(planner_locked_ids),
                        "banned": sorted(planner_banned_ids),
                    }, sort_keys=True, default=str)

                    st.markdown("")

                    # Step 3: Expected blanks & doubles
                    st.markdown("**Step 3: Expected blank & double gameweeks**")

                    # === MISSING FIXTURE DETECTOR ===
                    # Every team plays every other team twice (home & away) = 38 matches
                    # Build the full expected fixture list and find what's missing

                    team_id_to_short = {t_id: t.get("short_name", "?") for t_id, t in teams.items()}
                    all_team_ids = sorted(teams.keys())

                    # Build set of all assigned fixtures: (home_id, away_id)
                    assigned_fixtures = set()
                    team_fixture_total = {}
                    team_fixture_per_gw = {}
                    fixture_gw_map = {}  # (home_id, away_id) -> gw

                    for t_id in all_team_ids:
                        team_fixture_total[t_id] = 0
                        team_fixture_per_gw[t_id] = {}

                    for f in fixtures_list:
                        ev = f.get("event")
                        h, a = f["team_h"], f["team_a"]
                        if ev:
                            assigned_fixtures.add((h, a))
                            fixture_gw_map[(h, a)] = ev
                            for t_id in [h, a]:
                                team_fixture_total[t_id] = team_fixture_total.get(t_id, 0) + 1
                                if t_id not in team_fixture_per_gw:
                                    team_fixture_per_gw[t_id] = {}
                                team_fixture_per_gw[t_id][ev] = team_fixture_per_gw[t_id].get(ev, 0) + 1
                        else:
                            # Fixture exists but has no GW assigned = postponed
                            assigned_fixtures.add((h, a))

                    # Find missing fixtures (not even in the API as unassigned)
                    # Each pair should have exactly 2 fixtures: (A,B) and (B,A)
                    missing_fixtures = []
                    for i, t1 in enumerate(all_team_ids):
                        for t2 in all_team_ids:
                            if t1 == t2:
                                continue
                            if (t1, t2) not in assigned_fixtures:
                                t1_short = team_id_to_short.get(t1, "?")
                                t2_short = team_id_to_short.get(t2, "?")
                                missing_fixtures.append({
                                    "home": t1_short, "away": t2_short,
                                    "home_id": t1, "away_id": t2,
                                })

                    # Also find fixtures with event=None (postponed but in the API)
                    postponed = []
                    for f in fixtures_list:
                        if f.get("event") is None:
                            h_short = team_id_to_short.get(f["team_h"], "?")
                            a_short = team_id_to_short.get(f["team_a"], "?")
                            postponed.append({"home": h_short, "away": a_short})

                    # Fixtures owed per team
                    fixtures_owed = {}
                    for t_id, count in team_fixture_total.items():
                        owed = LC["season_gws"] - count
                        if owed > 0:
                            fixtures_owed[team_id_to_short.get(t_id, "?")] = owed

                    # === FREE MIDWEEK FINDER ===
                    # Look at GW dates to find gaps where midweek fixtures could slot in
                    events_all = bootstrap.get("events", [])
                    gw_deadlines = {}
                    for ev in events_all:
                        if ev.get("id") and ev.get("deadline_time"):
                            try:
                                dt = datetime.fromisoformat(ev["deadline_time"].replace("Z", "+00:00"))
                                gw_deadlines[ev["id"]] = dt
                            except Exception:
                                pass

                    free_midweeks = []
                    for gw_num in range(planning_gw_id, LC["season_gws"]):
                        dt_this = gw_deadlines.get(gw_num)
                        dt_next = gw_deadlines.get(gw_num + 1)
                        if dt_this and dt_next:
                            gap_days = (dt_next - dt_this).days
                            # A gap of 7+ days between GW deadlines means there's a free midweek
                            if gap_days >= 7:
                                midweek_date = dt_this + timedelta(days=3)
                                free_midweeks.append({
                                    "after_gw": gw_num,
                                    "gap_days": gap_days,
                                    "likely_dgw": gw_num + 1,  # postponed match goes into next GW
                                    "midweek_approx": midweek_date.strftime("%d %b"),
                                })

                    # === DISPLAY ===
                    if missing_fixtures or postponed or fixtures_owed:
                        st.markdown(
                            '<div style="background:#1a2e1a;border-radius:8px;padding:0.8rem;margin-bottom:0.5rem;">',
                            unsafe_allow_html=True,
                        )

                        if fixtures_owed:
                            owed_str = ", ".join([f"**{t}** ({n})" for t, n in sorted(fixtures_owed.items(), key=lambda x: -x[1])])
                            st.markdown(
                                f'<span style="color:#34d399;font-weight:600;">📊 Fixtures owed:</span> '
                                f'<span style="color:#8892a8;font-size:0.82rem;">{owed_str}</span>',
                                unsafe_allow_html=True,
                            )

                        if missing_fixtures or postponed:
                            all_unscheduled = postponed + missing_fixtures
                            match_strs = [f"{m['home']} vs {m['away']}" for m in all_unscheduled[:10]]
                            st.markdown(
                                f'<span style="color:#34d399;font-weight:600;">⚽ Unscheduled matches:</span> '
                                f'<span style="color:#8892a8;font-size:0.82rem;">{" · ".join(match_strs)}'
                                f'{"..." if len(all_unscheduled) > 10 else ""}</span>',
                                unsafe_allow_html=True,
                            )

                        if free_midweeks:
                            mw_strs = [f"After GW{m['after_gw']} ({m['gap_days']}d gap → likely DGW in GW{m['likely_dgw']}, ~{m['midweek_approx']})"
                                       for m in free_midweeks]
                            st.markdown(
                                f'<span style="color:#34d399;font-weight:600;">📅 Free midweeks:</span> '
                                f'<span style="color:#8892a8;font-size:0.82rem;">{"<br>".join(mw_strs)}</span>',
                                unsafe_allow_html=True,
                            )

                        # Suggest likely DGW mapping
                        if fixtures_owed and free_midweeks:
                            st.markdown(
                                '<span style="color:#fbbf24;font-weight:600;font-size:0.82rem;">'
                                '💡 Suggestion: Unscheduled matches will likely be placed in GWs with free midweeks. '
                                'Flag those teams as doubling in the inputs below.</span>',
                                unsafe_allow_html=True,
                            )

                        st.markdown('</div>', unsafe_allow_html=True)
                    else:
                        st.markdown(
                            '<span style="color:#5a6580;font-size:0.8rem;">No missing fixtures detected — all 380 matches are assigned.</span>',
                            unsafe_allow_html=True,
                        )

                    st.caption("Flag teams blanking in specific GWs. For expected DGWs, flag the teams "
                                "that will double — their xPts will be boosted ×1.7 for that GW.")

                    # Build team list
                    all_team_shorts = sorted(df["team"].unique().tolist())

                    blank_overrides = {}  # {gw: set of team_short_names}
                    dgw_overrides = {}    # {gw: set of team_short_names}
                    gw_options_plan = [planning_gw_id + i for i in range(6)]

                    # Show compact inputs for blanks AND doubles per GW
                    for gw_b in gw_options_plan:
                        bcol, dcol = st.columns(2)
                        with bcol:
                            blanking = st.multiselect(
                                f"⚠️ GW{gw_b} blanks",
                                options=all_team_shorts,
                                placeholder="Teams blanking...",
                                key=f"blank_gw_{gw_b}",
                            )
                            if blanking:
                                blank_overrides[gw_b] = set(blanking)
                        with dcol:
                            doubling = st.multiselect(
                                f"🔥 GW{gw_b} doubles",
                                options=all_team_shorts,
                                placeholder="Teams doubling...",
                                key=f"dgw_gw_{gw_b}",
                            )
                            if doubling:
                                dgw_overrides[gw_b] = set(doubling)

                    # Build modified xpts_map that zeros blanks and boosts doubles
                    DGW_BOOST = 1.7  # DGW players get ~1.7x xPts (two fixtures)
                    xpts_map_adjusted = {}
                    player_teams = dict(zip(df["id"], df["team"]))
                    for pid, gw_dict in xpts_map.items():
                        player_team = player_teams.get(pid, "")
                        adjusted = {}
                        for gw, xpts_val in gw_dict.items():
                            blanking_teams = blank_overrides.get(gw, set())
                            doubling_teams = dgw_overrides.get(gw, set())
                            if player_team in blanking_teams:
                                adjusted[gw] = 0.0
                            elif player_team in doubling_teams:
                                adjusted[gw] = round(xpts_val * DGW_BOOST, 2)
                            else:
                                adjusted[gw] = xpts_val
                        xpts_map_adjusted[pid] = adjusted

                    # Show summary of overrides
                    override_parts = []
                    if blank_overrides:
                        for gw_o in sorted(blank_overrides.keys()):
                            teams_str = ", ".join(sorted(blank_overrides[gw_o]))
                            override_parts.append(f"GW{gw_o} blank: {teams_str}")
                    if dgw_overrides:
                        for gw_o in sorted(dgw_overrides.keys()):
                            teams_str = ", ".join(sorted(dgw_overrides[gw_o]))
                            override_parts.append(f"GW{gw_o} DGW: {teams_str}")
                    if override_parts:
                        st.info(" · ".join(override_parts))

                    st.markdown("")

                    # Step 4: Generate plan
                    st.markdown("**Step 4: Generate your optimal plan**")
                    generate = st.button("🚀 Generate 6-Gameweek Plan", use_container_width=True, type="primary")

                    # Auto-run if triggered from Chip Strategy tab's Apply button
                    auto_run = st.session_state.pop("plan_auto_run", False)
                    if auto_run:
                        generate = True

                    if generate or st.session_state.get("plan_generated"):
                        st.session_state["plan_generated"] = True
                        plan = st.session_state.get("last_plan")

                        if generate:
                            loading_placeholder = st.empty()

                            # Build pixel-art DATUMLY with pure CSS (no JavaScript)
                            _px_letters = {
                                'D': [1,1,1,0,0,1,0,0,1,0,1,0,0,0,1,1,0,0,0,1,1,0,0,0,1,1,0,0,1,0,1,1,1,0,0],
                                'A': [0,1,1,1,0,1,0,0,0,1,1,0,0,0,1,1,1,1,1,1,1,0,0,0,1,1,0,0,0,1,1,0,0,0,1],
                                'T': [1,1,1,1,1,0,0,1,0,0,0,0,1,0,0,0,0,1,0,0,0,0,1,0,0,0,0,1,0,0,0,0,1,0,0],
                                'U': [1,0,0,0,1,1,0,0,0,1,1,0,0,0,1,1,0,0,0,1,1,0,0,0,1,1,0,0,0,1,0,1,1,1,0],
                                'M': [1,0,0,0,1,1,1,0,1,1,1,0,1,0,1,1,0,0,0,1,1,0,0,0,1,1,0,0,0,1,1,0,0,0,1],
                                'L': [1,0,0,0,0,1,0,0,0,0,1,0,0,0,0,1,0,0,0,0,1,0,0,0,0,1,0,0,0,0,1,1,1,1,1],
                                'Y': [1,0,0,0,1,1,0,0,0,1,0,1,0,1,0,0,0,1,0,0,0,0,1,0,0,0,0,1,0,0,0,0,1,0,0],
                            }
                            _px_word = 'DATUMLY'
                            _px_cols = ['#00b46e','#f02d6e','#00b46e','#f02d6e','#00b46e','#f02d6e','#00b46e']
                            _px_html = []
                            _px_idx = 0
                            for _li, _ch in enumerate(_px_word):
                                _pat = _px_letters[_ch]
                                _divs = []
                                for _on in _pat:
                                    if _on:
                                        _divs.append(f'<div class="dl-px" style="background:{_px_cols[_li]};box-shadow:0 0 5px {_px_cols[_li]};animation-delay:{_px_idx * 22}ms"></div>')
                                        _px_idx += 1
                                    else:
                                        _divs.append('<div class="dl-px dl-off"></div>')
                                _px_html.append(f'<div class="dl-let">{"".join(_divs)}</div>')

                            loading_placeholder.markdown(
                                '<style>'
                                '@keyframes pxBlink{0%,20%{opacity:0}25%{opacity:1}40%{opacity:0}60%{opacity:1}75%{opacity:0}100%{opacity:1}}'
                                '@keyframes barGrow{0%{width:4%}100%{width:96%}}'
                                '.dl-wrap{display:flex;flex-direction:column;align-items:center;padding:2rem 0;position:relative}'
                                '.dl-scanlines{position:absolute;inset:0;background:repeating-linear-gradient(0deg,transparent,transparent 3px,rgba(0,0,0,0.12) 3px,rgba(0,0,0,0.12) 4px);pointer-events:none}'
                                '.dl-grid{display:flex;gap:4px;margin-bottom:1.5rem}'
                                '.dl-let{display:grid;grid-template-columns:repeat(5,12px);grid-template-rows:repeat(7,12px);gap:2px;margin:0 6px}'
                                '.dl-px{width:12px;height:12px;opacity:0;animation:pxBlink 0.5s steps(5,end) forwards}'
                                '.dl-off{background:transparent!important;animation:none!important;opacity:0!important}'
                                '.dl-status{color:#8892a8;font-size:0.75rem;margin-top:0.8rem;font-family:"Courier New",monospace;letter-spacing:3px;text-transform:uppercase}'
                                '.dl-bar{width:252px;height:10px;background:#0d1117;outline:2px solid #2a3040;margin-top:1rem}'
                                '.dl-bar-fill{height:100%;background:repeating-linear-gradient(90deg,#00b46e 0,#00b46e 14px,transparent 14px,transparent 16px);width:4%;animation:barGrow 3s steps(16,end) forwards}'
                                '</style>'
                                '<div class="dl-wrap">'
                                '<div class="dl-scanlines"></div>'
                                '<div class="dl-grid">' + "".join(_px_html) + '</div>'
                                '<div class="dl-status">&#9658; OPTIMISING SQUAD...</div>'
                                '<div class="dl-bar"><div class="dl-bar-fill"></div></div>'
                                '</div>',
                                unsafe_allow_html=True,
                            )

                            plan = build_rolling_plan(
                                my_squad, df,
                                    bank=team_data["bank"],
                                    free_transfers=ft_available,
                                    purchase_prices=team_data.get("purchase_prices", {}),
                                    selling_prices_api=team_data.get("selling_prices_api", {}),
                                    xpts_map=xpts_map_adjusted,
                                    planning_gw_id=planning_gw_id,
                                    n_gws=6,
                                    chip_schedule=chip_schedule,
                                    team_fixture_counts=team_fixture_counts,
                                    locked_ids=planner_locked_ids,
                                    banned_ids=planner_banned_ids,
                                )
                            st.session_state["last_plan"] = plan
                            st.session_state["last_plan_fingerprint"] = _settings_fp

                            # Clear the loading animation
                            loading_placeholder.empty()

                        # Warn if settings have changed since plan was generated
                        if plan and not generate and st.session_state.get("last_plan_fingerprint") != _settings_fp:
                            st.warning("⚠️ Settings have changed since this plan was generated. Click **🚀 Generate 6-Gameweek Plan** above to refresh.")

                        # Plan summary
                        if plan:
                            total_xpts = 0
                            total_hits = 0
                            total_transfers = 0
                            for gw_e in plan:
                                xi = gw_e.get("xi")
                                if xi is not None and "xpts_gw" in xi.columns:
                                    cap = gw_e.get("captain")
                                    cap_mult = gw_e.get("captain_multiplier", 2)
                                    xi_pts = xi["xpts_gw"].sum()
                                    # Add captain bonus (extra x1 or x2)
                                    if cap is not None:
                                        cap_xpts = xpts_map.get(cap.get("id", 0), {}).get(gw_e["gw"], 0)
                                        xi_pts += cap_xpts * (cap_mult - 1)
                                    # Add bench if bench boost
                                    if gw_e.get("bench_boost") and gw_e.get("bench") is not None:
                                        xi_pts += gw_e["bench"]["xpts_gw"].sum() if "xpts_gw" in gw_e["bench"].columns else 0
                                    total_xpts += xi_pts
                                total_hits += gw_e.get("hit", 0)
                                total_transfers += len(gw_e.get("transfers", []))

                            sc1, sc2, sc3 = st.columns(3)
                            with sc1:
                                st.markdown(f"""<div class="metric-card">
                                    <div class="metric-label">Projected Total (6GW)</div>
                                    <div class="metric-value">{total_xpts:.0f} xPts</div>
                                    <div class="metric-sub">After captain bonus</div>
                                </div>""", unsafe_allow_html=True)
                            with sc2:
                                st.markdown(f"""<div class="metric-card">
                                    <div class="metric-label">Transfers Planned</div>
                                    <div class="metric-value">{total_transfers}</div>
                                    <div class="metric-sub">{total_hits}pts in hits</div>
                                </div>""", unsafe_allow_html=True)
                            with sc3:
                                chips_used = sum(1 for g in plan if g.get("chip"))
                                st.markdown(f"""<div class="metric-card">
                                    <div class="metric-label">Chips Used</div>
                                    <div class="metric-value">{chips_used}</div>
                                    <div class="metric-sub">of {len(chip_schedule)} planned</div>
                                </div>""", unsafe_allow_html=True)

                            st.markdown("")

                            for gw_entry in plan:
                                gw = gw_entry["gw"]
                                transfer = gw_entry["transfer"]
                                xi = gw_entry["xi"]
                                bench = gw_entry["bench"]
                                hit = gw_entry["hit"]
                                chip = gw_entry.get("chip")
                                captain = gw_entry.get("captain")
                                cap_mult = gw_entry.get("captain_multiplier", 2)
                                is_bb = gw_entry.get("bench_boost", False)

                                # Expander label with chip badge
                                chip_labels = {
                                    "wildcard": "🃏 WILDCARD",
                                    "free_hit": "⚡ FREE HIT",
                                    "triple_captain": "👑 TRIPLE CAPTAIN",
                                    "bench_boost": "💪 BENCH BOOST",
                                    "park_the_bus": "🚌 PARK THE BUS",
                                    "dynamic_duo": "👥 DYNAMIC DUO",
                                    "loan_rangers": "🔄 LOAN RANGERS",
                                    "frikort": "🎫 FRIKORT",
                                }
                                chip_str = f" — {chip_labels.get(chip, '')}" if chip else ""
                                with st.expander(f"**Gameweek {gw}**{chip_str}", expanded=(gw == planning_gw_id)):

                                    # Transfer / chip action
                                    if chip == "wildcard" or chip == "frikort":
                                        squad_count = len(gw_entry.get("squad", []))
                                        chip_icon = "🃏 WILDCARD" if chip == "wildcard" else "🎫 FRIKORT"
                                        st.markdown(
                                            f"<div class='transfer-card'>"
                                            f"<span style='color:#a78bfa;font-weight:700;'>{chip_icon} ACTIVE</span>"
                                            f"<br><span style='color:#8892a8;font-size:0.72rem;'>"
                                            f"Full squad rebuilt via MILP solver ({squad_count} players) — optimised for remaining GWs</span>"
                                            f"</div>",
                                            unsafe_allow_html=True,
                                        )
                                    elif chip == "free_hit":
                                        st.markdown(
                                            f"<div class='transfer-card'>"
                                            f"<span style='color:#38bdf8;font-weight:700;'>⚡ FREE HIT ACTIVE</span>"
                                            f"<br><span style='color:#8892a8;font-size:0.72rem;'>"
                                            f"Best possible squad for this single GW — reverts to your team next week</span>"
                                            f"</div>",
                                            unsafe_allow_html=True,
                                        )
                                    elif chip == "loan_rangers":
                                        st.markdown(
                                            f"<div class='transfer-card'>"
                                            f"<span style='color:#38bdf8;font-weight:700;'>🔄 LOAN RANGERS ACTIVE</span>"
                                            f"<br><span style='color:#8892a8;font-size:0.72rem;'>"
                                            f"Best possible squad with no team limit — reverts next week</span>"
                                            f"</div>",
                                            unsafe_allow_html=True,
                                        )
                                    elif chip == "park_the_bus":
                                        st.markdown(
                                            f"<div class='transfer-card'>"
                                            f"<span style='color:#34d399;font-weight:700;'>🚌 PARK THE BUS ACTIVE</span>"
                                            f"<br><span style='color:#8892a8;font-size:0.72rem;'>"
                                            f"All defender points doubled — no captain this GW</span>"
                                            f"</div>",
                                            unsafe_allow_html=True,
                                        )
                                    elif chip == "dynamic_duo":
                                        vice = gw_entry.get("vice_captain")
                                        vice_name = "?"
                                        if vice is not None:
                                            try:
                                                vice_name = vice["name"] if hasattr(vice, "__getitem__") else "?"
                                            except Exception:
                                                pass
                                        st.markdown(
                                            f"<div class='transfer-card'>"
                                            f"<span style='color:#fbbf24;font-weight:700;'>👥 DYNAMIC DUO ACTIVE</span>"
                                            f"<br><span style='color:#8892a8;font-size:0.72rem;'>"
                                            f"Captain ×3 + Vice Captain ({vice_name}) ×2</span>"
                                            f"</div>",
                                            unsafe_allow_html=True,
                                        )
                                    else:
                                        transfers_list = gw_entry.get("transfers", [])
                                        gw_hit = gw_entry.get("hit", 0)
                                        ft_used = gw_entry.get("ft_used", 0)
                                        n_total = len(transfers_list)

                                        if n_total > 0:
                                            # Summary line
                                            hit_parts = []
                                            if ft_used > 0:
                                                hit_parts.append(f"{ft_used} free")
                                            paid = n_total - ft_used
                                            if paid > 0:
                                                hit_parts.append(f"{paid} hit (-{paid * 4}pts)")
                                            summary = " + ".join(hit_parts)
                                            total_gain = sum(t["xpts_gain"] for t in transfers_list)

                                            st.markdown(
                                                f"<span style='color:#8892a8;font-size:0.75rem;'>"
                                                f"**{n_total} transfer{'s' if n_total > 1 else ''}** ({summary}) · "
                                                f"Horizon xPts gain: <span style='color:#34d399;font-weight:600;'>+{total_gain:.1f}</span>"
                                                f"{f' · Net after hit: +{total_gain - gw_hit:.1f}' if gw_hit > 0 else ''}"
                                                f"</span>",
                                                unsafe_allow_html=True,
                                            )

                                            for t_idx, t in enumerate(transfers_list):
                                                o = t["out"]
                                                i_p = t["in"]
                                                sp = calculate_selling_price(
                                                    o["id"], o["now_cost"],
                                                    team_data.get("purchase_prices", {}),
                                                    team_data.get("selling_prices_api", {})
                                                )
                                                is_free = (t_idx < ft_used)
                                                tag = "Free" if is_free else "-4pt hit"
                                                tag_color = "#34d399" if is_free else "#f87171"
                                                gw_gain_str = f" · This GW: +{t.get('xpts_gw_gain', 0):.1f}" if "xpts_gw_gain" in t else ""

                                                st.markdown(f"""<div class="transfer-card">
                                                    <span class="transfer-out">▼ {o['name']}</span>
                                                    <span style="color:#5a6580;font-size:0.7rem;">
                                                        {o.get('pos','?')} · {o.get('team','?')} · SP £{sp/10:.1f}m
                                                    </span>
                                                    &nbsp;<span class="transfer-arrow">→</span>&nbsp;
                                                    <span class="transfer-in">▲ {i_p['name']}</span>
                                                    <span style="color:#5a6580;font-size:0.7rem;">
                                                        {i_p.get('pos','?')} · {i_p.get('team','?')} · £{i_p['now_cost']/10:.1f}m
                                                    </span>
                                                    <br>
                                                    <span style="color:#34d399;font-size:0.72rem;font-weight:600;">
                                                        +{t['xpts_gain']:.1f} xPts remaining horizon{gw_gain_str}
                                                    </span>
                                                    <span style="color:{tag_color};font-size:0.7rem;font-weight:600;"> · {tag}</span>
                                                    <span style="color:#5a6580;font-size:0.7rem;"> ·
                                                        £{t['new_bank']/10:.1f}m ITB
                                                    </span>
                                                </div>""", unsafe_allow_html=True)
                                        else:
                                            st.markdown(
                                                "<div class='transfer-card'>"
                                                "<span style='color:#8892a8;'>No transfer — banking free transfer</span>"
                                                "</div>",
                                                unsafe_allow_html=True,
                                            )
                                    # Best XI pitch view
                                    if xi is not None and len(xi) >= 11:
                                        formation = get_formation_str(xi)
                                        xi_total = xi["xpts_gw"].sum() if "xpts_gw" in xi.columns else 0
                                        st.markdown(
                                            f"<span style='color:#8892a8;font-size:0.78rem;'>"
                                            f"Best XI: {formation} · Projected {xi_total:.1f} xPts</span>",
                                            unsafe_allow_html=True,
                                        )

                                        for pid_val, plabel in [(4, "FWD"), (3, "MID"), (2, "DEF"), (1, "GK")]:
                                            pp = xi[xi["pos_id"] == pid_val]
                                            if len(pp) > 0:
                                                cols = st.columns(max(len(pp), 1))
                                                for ci, (_, p) in enumerate(pp.iterrows()):
                                                    is_gk = (p["pos_id"] == 1)
                                                    is_cap = (captain is not None and p["id"] == captain.get("id"))
                                                    gw_xpts = p.get("xpts_gw", 0)
                                                    shirt_svg = make_shirt_svg(p.get("team", "???"), f"{gw_xpts:.1f}", is_gk=is_gk, is_captain=is_cap, player_name=p.get("name", ""), team_code=int(p.get("team_code", 0) or 0))
                                                    with cols[ci]:
                                                        html = f'<div style="text-align:center;">{shirt_svg}<div class="pitch-name">{p["name"]}</div><div class="pitch-price">£{p["price"]:.1f}m</div></div>'
                                                        st.markdown(html, unsafe_allow_html=True)

                                        if bench is not None and len(bench) > 0:
                                            if is_bb:
                                                bench_label = "💪 BENCH BOOST — all bench players score:"
                                            else:
                                                bench_label = "Bench:"
                                            bench_names = ", ".join([
                                                f"{r['name']} ({r.get('xpts_gw', 0):.1f})"
                                                for _, r in bench.iterrows()
                                            ])
                                            st.markdown(
                                                f"<span style='color:#5a6580;font-size:0.68rem;'>"
                                                f"{'💪 ' if is_bb else ''}{bench_label} {bench_names}</span>",
                                                unsafe_allow_html=True,
                                            )

                                    # Captain recommendation
                                    if captain is not None:
                                        cap_xpts = xpts_map.get(captain.get("id", 0), {}).get(gw, 0)
                                        cap_label = "Triple Captain" if cap_mult == 3 else "Captain"
                                        cap_emoji = "👑" if cap_mult == 3 else "©️"
                                        st.markdown(
                                            f"<span style='color:#fbbf24;font-size:0.78rem;font-weight:600;'>"
                                            f"{cap_emoji} {cap_label}: {captain.get('name', '?')} "
                                            f"({cap_xpts:.1f} xPts × {cap_mult} = {cap_xpts * cap_mult:.1f})</span>",
                                            unsafe_allow_html=True,
                                        )
                        else:
                            st.info("Could not build a rolling plan — not enough fixture data.")

                        # === EXPORT PLAN TO EXCEL ===
                        if plan and len(plan) > 0:
                            st.markdown("")
                            if st.button("📥 Export Plan to Excel", use_container_width=True):
                                import io
                                from openpyxl import Workbook
                                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                                wb = Workbook()

                                # --- Sheet 1: Plan Summary ---
                                ws_summary = wb.active
                                ws_summary.title = "Plan Summary"

                                header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
                                header_fill = PatternFill("solid", fgColor="2D2D3D")
                                pink_fill = PatternFill("solid", fgColor="F02D6E")
                                green_font = Font(color="34D399", bold=True, name="Arial")
                                red_font = Font(color="F87171", bold=True, name="Arial")
                                default_font = Font(name="Arial", size=10)
                                thin_border = Border(
                                    bottom=Side(style="thin", color="3A3A4A")
                                )

                                # Headers
                                summary_headers = ["GW", "Chip", "Transfers", "Hits (pts)", "FTs Used",
                                                   "Formation", "XI xPts", "Captain", "Cap xPts", "Bench xPts"]
                                for col, h in enumerate(summary_headers, 1):
                                    cell = ws_summary.cell(row=1, column=col, value=h)
                                    cell.font = header_font
                                    cell.fill = header_fill
                                    cell.alignment = Alignment(horizontal="center")

                                for row_idx, gw_e in enumerate(plan, 2):
                                    gw = gw_e["gw"]
                                    chip = gw_e.get("chip", "—") or "—"
                                    transfers_list = gw_e.get("transfers") or []
                                    hit = gw_e.get("hit", 0)
                                    ft_used = gw_e.get("ft_used", 0)
                                    xi_data = gw_e.get("xi")
                                    bench_data = gw_e.get("bench")
                                    captain = gw_e.get("captain")
                                    cap_mult = gw_e.get("captain_multiplier", 2)

                                    formation = get_formation_str(xi_data) if xi_data is not None and len(xi_data) >= 11 else "?"
                                    xi_xpts = xi_data["xpts_gw"].sum() if xi_data is not None and "xpts_gw" in xi_data.columns else 0
                                    bench_xpts = bench_data["xpts_gw"].sum() if bench_data is not None and "xpts_gw" in bench_data.columns else 0

                                    # Safe captain extraction (could be pandas Series or dict or None)
                                    cap_name = "?"
                                    cap_xpts = 0
                                    if captain is not None:
                                        try:
                                            cap_id = captain["id"] if hasattr(captain, "__getitem__") else 0
                                            cap_name = captain["name"] if hasattr(captain, "__getitem__") else "?"
                                            cap_xpts = xpts_map_adjusted.get(cap_id, {}).get(gw, 0) * cap_mult
                                        except Exception:
                                            pass

                                    ws_summary.cell(row=row_idx, column=1, value=f"GW{gw}").font = default_font
                                    ws_summary.cell(row=row_idx, column=2, value=chip).font = default_font
                                    ws_summary.cell(row=row_idx, column=3, value=len(transfers_list)).font = default_font
                                    c_hit = ws_summary.cell(row=row_idx, column=4, value=f"-{hit}" if hit > 0 else "0")
                                    c_hit.font = red_font if hit > 0 else default_font
                                    ws_summary.cell(row=row_idx, column=5, value=ft_used).font = default_font
                                    ws_summary.cell(row=row_idx, column=6, value=formation).font = default_font
                                    ws_summary.cell(row=row_idx, column=7, value=round(xi_xpts, 1)).font = green_font
                                    ws_summary.cell(row=row_idx, column=8, value=cap_name).font = default_font
                                    ws_summary.cell(row=row_idx, column=9, value=round(cap_xpts, 1)).font = default_font
                                    ws_summary.cell(row=row_idx, column=10, value=round(bench_xpts, 1)).font = default_font

                                for col in range(1, 11):
                                    ws_summary.column_dimensions[chr(64 + col)].width = 14

                                # --- Sheet 2: Transfers ---
                                ws_transfers = wb.create_sheet("Transfers")
                                t_headers = ["GW", "Out", "Out Team", "Out Pos", "Out Price",
                                            "In", "In Team", "In Pos", "In Price",
                                            "Horizon Gain", "GW Gain", "Type", "Bank After"]
                                for col, h in enumerate(t_headers, 1):
                                    cell = ws_transfers.cell(row=1, column=col, value=h)
                                    cell.font = header_font
                                    cell.fill = header_fill
                                    cell.alignment = Alignment(horizontal="center")

                                t_row = 2
                                for gw_e in plan:
                                    gw = gw_e["gw"]
                                    ft_used = gw_e.get("ft_used", 0)
                                    for t_idx, t in enumerate(gw_e.get("transfers", [])):
                                        o = t["out"]
                                        i_p = t["in"]
                                        is_free = t_idx < ft_used
                                        ws_transfers.cell(row=t_row, column=1, value=f"GW{gw}").font = default_font
                                        ws_transfers.cell(row=t_row, column=2, value=o.get("name", "?")).font = red_font
                                        ws_transfers.cell(row=t_row, column=3, value=o.get("team", "?")).font = default_font
                                        ws_transfers.cell(row=t_row, column=4, value=o.get("pos", "?")).font = default_font
                                        ws_transfers.cell(row=t_row, column=5, value=round(o.get("now_cost", 0) / 10, 1)).font = default_font
                                        ws_transfers.cell(row=t_row, column=6, value=i_p.get("name", "?")).font = green_font
                                        ws_transfers.cell(row=t_row, column=7, value=i_p.get("team", "?")).font = default_font
                                        ws_transfers.cell(row=t_row, column=8, value=i_p.get("pos", "?")).font = default_font
                                        ws_transfers.cell(row=t_row, column=9, value=round(i_p.get("now_cost", 0) / 10, 1)).font = default_font
                                        ws_transfers.cell(row=t_row, column=10, value=t.get("xpts_gain", 0)).font = default_font
                                        ws_transfers.cell(row=t_row, column=11, value=t.get("xpts_gw_gain", 0)).font = default_font
                                        ws_transfers.cell(row=t_row, column=12, value="Free" if is_free else "-4pt Hit").font = default_font
                                        ws_transfers.cell(row=t_row, column=13, value=round(t.get("new_bank", 0) / 10, 1)).font = default_font
                                        t_row += 1

                                for col in range(1, 14):
                                    ws_transfers.column_dimensions[chr(64 + col) if col <= 26 else "A" + chr(64 + col - 26)].width = 14

                                # --- Sheet 3: Starting XIs per GW ---
                                ws_xi = wb.create_sheet("Starting XIs")
                                xi_headers = ["GW", "Player", "Team", "Pos", "Price", "xPts", "Captain"]
                                for col, h in enumerate(xi_headers, 1):
                                    cell = ws_xi.cell(row=1, column=col, value=h)
                                    cell.font = header_font
                                    cell.fill = header_fill
                                    cell.alignment = Alignment(horizontal="center")

                                xi_row = 2
                                for gw_e in plan:
                                    gw = gw_e["gw"]
                                    xi_data = gw_e.get("xi")
                                    captain = gw_e.get("captain")
                                    cap_id = None
                                    if captain is not None:
                                        try:
                                            cap_id = captain["id"] if hasattr(captain, "__getitem__") else None
                                        except Exception:
                                            pass

                                    if xi_data is not None and len(xi_data) > 0:
                                        for _, p in xi_data.sort_values("pos_id").iterrows():
                                            is_cap = p["id"] == cap_id
                                            ws_xi.cell(row=xi_row, column=1, value=f"GW{gw}").font = default_font
                                            ws_xi.cell(row=xi_row, column=2, value=p["name"]).font = default_font
                                            ws_xi.cell(row=xi_row, column=3, value=p.get("team", "?")).font = default_font
                                            ws_xi.cell(row=xi_row, column=4, value=p.get("pos", "?")).font = default_font
                                            ws_xi.cell(row=xi_row, column=5, value=round(p.get("price", 0), 1)).font = default_font
                                            ws_xi.cell(row=xi_row, column=6, value=round(p.get("xpts_gw", 0), 1)).font = green_font
                                            ws_xi.cell(row=xi_row, column=7, value="(C)" if is_cap else "").font = Font(name="Arial", color="FFD700", bold=True) if is_cap else default_font
                                            xi_row += 1

                                for col in range(1, 8):
                                    ws_xi.column_dimensions[chr(64 + col)].width = 14

                                # Save to buffer
                                buf = io.BytesIO()
                                wb.save(buf)
                                buf.seek(0)

                                st.download_button(
                                    label="⬇️ Download Plan (.xlsx)",
                                    data=buf,
                                    file_name=f"datumly_plan_GW{planning_gw_id}-{planning_gw_id + 5}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True,
                                )

                    # Full squad table
                    st.markdown("")
                    st.subheader("Full Squad Breakdown")
                    sq_show = my_squad.sort_values(["is_starter", "pos_id", "xpts_total"],
                                                    ascending=[False, True, False])
                    # Add selling price column
                    sq_show = sq_show.copy()
                    sq_show["sell_price"] = sq_show.apply(
                        lambda r: calculate_selling_price(
                            r["id"], r["now_cost"],
                            team_data.get("purchase_prices", {}),
                            team_data.get("selling_prices_api", {})
                        ) / 10, axis=1
                    )
                    sq_display = sq_show[["name", "team", "pos", "price", "sell_price", "total_points",
                                          "form_str", "xpts_next_gw", "xpts_total", "is_starter"]].copy()
                    sq_display.columns = ["Player", "Team", "Pos", "Mkt Price", "Sell Price", "Pts",
                                          "Form", "xPts GW", "xPts 6GW", "Starter"]
                    sq_display["Starter"] = sq_display["Starter"].map({True: "XI", False: "Bench"})
                    sq_display = sq_display.reset_index(drop=True)
                    sq_display.index += 1
                    st.dataframe(sq_display, use_container_width=True)
                else:
                    st.warning("Could not match squad players to current data.")
        else:
            st.info("Enter your FPL Team ID above to get personalised transfer suggestions. "
                    "You can find it in the URL when you view your team on the FPL website "
                    "(e.g. fantasy.premierleague.com/entry/**123456**/event/1)")

    # ==================== DASHBOARD ====================
    with tab2:
        if len(active) > 0:
            c1, c2, c3, c4 = st.columns(4)
            top_xpts = qualified.loc[qualified["xpts_total"].idxmax()] if len(qualified) > 0 else None
            top_scorer = active.loc[active["total_points"].idxmax()]
            top_value = qualified.loc[qualified["value"].idxmax()] if len(qualified) > 0 else None
            top_form = active.loc[active["form"].idxmax()]

            cards = [
                (c1, "Highest xPts (6GW)", f"{top_xpts['xpts_total']:.1f}" if top_xpts is not None else "-",
                 f"{top_xpts['name']} (£{top_xpts['price']:.1f}m)" if top_xpts is not None else ""),
                (c2, "Top Scorer", str(int(top_scorer["total_points"])),
                 f"{top_scorer['name']} ({top_scorer['team']})"),
                (c3, "Best Value (xPts/£)", f"{top_value['value']:.1f}" if top_value is not None else "-",
                 f"{top_value['name']} (£{top_value['price']:.1f}m)" if top_value is not None else ""),
                (c4, "Hottest Form", f"{top_form['form']:.1f}",
                 f"{top_form['name']} ({top_form['team']})"),
            ]
            for col, label, val, sub in cards:
                with col:
                    st.markdown(f"""<div class="metric-card">
                        <div class="metric-label">{label}</div>
                        <div class="metric-value">{val}</div>
                        <div class="metric-sub">{sub}</div>
                    </div>""", unsafe_allow_html=True)

        st.markdown("")
        st.markdown('<div class="section-header">Top xPts Picks — Next 6 Gameweeks <span class="source-tag src-model">xPts Model</span></div>', unsafe_allow_html=True)
        if len(qualified) > 0:
            tp = qualified.nlargest(15, "xpts_total")[
                ["name", "team", "pos", "price", "total_points", "xpts_next_gw", "xpts_total", "value"]
            ].copy()
            tp.columns = ["Player", "Team", "Pos", "Price (£m)", "Actual Pts", "xPts Next GW", "xPts 6GW", "Value"]
            tp = tp.reset_index(drop=True)
            tp.index += 1
            st.dataframe(tp, use_container_width=True, height=540)

        st.markdown("")
        st.markdown('<div class="section-header">Differentials — Under 10% Ownership</div>', unsafe_allow_html=True)
        diffs = qualified[(qualified["selected_pct"] < 10) & (qualified["xpts_total"] > 0)].nlargest(10, "xpts_total")
        if len(diffs) > 0:
            dd = diffs[["name", "team", "pos", "price", "selected_pct", "xpts_total", "value"]].copy()
            dd.columns = ["Player", "Team", "Pos", "Price", "Own%", "xPts 6GW", "Value"]
            dd = dd.reset_index(drop=True)
            dd.index += 1
            st.dataframe(dd, use_container_width=True, height=380)

        # ==================== CAPTAIN OPTIMISER ====================
        st.markdown("")
        st.markdown(
            '<div class="section-header">👑 Captain Optimiser '
            '<span class="source-tag src-model">Next GW</span></div>',
            unsafe_allow_html=True,
        )
        st.caption("Ranks captain candidates by expected captaincy points (xPts × 2). "
                    "Shows ownership for differential edge — captaining a low-ownership player "
                    "who hauls gives you a bigger rank boost than a highly-owned pick.")

        if len(qualified) > 0:
            cap_candidates = qualified[qualified["xpts_next_gw"] > 0].copy()
            cap_candidates["cap_xpts"] = cap_candidates["xpts_next_gw"] * 2
            cap_candidates["differential_score"] = cap_candidates["cap_xpts"] * (1 + (100 - cap_candidates["selected_pct"]) / 100)

            # Get fixture info for next GW
            cap_candidates["next_fixture"] = cap_candidates["upcoming"].apply(
                lambda u: f"{'(H)' if u[0]['home'] else '(A)'} vs {teams.get(u[0]['opp'], {}).get('short_name', '?')}" if u else "?"
            )
            cap_candidates["fixture_diff"] = cap_candidates["upcoming"].apply(
                lambda u: u[0].get("difficulty", 3) if u else 3
            )

            cap_top = cap_candidates.nlargest(15, "cap_xpts")

            cap_display = cap_top[[
                "name", "team", "pos", "price", "next_fixture", "fixture_diff",
                "xpts_next_gw", "cap_xpts", "selected_pct", "rotation_risk", "start_prob"
            ]].copy()
            cap_display.columns = [
                "Player", "Team", "Pos", "Price", "Fixture", "FDR",
                "xPts", "Cap xPts", "Own%", "Rotation", "Start%"
            ]
            cap_display["Start%"] = (cap_display["Start%"] * 100).round(0).astype(int).astype(str) + "%"
            cap_display = cap_display.reset_index(drop=True)
            cap_display.index += 1
            st.dataframe(cap_display, use_container_width=True, height=540)

            # Highlight top pick
            if len(cap_top) > 0:
                best_cap = cap_top.iloc[0]
                st.markdown(
                    f"<div style='background:#1a1e2e;border-radius:8px;padding:0.8rem;margin:0.5rem 0;'>"
                    f"<span style='color:#fbbf24;font-weight:700;font-size:1.1rem;'>👑 Recommended Captain: "
                    f"{best_cap['name']}</span><br>"
                    f"<span style='color:#8892a8;font-size:0.8rem;'>"
                    f"{best_cap['team']} · {best_cap['next_fixture']} · "
                    f"{best_cap['xpts_next_gw']:.1f} xPts × 2 = {best_cap['cap_xpts']:.1f} · "
                    f"Owned by {best_cap['selected_pct']:.1f}% · "
                    f"Rotation: {best_cap['rotation_risk']} ({best_cap['start_prob']:.0%})</span></div>",
                    unsafe_allow_html=True,
                )

                # Differential captain suggestion
                diff_caps = cap_candidates[cap_candidates["selected_pct"] < 15].nlargest(3, "cap_xpts")
                if len(diff_caps) > 0:
                    diff_best = diff_caps.iloc[0]
                    st.markdown(
                        f"<div style='background:#1a2e1a;border-radius:8px;padding:0.8rem;margin:0.5rem 0;'>"
                        f"<span style='color:#34d399;font-weight:700;'>🎯 Differential Captain: "
                        f"{diff_best['name']}</span><br>"
                        f"<span style='color:#8892a8;font-size:0.8rem;'>"
                        f"{diff_best['team']} · {diff_best['next_fixture']} · "
                        f"{diff_best['xpts_next_gw']:.1f} xPts × 2 = {diff_best['cap_xpts']:.1f} · "
                        f"Only {diff_best['selected_pct']:.1f}% ownership — huge upside if they haul</span></div>",
                        unsafe_allow_html=True,
                    )

        # ==================== OWNERSHIP ANALYSIS ====================
        st.markdown("")
        st.markdown(
            '<div class="section-header">📊 Ownership Analysis '
            '<span class="source-tag src-fpl">Effective Ownership</span></div>',
            unsafe_allow_html=True,
        )
        st.caption("High ownership + high xPts = safe pick (everyone has them). "
                    "Low ownership + high xPts = differential (big rank gains if they score). "
                    "High ownership + low xPts = trap (everyone has them but they won't score).")

        if len(qualified) > 0:
            own_df = qualified[qualified["xpts_next_gw"] > 0].copy()

            # Template players
            col_safe, col_diff, col_trap = st.columns(3)

            with col_safe:
                st.markdown("**🛡️ Must-Haves** (high own%, high xPts)")
                must_haves = own_df[(own_df["selected_pct"] >= 20) & (own_df["xpts_next_gw"] > 0)].nlargest(8, "xpts_next_gw")
                for _, p in must_haves.iterrows():
                    st.markdown(
                        f"<span style='color:#8892a8;font-size:0.78rem;'>"
                        f"**{p['name']}** ({p['team']}) · {p['xpts_next_gw']:.1f} xPts · "
                        f"{p['selected_pct']:.0f}% owned</span>",
                        unsafe_allow_html=True,
                    )

            with col_diff:
                st.markdown("**🎯 Differentials** (low own%, high xPts)")
                differentials = own_df[(own_df["selected_pct"] < 15)].nlargest(8, "xpts_next_gw")
                for _, p in differentials.iterrows():
                    st.markdown(
                        f"<span style='color:#34d399;font-size:0.78rem;'>"
                        f"**{p['name']}** ({p['team']}) · {p['xpts_next_gw']:.1f} xPts · "
                        f"Only {p['selected_pct']:.1f}% owned</span>",
                        unsafe_allow_html=True,
                    )

            with col_trap:
                st.markdown("**⚠️ Traps** (high own%, low xPts next GW)")
                traps = own_df[own_df["selected_pct"] >= 15].nsmallest(8, "xpts_next_gw")
                for _, p in traps.iterrows():
                    st.markdown(
                        f"<span style='color:#f87171;font-size:0.78rem;'>"
                        f"**{p['name']}** ({p['team']}) · {p['xpts_next_gw']:.1f} xPts · "
                        f"{p['selected_pct']:.0f}% owned</span>",
                        unsafe_allow_html=True,
                    )

        # ==================== FIXTURE TICKER ====================
        st.markdown("")
        st.markdown(
            '<div class="section-header">📅 Fixture Sequence Ticker '
            '<span class="source-tag src-fpl">Next 6 GWs</span></div>',
            unsafe_allow_html=True,
        )
        st.caption("Green = easy fixture (FDR 1-2), amber = medium (3), red = hard (4-5). "
                    "Target players with runs of green fixtures.")

        if len(qualified) > 0:
            # Show top 30 players by xPts with their fixture sequence
            ticker_players = qualified.nlargest(30, "xpts_total")

            ticker_rows = []
            for _, p in ticker_players.iterrows():
                row = {"Player": p["name"], "Team": p["team"], "Pos": p["pos"],
                       "Price": f"£{p['price']:.1f}m", "xPts 6GW": round(p["xpts_total"], 1)}
                for i, fix in enumerate(p["upcoming"][:6]):
                    opp = teams.get(fix["opp"], {}).get("short_name", "?")
                    venue = "H" if fix["home"] else "A"
                    diff = fix.get("difficulty", 3)
                    row[f"GW{planning_gw_id + i}"] = f"{opp}({venue})"
                ticker_rows.append(row)

            if ticker_rows:
                ticker_df = pd.DataFrame(ticker_rows)
                ticker_df = ticker_df.reset_index(drop=True)
                ticker_df.index += 1
                st.dataframe(ticker_df, use_container_width=True, height=700)

    # ==================== PLAYER PROJECTIONS ====================
    with tab3:
        fc1, fc2, fc3, fc4, fc5 = st.columns([2, 1, 1, 1, 1])
        with fc1:
            search = st.text_input("🔍 Search", "", key="ps2")
        with fc2:
            pos_f = st.selectbox("Position", ["All"] + list(POS_FULL.values()), key="pf2")
        with fc3:
            team_f = st.selectbox("Team", ["All"] + sorted(df["team_name"].unique().tolist()), key="tf2")
        with fc4:
            price_f = st.selectbox("Price", ["All", "Under £5m", "£5-7m", "£7-10m", "Over £10m"], key="prf2")
        with fc5:
            so = {"xPts (6GW)": "xpts_total", "xPts Next GW": "xpts_next_gw", "Total Pts": "total_points",
                  "Form": "form", "Value": "value", "ICT": "ict_index"}
            sort_f = st.selectbox("Sort by", list(so.keys()), key="sf2")

        fl = active.copy()
        if search:
            sl = search.lower()
            fl = fl[fl["name"].str.lower().str.contains(sl, na=False) |
                     fl["second_name"].str.lower().str.contains(sl, na=False)]
        if pos_f != "All":
            fl = fl[fl["pos_id"] == {v: k for k, v in POS_FULL.items()}[pos_f]]
        if team_f != "All":
            fl = fl[fl["team_name"] == team_f]
        if price_f == "Under £5m": fl = fl[fl["price"] < 5]
        elif price_f == "£5-7m": fl = fl[(fl["price"] >= 5) & (fl["price"] < 7)]
        elif price_f == "£7-10m": fl = fl[(fl["price"] >= 7) & (fl["price"] < 10)]
        elif price_f == "Over £10m": fl = fl[fl["price"] >= 10]

        fl = fl.sort_values(so[sort_f], ascending=False)
        sd = fl.head(80)[["name", "team", "pos", "price", "total_points", "form_str",
                           "xpts_next_gw", "xpts_total", "xg_per90", "xa_per90",
                           "selected_pct", "value"]].copy()
        sd.columns = ["Player", "Team", "Pos", "Price", "Pts", "Form",
                       "xPts GW", "xPts 6GW", "xG/90", "xA/90", "Own%", "Value"]
        sd = sd.reset_index(drop=True)
        sd.index += 1
        st.dataframe(sd, use_container_width=True, height=700)
        st.caption(f"Showing {min(80, len(fl))} of {len(fl)} players · xPts model blends FPL xG/xA + betting odds + Club Elo + DefCon")

        # === xPts BREAKDOWN INSPECTOR ===
        st.markdown("")
        st.markdown("**🔬 xPts Breakdown Inspector**")
        inspect_labels = {
            row["id"]: f"{row['name']} ({row['team']}, {row['pos']}, £{row['price']:.1f}m)"
            for _, row in active.sort_values("xpts_total", ascending=False).head(200).iterrows()
        }
        inspect_pid = st.selectbox(
            "Select a player to inspect",
            options=list(inspect_labels.keys()),
            format_func=lambda pid: inspect_labels.get(pid, str(pid)),
            key="inspect_player",
        )

        if inspect_pid and inspect_pid in xpts_breakdown:
            player_bd = xpts_breakdown[inspect_pid]
            player_row = df[df["id"] == inspect_pid].iloc[0] if len(df[df["id"] == inspect_pid]) > 0 else None

            if player_row is not None:
                st.markdown(
                    f"<span style='color:#8892a8;font-size:0.8rem;'>"
                    f"{player_row['name']} · {player_row['team']} · {player_row['pos']} · "
                    f"£{player_row['price']:.1f}m · {player_row['minutes']} mins · "
                    f"xG/90: {player_row['xg_per90']:.3f} · xA/90: {player_row['xa_per90']:.3f} · "
                    f"DefCon/90: {player_row.get('defcon_per90', 0):.2f} · "
                    f"Rotation: {player_row.get('rotation_risk', '?')} ({player_row.get('start_prob', 0):.0%})"
                    f"</span>",
                    unsafe_allow_html=True,
                )

            for gw in sorted(player_bd.keys()):
                bd = player_bd[gw]
                venue = "🏠 Home" if bd["home"] else "✈️ Away"
                st.markdown(
                    f"<div style='background:#1a1e2e;border-radius:8px;padding:0.7rem;margin:0.4rem 0;'>"
                    f"<span style='color:#f02d6e;font-weight:700;'>GW{gw}</span> "
                    f"<span style='color:#8892a8;'>vs {bd['opponent']} {venue}</span> "
                    f"<span style='color:#34d399;font-weight:700;font-size:1.1rem;'>"
                    f"= {bd['total']} xPts</span><br>"
                    f"<span style='color:#5a6580;font-size:0.75rem;'>"
                    f"Appearance: {bd['appearance_pts']} · "
                    f"Goals: {bd['goal_pts']} (adj xG: {bd['adj_xg']}) · "
                    f"Assists: {bd['assist_pts']} (adj xA: {bd['adj_xa']}) · "
                    f"CS: {bd['cs_pts']} (prob: {bd['cs_prob']:.0%}) · "
                    f"Bonus: {bd['bonus_pts']} · "
                    f"Conceded: {bd['conceded_pts']} · "
                    f"DefCon: {bd['defcon_pts']}<br>"
                    f"Play prob: {bd['play_prob']:.0%} · "
                    f"Full game: {bd['full_game_prob']:.0%} · "
                    f"Exp 90s: {bd['expected_90s']} · "
                    f"Opp def str: {bd['opp_def_str']} · "
                    f"Team atk str: {bd['team_atk_str']} · "
                    f"Opp atk str: {bd['opp_atk_str']}"
                    f"</span></div>",
                    unsafe_allow_html=True,
                )
        elif inspect_pid:
            st.info("No breakdown available for this player (may not have enough minutes or upcoming fixtures).")

    # ==================== OPTIMAL SQUAD (MILP) ====================
    with tab4:
        st.markdown(
            '<div class="section-header">⭐ MILP-Optimised Squad '
            '<span class="source-tag src-model">PuLP Solver</span></div>',
            unsafe_allow_html=True,
        )
        st.caption("XI-aware optimal squad: maximises starting XI xPts with bench cost penalty. "
                    "Budget saved on bench is redirected to XI upgrades. "
                    "Constraints: £100m, 2 GK / 5 DEF / 5 MID / 3 FWD, max 3 per team.")

        if len(qualified) > 0:
            # Player lock/ban controls
            all_player_options = qualified.sort_values("xpts_total", ascending=False)
            player_labels = {
                row["id"]: f"{row['name']} ({row['team']}, {row['pos']}, £{row['price']:.1f}m)"
                for _, row in all_player_options.iterrows()
            }

            col_lock, col_ban = st.columns(2)
            with col_lock:
                locked_selections = st.multiselect(
                    "🔒 Lock players (must include)",
                    options=list(player_labels.keys()),
                    format_func=lambda pid: player_labels.get(pid, str(pid)),
                    placeholder="e.g. Salah, Haaland...",
                )
            with col_ban:
                banned_selections = st.multiselect(
                    "🚫 Ban players (exclude)",
                    options=list(player_labels.keys()),
                    format_func=lambda pid: player_labels.get(pid, str(pid)),
                    placeholder="e.g. injured players, avoid...",
                )

            locked_ids = set(locked_selections)
            banned_ids = set(banned_selections)

            # Warn if too many locked
            if len(locked_ids) > 15:
                st.warning("You can't lock more than 15 players.")
                locked_ids = set()

            with st.spinner("Running MILP solver..."):
                squad, solve_err = solve_optimal_squad(
                    qualified, "xpts_total", 1000,
                    locked_ids=locked_ids, banned_ids=banned_ids,
                )

            if squad is not None and len(squad) == 15:
                # Solve best XI
                xi, bench = solve_best_xi(squad, "xpts_next_gw")

                total_cost = squad["now_cost"].sum() / 10
                total_xpts = squad["xpts_total"].sum()
                xi_xpts = xi["xpts_next_gw"].sum() if xi is not None else 0
                xi_cost = xi["now_cost"].sum() / 10 if xi is not None else 0
                bench_cost = bench["now_cost"].sum() / 10 if bench is not None else 0
                formation = get_formation_str(xi)

                c1, c2, c3, c4, c5 = st.columns(5)
                c1.metric("Total Cost", f"£{total_cost:.1f}m")
                c2.metric("XI Cost", f"£{xi_cost:.1f}m")
                c3.metric("Bench Cost", f"£{bench_cost:.1f}m")
                c4.metric("XI xPts (Next GW)", f"{xi_xpts:.1f}")
                c5.metric("Formation", formation)

                st.markdown("")

                # Render pitch view
                if xi is not None:
                    for pid, plabel in [(4, "Forwards"), (3, "Midfielders"), (2, "Defenders"), (1, "Goalkeeper")]:
                        pp = xi[xi["pos_id"] == pid]
                        if len(pp) > 0:
                            st.markdown(f"<div class='pitch-row-label'>{plabel}</div>", unsafe_allow_html=True)
                            cols = st.columns(max(len(pp), 1))
                            for i, (_, p) in enumerate(pp.iterrows()):
                                is_gk = (p["pos_id"] == 1)
                                shirt_svg = make_shirt_svg(p["team"], f"{p['xpts_next_gw']:.1f}", is_gk=is_gk, player_name=p.get("name", ""), team_code=int(p.get("team_code", 0) or 0))
                                with cols[i]:
                                    html = f'<div style="text-align:center;">{shirt_svg}<div class="pitch-name">{p["name"]}</div><div class="pitch-price">£{p["price"]:.1f}m · {p["form_str"]}</div></div>'
                                    st.markdown(html, unsafe_allow_html=True)

                if bench is not None and len(bench) > 0:
                    st.markdown("**Bench**")
                    bcols = st.columns(len(bench))
                    for i, (_, p) in enumerate(bench.iterrows()):
                        is_gk = (p["pos_id"] == 1)
                        shirt_svg = make_shirt_svg(p["team"], f"{p['xpts_next_gw']:.1f}", is_gk=is_gk, width=44, height=44, player_name=p.get("name", ""), team_code=int(p.get("team_code", 0) or 0))
                        with bcols[i]:
                            html = f'<div style="text-align:center;opacity:0.65;">{shirt_svg}<div class="pitch-name">{p["name"]}</div><div class="pitch-price">{p["pos"]} · £{p["price"]:.1f}m</div></div>'
                            st.markdown(html, unsafe_allow_html=True)

                st.markdown("")
                st.subheader("Full Squad Breakdown")
                sq = squad.sort_values(["pos_id", "xpts_total"], ascending=[True, False])
                sq_show = sq[["name", "team", "pos", "price", "total_points", "xpts_next_gw", "xpts_total", "value"]].copy()
                sq_show.columns = ["Player", "Team", "Pos", "Price", "Actual Pts", "xPts GW", "xPts 6GW", "Value"]
                sq_show = sq_show.reset_index(drop=True)
                sq_show.index += 1
                st.dataframe(sq_show, use_container_width=True)
            else:
                st.warning(f"Could not find optimal squad: {solve_err}")

    # ==================== TRANSFER PLANNER ====================
    with tab5:
        st.markdown('<div class="section-header">🔄 Transfer Suggestions <span class="source-tag src-model">xPts Model</span></div>', unsafe_allow_html=True)
        st.caption("Finds the highest-xPts replacements for underperforming popular players, matched by position.")

        if len(qualified) > 0:
            cands = qualified.nlargest(30, "xpts_total")
            outs = qualified[qualified["selected_pct"] > 15].nsmallest(15, "xpts_total")
            transfers, ui, uo = [], set(), set()

            for _, ip in cands.iterrows():
                for _, op in outs.iterrows():
                    if ip["id"] in ui or op["id"] in uo:
                        continue
                    if ip["pos_id"] != op["pos_id"] or ip["id"] == op["id"]:
                        continue
                    if ip["xpts_total"] <= op["xpts_total"] * 1.15:
                        continue
                    reasons = []
                    if ip["xpts_next_gw"] > op["xpts_next_gw"]:
                        reasons.append(f"+{ip['xpts_next_gw'] - op['xpts_next_gw']:.1f} xPts next GW")
                    if ip["avg_difficulty"] < op["avg_difficulty"]:
                        reasons.append("Easier fixtures")
                    if ip["form"] > op["form"]:
                        reasons.append("Better form")
                    if not reasons:
                        reasons.append(f"+{ip['xpts_total'] - op['xpts_total']:.1f} xPts over 6GW")
                    transfers.append({"out": op, "in": ip, "reasons": reasons})
                    ui.add(ip["id"])
                    uo.add(op["id"])
                    if len(transfers) >= 8:
                        break
                if len(transfers) >= 8:
                    break

            if transfers:
                for t in transfers:
                    o, i = t["out"], t["in"]
                    rs = " · ".join(t["reasons"])
                    st.markdown(f"""<div class="transfer-card">
                        <span class="transfer-out">▼ {o['name']}</span>
                        <span style="color:#5a6580;font-size:0.72rem;"> {o['pos']} · {o['team']} · £{o['price']:.1f}m · {o['xpts_total']:.1f}xPts</span>
                        &nbsp;<span class="transfer-arrow">→</span>&nbsp;
                        <span class="transfer-in">▲ {i['name']}</span>
                        <span style="color:#5a6580;font-size:0.72rem;"> {i['pos']} · {i['team']} · £{i['price']:.1f}m · {i['xpts_total']:.1f}xPts</span>
                        <br><span style="color:#8892a8;font-size:0.7rem;">{rs}</span>
                    </div>""", unsafe_allow_html=True)
            else:
                st.info("No clear transfer improvements found.")

        st.markdown("")
        cr, cf = st.columns(2)
        with cr:
            st.subheader("📈 Most Transferred In")
            ri = active.nlargest(10, "transfers_in")[["name", "team", "pos", "price", "transfers_in", "xpts_total"]].copy()
            ri.columns = ["Player", "Team", "Pos", "Price", "In", "xPts 6GW"]
            ri = ri.reset_index(drop=True); ri.index += 1
            st.dataframe(ri, use_container_width=True)
        with cf:
            st.subheader("📉 Most Transferred Out")
            fo = active.nlargest(10, "transfers_out")[["name", "team", "pos", "price", "transfers_out", "xpts_total"]].copy()
            fo.columns = ["Player", "Team", "Pos", "Price", "Out", "xPts 6GW"]
            fo = fo.reset_index(drop=True); fo.index += 1
            st.dataframe(fo, use_container_width=True)

    # ==================== FIXTURES ====================
    with tab6:
        st.markdown('<div class="section-header">Fixture Difficulty — Next 6 Gameweeks <span class="source-tag src-odds">Odds-enhanced</span></div>', unsafe_allow_html=True)
        gw_range = list(range(planning_gw_id, planning_gw_id + 6))

        fm = {t_id: {} for t_id in teams}
        for f in fixtures_list:
            if f.get("event") in gw_range:
                if f["team_h"] in fm:
                    fm[f["team_h"]][f["event"]] = {"opp": f["team_a"], "home": True, "diff": f.get("team_h_difficulty", 3)}
                if f["team_a"] in fm:
                    fm[f["team_a"]][f["event"]] = {"opp": f["team_h"], "home": False, "diff": f.get("team_a_difficulty", 3)}

        rows = []
        for t_id, td in teams.items():
            row = {"Team": td["short_name"]}
            diffs = []
            t_short = td["short_name"]
            t_odds = {v: team_odds.get(k) for k, v in TEAM_NAME_MAP.items()}.get(t_short, {})

            for gw in gw_range:
                fix = fm.get(t_id, {}).get(gw)
                fc = team_fixture_counts.get(t_id, {}).get(gw, 1)
                if fc == 0:
                    row[f"GW{gw}"] = "BLANK"
                elif fix:
                    opp = teams.get(fix["opp"], {}).get("short_name", "???")
                    pre = "" if fix["home"] else "@"
                    dgw = " [DGW]" if fc >= 2 else ""
                    row[f"GW{gw}"] = f"{pre}{opp} ({fix['diff']}){dgw}"
                    diffs.append(fix["diff"])
                else:
                    row[f"GW{gw}"] = "-"

            row["Avg FDR"] = round(np.mean(diffs), 1) if diffs else 3.0

            # Add odds-derived CS probability if available
            if t_odds and isinstance(t_odds, dict):
                row["CS%"] = f"{t_odds.get('cs_prob', 0)*100:.0f}%"
                row["Atk Str"] = f"{t_odds.get('attack_strength', 1):.2f}"
            else:
                row["CS%"] = "-"
                row["Atk Str"] = "-"

            rows.append(row)

        fdf = pd.DataFrame(rows).sort_values("Avg FDR").reset_index(drop=True)
        fdf.index += 1
        st.dataframe(fdf, use_container_width=True, height=740)

    # ==================== BACKTEST ====================
    with tab7:
        st.markdown(
            '<div class="section-header">🔬 Model Backtesting '
            '<span class="source-tag src-model">Accuracy Analysis</span></div>',
            unsafe_allow_html=True,
        )
        st.caption("Compares Datumly's predicted xPts against actual FPL points for recent gameweeks. "
                    "This tells you how accurate the model is and where it's failing.")

        # Fetch actual points for recent completed GWs
        events = bootstrap.get("events", [])
        completed_gws = [e for e in events if e.get("finished") and e.get("id", 0) >= max(1, planning_gw_id - 7)]
        completed_gws = sorted(completed_gws, key=lambda e: e["id"], reverse=True)[:6]

        if completed_gws and len(active) > 0:
            if st.button("🔬 Run Backtest", use_container_width=True, type="primary"):
                headers = {"User-Agent": "FPL-Optimizer/2.0"}

                all_comparisons = []
                gw_summaries = []

                with st.spinner("Building retroactive predictions for completed GWs..."):
                    # Rebuild upcoming fixtures for PAST GWs
                    backtest_gw_ids = [e["id"] for e in completed_gws]
                    min_bt_gw = min(backtest_gw_ids)
                    max_bt_gw = max(backtest_gw_ids)

                    # Build opponent map for past GWs
                    bt_upcoming = {}
                    for t_id in teams:
                        bt_upcoming[t_id] = []
                    for f in fixtures_list:
                        ev = f.get("event")
                        if ev and min_bt_gw <= ev <= max_bt_gw:
                            if f["team_h"] in bt_upcoming:
                                bt_upcoming[f["team_h"]].append({
                                    "gw": ev, "opp_id": f["team_a"], "home": True,
                                    "difficulty": f.get("team_h_difficulty", 3)
                                })
                            if f["team_a"] in bt_upcoming:
                                bt_upcoming[f["team_a"]].append({
                                    "gw": ev, "opp_id": f["team_h"], "home": False,
                                    "difficulty": f.get("team_a_difficulty", 3)
                                })

                    # Temporarily monkey-patch the upcoming map in the model
                    # by calling build_xpts_model with a shifted current_gw_id
                    bt_xpts_map, _ = build_xpts_model(
                        df, team_odds, teams, fixtures_list, min_bt_gw,
                        form_xg_data=None,
                        team_fixture_counts=team_fixture_counts,
                        elo_ratings=None,
                        live_odds=None,
                        rotation_data=None,
                    )

                with st.spinner("Fetching actual points data..."):
                    for gw_event in completed_gws:
                        gw_id = gw_event["id"]
                        try:
                            resp = requests.get(
                                f"{FPL_BASE}/event/{gw_id}/live/",
                                headers=headers, timeout=15,
                            )
                            if resp.status_code != 200:
                                continue
                            live_data = resp.json()
                            elements = live_data.get("elements", [])

                            gw_errors = []
                            for el in elements:
                                pid = el["id"]
                                actual_pts = el.get("stats", {}).get("total_points", 0)
                                predicted_pts = bt_xpts_map.get(pid, {}).get(gw_id, None)

                                if predicted_pts is None or predicted_pts == 0:
                                    continue

                                mins = el.get("stats", {}).get("minutes", 0)
                                if mins == 0:
                                    continue

                                player_info = df[df["id"] == pid]
                                if len(player_info) == 0:
                                    continue
                                p = player_info.iloc[0]

                                error = actual_pts - predicted_pts
                                abs_error = abs(error)
                                gw_errors.append(abs_error)

                                all_comparisons.append({
                                    "GW": gw_id,
                                    "Player": p["name"],
                                    "Team": p["team"],
                                    "Pos": p["pos"],
                                    "Predicted": round(predicted_pts, 1),
                                    "Actual": actual_pts,
                                    "Error": round(error, 1),
                                    "Abs Error": round(abs_error, 1),
                                })

                            if gw_errors:
                                gw_summaries.append({
                                    "GW": gw_id,
                                    "Players": len(gw_errors),
                                    "MAE": round(np.mean(gw_errors), 2),
                                    "Median AE": round(np.median(gw_errors), 2),
                                    "Within 2pts": f"{sum(1 for e in gw_errors if e <= 2) / len(gw_errors):.0%}",
                                    "Within 4pts": f"{sum(1 for e in gw_errors if e <= 4) / len(gw_errors):.0%}",
                                })
                        except Exception:
                            continue

                if all_comparisons:
                    comp_df = pd.DataFrame(all_comparisons)
                    summ_df = pd.DataFrame(gw_summaries)

                    # Overall metrics
                    overall_mae = comp_df["Abs Error"].mean()
                    overall_median = comp_df["Abs Error"].median()
                    within_2 = (comp_df["Abs Error"] <= 2).mean()
                    within_4 = (comp_df["Abs Error"] <= 4).mean()
                    correlation = comp_df["Predicted"].corr(comp_df["Actual"])

                    st.markdown("### Overall Model Accuracy")
                    mc1, mc2, mc3, mc4, mc5 = st.columns(5)
                    mc1.metric("MAE", f"{overall_mae:.2f} pts")
                    mc2.metric("Median Error", f"{overall_median:.2f} pts")
                    mc3.metric("Within 2pts", f"{within_2:.0%}")
                    mc4.metric("Within 4pts", f"{within_4:.0%}")
                    mc5.metric("Correlation", f"{correlation:.3f}")

                    st.markdown(
                        "<span style='color:#5a6580;font-size:0.75rem;'>"
                        "MAE (Mean Absolute Error) = average difference between predicted and actual points. "
                        "Lower is better. A MAE of ~2.5 is good for FPL. "
                        "Correlation measures how well predicted rankings match actual rankings (1.0 = perfect)."
                        "</span>",
                        unsafe_allow_html=True,
                    )

                    st.markdown("")
                    st.markdown("### Per-Gameweek Breakdown")
                    st.dataframe(summ_df, use_container_width=True)

                    # Error by position
                    st.markdown("")
                    st.markdown("### Accuracy by Position")
                    pos_errors = comp_df.groupby("Pos").agg(
                        Players=("Player", "count"),
                        MAE=("Abs Error", "mean"),
                        Median=("Abs Error", "median"),
                        Avg_Predicted=("Predicted", "mean"),
                        Avg_Actual=("Actual", "mean"),
                    ).round(2).reset_index()
                    pos_errors.columns = ["Position", "Players", "MAE", "Median Error", "Avg Predicted", "Avg Actual"]
                    st.dataframe(pos_errors, use_container_width=True)

                    # Biggest misses (for debugging)
                    st.markdown("")
                    st.markdown("### Biggest Misses (model vs reality)")
                    biggest = comp_df.nlargest(20, "Abs Error")[
                        ["GW", "Player", "Team", "Pos", "Predicted", "Actual", "Error"]
                    ].reset_index(drop=True)
                    biggest.index += 1
                    st.dataframe(biggest, use_container_width=True)

                    # Best predictions
                    st.markdown("")
                    st.markdown("### Most Accurate Predictions")
                    best = comp_df.nsmallest(20, "Abs Error")[
                        ["GW", "Player", "Team", "Pos", "Predicted", "Actual", "Error"]
                    ].reset_index(drop=True)
                    best.index += 1
                    st.dataframe(best, use_container_width=True)

                    # Direction accuracy — did we correctly rank players?
                    st.markdown("")
                    st.markdown("### Ranking Accuracy")
                    st.caption("For each GW, did our top 20 predicted players actually outscore the average?")
                    for gw_id in sorted(comp_df["GW"].unique()):
                        gw_data = comp_df[comp_df["GW"] == gw_id]
                        top20_predicted = gw_data.nlargest(20, "Predicted")
                        top20_actual_avg = top20_predicted["Actual"].mean()
                        all_avg = gw_data["Actual"].mean()
                        edge = top20_actual_avg - all_avg
                        emoji = "✅" if edge > 0 else "⚠️"
                        st.markdown(
                            f"<span style='color:#8892a8;font-size:0.8rem;'>"
                            f"GW{gw_id}: Top 20 predicted averaged "
                            f"<span style='color:{'#34d399' if edge > 0 else '#f87171'};font-weight:600;'>"
                            f"{top20_actual_avg:.1f}pts</span> vs all-player avg "
                            f"{all_avg:.1f}pts {emoji} "
                            f"(edge: {'+' if edge > 0 else ''}{edge:.1f}pts)</span>",
                            unsafe_allow_html=True,
                        )

                else:
                    st.warning("Could not generate backtest — no comparison data available.")

            # ==================== PARAMETER TUNER ====================
            st.markdown("---")
            st.markdown(
                '<div class="section-header">⚙️ Model Parameter Tuner '
                '<span class="source-tag src-model">Auto-Calibration</span></div>',
                unsafe_allow_html=True,
            )
            st.caption("Current model parameters and recommended values. "
                        "Adjust these to see how they affect xPts projections. "
                        "Optimal values are found by minimising prediction error against actual GW points.")

            st.markdown("**Current Parameters:**")

            # Editable parameters with current defaults
            tune_col1, tune_col2 = st.columns(2)
            with tune_col1:
                t_form_blend = st.slider(
                    "Form vs Season xG blend (% recent form)",
                    min_value=0.0, max_value=1.0, value=0.60, step=0.05,
                    help="0.6 = 60% recent form, 40% season average. Higher = more reactive to form.",
                    key="tune_form",
                )
                t_regression = st.slider(
                    "Over/underperformance regression strength",
                    min_value=0.0, max_value=0.60, value=0.30, step=0.05,
                    help="How aggressively to regress overperformers towards their xG. 0.3 = 30% regression.",
                    key="tune_regression",
                )
                t_home_boost = st.slider(
                    "Home advantage multiplier",
                    min_value=1.0, max_value=1.25, value=1.10, step=0.01,
                    help="xG multiplier for home teams. 1.10 = 10% boost at home.",
                    key="tune_home",
                )
                t_away_factor = st.slider(
                    "Away penalty multiplier",
                    min_value=0.80, max_value=1.0, value=0.95, step=0.01,
                    help="xG multiplier for away teams. 0.95 = 5% reduction away.",
                    key="tune_away",
                )

            with tune_col2:
                t_bonus = st.slider(
                    "Average bonus points per appearance",
                    min_value=0.0, max_value=1.5, value=0.50, step=0.05,
                    help="Expected bonus points per game for an average player.",
                    key="tune_bonus",
                )
                t_defcon_def = st.slider(
                    "DefCon scaling — Defenders",
                    min_value=0.0, max_value=1.0, value=0.60, step=0.05,
                    help="How aggressively to project DefCon points for defenders.",
                    key="tune_defcon_def",
                )
                t_defcon_mid = st.slider(
                    "DefCon scaling — Midfielders",
                    min_value=0.0, max_value=0.6, value=0.35, step=0.05,
                    help="How aggressively to project DefCon points for midfielders.",
                    key="tune_defcon_mid",
                )
                t_xa_regression = st.slider(
                    "Assist regression strength",
                    min_value=0.0, max_value=0.50, value=0.25, step=0.05,
                    help="How aggressively to regress assist overperformers.",
                    key="tune_xa_reg",
                )

            # Show what would change
            changes = []
            if t_form_blend != 0.60: changes.append(f"Form blend: 0.60 → {t_form_blend:.2f}")
            if t_regression != 0.30: changes.append(f"Regression: 0.30 → {t_regression:.2f}")
            if t_home_boost != 1.10: changes.append(f"Home boost: 1.10 → {t_home_boost:.2f}")
            if t_away_factor != 0.95: changes.append(f"Away factor: 0.95 → {t_away_factor:.2f}")
            if t_bonus != 0.50: changes.append(f"Bonus avg: 0.50 → {t_bonus:.2f}")
            if t_defcon_def != 0.60: changes.append(f"DefCon DEF: 0.60 → {t_defcon_def:.2f}")
            if t_defcon_mid != 0.35: changes.append(f"DefCon MID: 0.35 → {t_defcon_mid:.2f}")
            if t_xa_regression != 0.25: changes.append(f"Assist reg: 0.25 → {t_xa_regression:.2f}")

            if changes:
                st.info("Parameter changes: " + " · ".join(changes))
                st.markdown(
                    "<span style='color:#fbbf24;font-size:0.8rem;'>"
                    "⚡ To apply these changes, they need to be hardcoded into the model. "
                    "Use the backtest results above to determine which parameter values "
                    "minimise MAE, then update the code accordingly.</span>",
                    unsafe_allow_html=True,
                )

            # Parameter reference guide
            with st.expander("📖 Parameter Guide — what each one does"):
                st.markdown("""
**Form vs Season blend** — Controls how much weight the model gives to recent performance (last 7 GWs) versus the full season average. Higher values make the model more reactive to hot/cold streaks. Too high = overreacts to small samples. Too low = misses form changes.

**Regression strength** — When a player scores more goals than their xG suggests, this parameter pulls their projected xG back towards the mean. 0.30 means 30% regression. Higher = more conservative projections for streaky scorers.

**Home advantage** — Multiplier applied to xG/xA for home teams. The PL average is roughly +10% at home, but this varies by team.

**Away penalty** — Multiplier for away teams. Typically 5-10% reduction in xG.

**Bonus average** — Expected bonus points per appearance for an average player. The FPL average is roughly 0.4-0.6 per game across all players.

**DefCon scaling** — Controls how aggressively the model projects defensive contribution points. Defenders earn DefCon more reliably than midfielders, hence the separate sliders.

**Assist regression** — Same as goal regression but for assists. Assists are slightly more "sticky" (less random) than goals, so lower regression is appropriate.
                """)

        else:
            st.info("Need completed gameweeks to run backtest.")

    # ==================== CHIP STRATEGY ====================
    with tab8:
        st.markdown(
            '<div class="section-header">🎯 Chip Strategy Optimiser '
            '<span class="source-tag src-model">Brute-Force</span></div>',
            unsafe_allow_html=True,
        )
        st.caption("Evaluates every possible chip-to-gameweek combination and picks the strategy "
                    "that maximises your total xPts over the next 6 GWs. Requires your FPL ID to be loaded.")

        # Need team data for this feature
        if "team_data" not in st.session_state or st.session_state.get("team_data") is None:
            st.info("Load your FPL team in the 🏠 My Team tab first, then come back here.")
        else:
            team_data_chip = st.session_state["team_data"]
            chips_rem_chip = team_data_chip.get("chips_remaining", {})
            has_chips = any(v > 0 for v in chips_rem_chip.values())

            if not has_chips:
                st.info("You've used all your chips for this half-season. No strategy to optimise.")
            else:
                chip_labels_chip = {
                    "wildcard": "🃏 Wildcard", "freehit": "⚡ Free Hit",
                    "3xc": "👑 Triple Captain", "bboost": "💪 Bench Boost"
                }
                remaining_str_chip = " · ".join([
                    chip_labels_chip.get(k, k) for k, v in chips_rem_chip.items() if v > 0
                ])
                st.markdown(
                    f'<div style="background:#1a1e2e;border-radius:8px;padding:0.8rem;margin-bottom:0.8rem;">'
                    f'<span style="color:#f02d6e;font-weight:700;">🎯 Chips Remaining:</span> '
                    f'<span style="color:#8892a8;">{remaining_str_chip}</span></div>',
                    unsafe_allow_html=True,
                )

                my_squad_chip = df[df["id"].isin(team_data_chip["squad_ids"])].copy()
                ft_chip = team_data_chip.get("free_transfers", 1)

                # User-defined expected DGWs and BGWs
                st.markdown("**📅 Expected Double & Blank Gameweeks**")
                st.caption("The FPL fixture feed often doesn't show DGWs/BGWs until they're confirmed. "
                            "Use community intel (Ben Crellin, FPL Scout) to flag expected DGWs and BGWs. "
                            "The chip strategy will weight these weeks heavily for TC/BB/FH.")

                remaining_gws = list(range(planning_gw_id, 39))
                dgw_col, bgw_col = st.columns(2)
                with dgw_col:
                    expected_dgws = st.multiselect(
                        "🔥 Expected DGW weeks",
                        options=remaining_gws,
                        default=[gw for gw in [33, 36] if gw in remaining_gws],
                        help="GWs where you expect teams to have double fixtures",
                        key="exp_dgws",
                    )
                with bgw_col:
                    expected_bgws = st.multiselect(
                        "⚠️ Expected BGW weeks",
                        options=remaining_gws,
                        default=[gw for gw in [34, 37] if gw in remaining_gws],
                        help="GWs where you expect teams to blank",
                        key="exp_bgws",
                    )

                st.markdown("")

                if st.button("🚀 Find Optimal Chip Strategy", use_container_width=True, type="primary"):
                    from itertools import combinations, permutations
                    from collections import Counter

                    # === FIXTURE ANALYSIS — show DGWs/blanks ===
                    st.markdown("### 📅 Fixture Analysis")
                    events_all = bootstrap.get("events", [])
                    future_all = [e for e in events_all if e.get("id", 0) >= planning_gw_id and e.get("id", 0) <= LC["season_gws"]]

                    dgw_gws = []
                    blank_gws = []
                    for gw_ev in sorted(future_all, key=lambda x: x["id"]):
                        gw_id_a = gw_ev["id"]
                        gw_fix = [f for f in fixtures_list if f.get("event") == gw_id_a]
                        tc = Counter()
                        teams_playing = set()
                        for f in gw_fix:
                            tc[f["team_h"]] += 1
                            tc[f["team_a"]] += 1
                            teams_playing.add(f["team_h"])
                            teams_playing.add(f["team_a"])
                        n_dgw = sum(1 for t, c in tc.items() if c >= 2)
                        n_blank = 20 - len(teams_playing)
                        n_matches = len(gw_fix)

                        # Combine API-detected and user-flagged DGWs/BGWs
                        is_dgw = n_dgw > 0 or gw_id_a in expected_dgws
                        is_bgw = n_blank > 0 or gw_id_a in expected_bgws

                        marker = ""
                        if is_dgw and n_dgw > 0:
                            marker = f"🔥 DGW (confirmed — {n_dgw} teams)"
                            dgw_gws.append(gw_id_a)
                        elif is_dgw:
                            marker = f"🔥 DGW (expected — user flagged)"
                            dgw_gws.append(gw_id_a)
                        elif is_bgw and n_blank > 0:
                            marker = f"⚠️ BGW (confirmed — {n_blank} teams blanking)"
                            blank_gws.append(gw_id_a)
                        elif is_bgw:
                            marker = f"⚠️ BGW (expected — user flagged)"
                            blank_gws.append(gw_id_a)
                        else:
                            marker = f"📋 Standard"

                        # Best captain xPts for this GW
                        best_cap = 0
                        for _, p in qualified.head(50).iterrows():
                            gx = xpts_map.get(p["id"], {}).get(gw_id_a, 0)
                            if gx > best_cap:
                                best_cap = gx

                        # If user-flagged DGW, estimate boosted captain xPts (×1.7 for DGW)
                        cap_display = best_cap
                        if is_dgw and n_dgw == 0:
                            cap_display = best_cap * 1.7  # rough DGW uplift

                        st.markdown(
                            f'<span style="color:#8892a8;font-size:0.8rem;">'
                            f'GW{gw_id_a}: {n_matches} matches · {marker} · '
                            f'Best captain: {cap_display:.1f} xPts'
                            f'{"*" if is_dgw and n_dgw == 0 else ""}</span>',
                            unsafe_allow_html=True,
                        )

                    if not dgw_gws and not blank_gws:
                        st.markdown(
                            '<div style="background:#1a1e2e;border-radius:8px;padding:0.8rem;margin:0.5rem 0;">'
                            '<span style="color:#fbbf24;font-weight:600;">⚠️ No DGWs or BGWs detected or flagged.</span><br>'
                            '<span style="color:#8892a8;font-size:0.8rem;">'
                            'Use the inputs above to flag expected DGW/BGW weeks based on cup results '
                            'and community intel. This significantly affects chip recommendations.</span></div>',
                            unsafe_allow_html=True,
                        )
                    elif dgw_gws or blank_gws:
                        dgw_str = f"DGWs: GW{', GW'.join(str(g) for g in dgw_gws)}" if dgw_gws else ""
                        bgw_str = f"BGWs: GW{', GW'.join(str(g) for g in blank_gws)}" if blank_gws else ""
                        st.markdown(
                            f'<div style="background:#1a2e1a;border-radius:8px;padding:0.8rem;margin:0.5rem 0;">'
                            f'<span style="color:#34d399;font-weight:600;">✅ Key weeks identified: '
                            f'{dgw_str}{" · " if dgw_str and bgw_str else ""}{bgw_str}</span><br>'
                            f'<span style="color:#8892a8;font-size:0.8rem;">'
                            f'TC/BB are best on DGW weeks. FH is best on BGW weeks. '
                            f'WC is best 1-2 GWs before a DGW to load up on DGW players.</span></div>',
                            unsafe_allow_html=True,
                        )

                    st.markdown("---")

                    available_chips = [k for k, v in chips_rem_chip.items() if v > 0]
                    chip_name_map = {
                        "wildcard": "wildcard", "freehit": "free_hit",
                        "3xc": "triple_captain", "bboost": "bench_boost",
                    }
                    plan_gws = [planning_gw_id + i for i in range(6)]

                    with st.spinner("Running baseline plan and calculating chip values..."):
                        # Step 1: Run baseline plan ONCE (no chips)
                        try:
                            baseline_plan = build_rolling_plan(
                                my_squad_chip, df,
                                bank=team_data_chip["bank"],
                                free_transfers=ft_chip,
                                purchase_prices=team_data_chip.get("purchase_prices", {}),
                                selling_prices_api=team_data_chip.get("selling_prices_api", {}),
                                xpts_map=xpts_map,
                                planning_gw_id=planning_gw_id,
                                n_gws=6,
                                chip_schedule={},
                                team_fixture_counts=team_fixture_counts,
                            )
                            baseline_gw = {g["gw"]: g for g in baseline_plan} if baseline_plan else {}
                            baseline_total = sum(g.get("total_xpts", 0) for g in baseline_plan) if baseline_plan else 0
                        except Exception:
                            baseline_plan = None
                            baseline_gw = {}
                            baseline_total = 0

                        # Step 2: Pre-compute chip gain for each GW analytically
                        # This is FAST — no planner calls needed
                        chip_gains = {}  # {gw: {chip_key: gain}}

                        for gw in plan_gws:
                            chip_gains[gw] = {}
                            gw_data = baseline_gw.get(gw, {})
                            xi = gw_data.get("xi")
                            bench = gw_data.get("bench")
                            captain = gw_data.get("captain")
                            is_dgw = gw in dgw_gws
                            is_bgw = gw in blank_gws
                            dgw_mult = 1.7 if is_dgw else 1.0

                            # TRIPLE CAPTAIN gain = captain's xPts × 1 (extra double)
                            cap_xpts = 0
                            if captain is not None:
                                cap_id = captain.get("id", 0) if isinstance(captain, dict) else getattr(captain, "id", 0)
                                cap_xpts = xpts_map.get(cap_id, {}).get(gw, 0)
                            tc_gain = cap_xpts * dgw_mult  # DGW captains score ~1.7x
                            chip_gains[gw]["3xc"] = round(tc_gain, 1)

                            # BENCH BOOST gain = sum of bench xPts
                            bb_gain = 0
                            if bench is not None and len(bench) > 0:
                                bb_gain = bench["id"].map(
                                    lambda pid: xpts_map.get(pid, {}).get(gw, 0)
                                ).sum()
                            bb_gain *= dgw_mult  # DGW bench players score ~1.7x
                            chip_gains[gw]["bboost"] = round(bb_gain, 1)

                            # FREE HIT gain = best possible squad - current squad xPts
                            current_gw_xpts = 0
                            if xi is not None and len(xi) > 0:
                                current_gw_xpts = xi["id"].map(
                                    lambda pid: xpts_map.get(pid, {}).get(gw, 0)
                                ).sum()
                            # Best possible XI: top players by xPts this GW
                            all_gw = [(p["id"], xpts_map.get(p["id"], {}).get(gw, 0))
                                      for _, p in qualified.iterrows()]
                            all_gw.sort(key=lambda x: x[1], reverse=True)
                            best_xi_xpts = sum(x[1] for x in all_gw[:11]) if len(all_gw) >= 11 else 0
                            fh_gain = max(best_xi_xpts - current_gw_xpts, 0)
                            if is_bgw:
                                fh_gain *= 1.5  # FH is extra valuable on BGWs
                            chip_gains[gw]["freehit"] = round(fh_gain, 1)

                            # === ALLSVENSKAN CHIP GAINS ===

                            # PARK THE BUS gain = sum of DEF xPts (doubled) - captain bonus lost
                            if xi is not None and len(xi) > 0:
                                def_xpts = xi[xi["pos_id"] == 2]["id"].map(
                                    lambda pid: xpts_map.get(pid, {}).get(gw, 0)
                                ).sum()
                                ptb_gain = def_xpts * dgw_mult - cap_xpts  # gain DEF double, lose captain
                            else:
                                ptb_gain = 0
                            chip_gains[gw]["park_the_bus"] = round(max(ptb_gain, 0), 1)

                            # DYNAMIC DUO gain = captain ×2 extra + vice captain ×1 extra
                            # (normal captain gives ×1 extra, DD gives ×2 extra = net +1 for cap)
                            # Plus vice captain gets ×1 extra (normally nothing)
                            vice_xpts = 0
                            if xi is not None and len(xi) > 0:
                                # Find 2nd highest xPts player in XI
                                xi_sorted = sorted(
                                    [(pid, xpts_map.get(pid, {}).get(gw, 0)) for pid in xi["id"]],
                                    key=lambda x: x[1], reverse=True
                                )
                                if len(xi_sorted) >= 2:
                                    vice_xpts = xi_sorted[1][1]
                            dd_gain = (cap_xpts + vice_xpts) * dgw_mult  # extra cap + vice bonus
                            chip_gains[gw]["dynamic_duo"] = round(dd_gain, 1)

                            # LOAN RANGERS gain = best possible squad (no team limit) - current XI
                            # Similar to Free Hit but potentially better due to stacking one team
                            lr_gain = fh_gain * 1.2  # rough estimate — stacking adds ~20%
                            if is_dgw:
                                lr_gain *= 1.5  # much more valuable on DGWs (stack DGW team)
                            chip_gains[gw]["loan_rangers"] = round(lr_gain, 1)

                            # FRIKORT gain = same as Wildcard
                            chip_gains[gw]["frikort"] = 0

                            # WILDCARD gain: run WC plan once to measure improvement
                            # Only compute for the top 2 candidate GWs to save time
                            chip_gains[gw]["wildcard"] = 0

                        # Run WC for the 2 most promising GWs (before DGWs or highest xPts weeks)
                        wc_candidate_gws = sorted(plan_gws,
                            key=lambda g: (1 if g + 1 in dgw_gws or g + 2 in dgw_gws else 0,
                                           sum(xpts_map.get(p["id"], {}).get(g, 0)
                                               for _, p in qualified.nlargest(15, "xpts_total").iterrows())),
                            reverse=True)[:2]

                        for wc_gw in wc_candidate_gws:
                            try:
                                wc_plan = build_rolling_plan(
                                    my_squad_chip, df,
                                    bank=team_data_chip["bank"],
                                    free_transfers=ft_chip,
                                    purchase_prices=team_data_chip.get("purchase_prices", {}),
                                    selling_prices_api=team_data_chip.get("selling_prices_api", {}),
                                    xpts_map=xpts_map,
                                    planning_gw_id=planning_gw_id,
                                    n_gws=6,
                                    chip_schedule={wc_gw: "wildcard"},
                                    team_fixture_counts=team_fixture_counts,
                                )
                                if wc_plan:
                                    wc_total = sum(g.get("total_xpts", 0) for g in wc_plan)
                                    chip_gains[wc_gw]["wildcard"] = round(max(wc_total - baseline_total, 0), 1)
                            except Exception:
                                pass

                        # Step 3: Build all valid chip assignments using pre-computed gains
                        # This is instant — just arithmetic
                        from itertools import combinations, permutations

                        all_chip_results = []
                        # No chips baseline
                        all_chip_results.append({
                            "schedule": {},
                            "total": baseline_total,
                            "gw_totals": {g["gw"]: g.get("total_xpts", 0) for g in baseline_plan} if baseline_plan else {},
                        })

                        for n_chips in range(1, len(available_chips) + 1):
                            for chip_subset in combinations(available_chips, n_chips):
                                for gw_perm in permutations(plan_gws, n_chips):
                                    schedule = {}
                                    total_gain = 0
                                    gw_totals = {}
                                    valid = True
                                    for c, g in zip(chip_subset, gw_perm):
                                        planner_name = chip_name_map.get(c, c)
                                        schedule[g] = planner_name
                                        gain = chip_gains.get(g, {}).get(c, 0)
                                        total_gain += gain

                                    # Build approximate gw_totals
                                    for gw in plan_gws:
                                        base = baseline_gw.get(gw, {}).get("total_xpts", 0)
                                        chip_on = schedule.get(gw)
                                        if chip_on:
                                            rev = {v: k for k, v in chip_name_map.items()}
                                            ck = rev.get(chip_on, chip_on)
                                            base += chip_gains.get(gw, {}).get(ck, 0)
                                        gw_totals[gw] = round(base, 1)

                                    all_chip_results.append({
                                        "schedule": schedule,
                                        "total": round(baseline_total + total_gain, 1),
                                        "gw_totals": gw_totals,
                                    })

                        all_chip_results.sort(key=lambda x: x["total"], reverse=True)

                    # Store results in session state so they persist
                    st.session_state["chip_results"] = all_chip_results
                    st.session_state["chip_baseline"] = baseline_total
                    st.session_state["chip_combos_count"] = len(all_chip_results)

                # Display results (from session state if available)
                if "chip_results" in st.session_state and st.session_state["chip_results"]:
                    all_chip_results = st.session_state["chip_results"]
                    baseline_total = st.session_state.get("chip_baseline", 0)
                    combos_count = st.session_state.get("chip_combos_count", 0)

                    st.markdown("### Results")
                    st.markdown(
                        f'<span style="color:#8892a8;">Evaluated <b>{combos_count}</b> combinations · '
                        f'Baseline (no chips): <b>{baseline_total:.1f}</b> xPts</span>',
                        unsafe_allow_html=True,
                    )
                    st.markdown("")

                    chip_name_map = {
                        "wildcard": "wildcard", "freehit": "free_hit",
                        "3xc": "triple_captain", "bboost": "bench_boost",
                    }
                    reverse_map = {v: k for k, v in chip_name_map.items()}
                    # Reverse map for setting session state keys
                    planner_key_map = {
                        "wildcard": "wc_gw", "free_hit": "fh_gw",
                        "triple_captain": "tc_gw", "bench_boost": "bb_gw",
                    }
                    top_results = all_chip_results[:5]

                    for rank, result in enumerate(top_results):
                        is_best = (rank == 0)
                        schedule = result["schedule"]
                        total = result["total"]
                        gain = total - baseline_total

                        if schedule:
                            chip_strs = []
                            for gw in sorted(schedule.keys()):
                                chip_name = schedule[gw]
                                chip_key = reverse_map.get(chip_name, chip_name)
                                label = chip_labels_chip.get(chip_key, chip_name)
                                chip_strs.append(f"{label} GW{gw}")
                            strategy_str = " · ".join(chip_strs)
                        else:
                            strategy_str = "No chips used — save for later"

                        border_col = "#f02d6e" if is_best else "#2a3550"
                        bg = "#111827" if is_best else "#0d1117"
                        rank_label = "⭐ BEST" if is_best else f"#{rank + 1}"

                        col_info, col_btn = st.columns([5, 1])
                        with col_info:
                            st.markdown(
                                f'<div style="background:{bg};border-left:3px solid {border_col};padding:0.7rem 0.8rem;border-radius:0 6px 6px 0;">'
                                f'<span style="color:{"#f02d6e" if is_best else "#5a6580"};font-weight:700;font-size:0.8rem;">{rank_label}</span> '
                                f'<span style="color:#e2e8f0;font-weight:700;font-size:1.1rem;">{total:.1f} xPts</span> '
                                f'<span style="color:#34d399;font-size:0.85rem;">(+{gain:.1f})</span><br>'
                                f'<span style="color:#8892a8;font-size:0.8rem;">{strategy_str}</span></div>',
                                unsafe_allow_html=True,
                            )
                        with col_btn:
                            if schedule and st.button("✅ Apply & Run", key=f"apply_chip_{rank}", use_container_width=True):
                                # Store applied strategy and flag to auto-run the planner
                                st.session_state["applied_chip_schedule"] = dict(schedule)
                                st.session_state["plan_auto_run"] = True
                                st.session_state["navigate_to_my_team"] = True
                                st.rerun()

                    # Per-GW breakdown: compare BEST vs BASELINE
                    if len(all_chip_results) >= 1:
                        st.markdown("")
                        st.markdown("### Per-GW Breakdown (Best Strategy vs No Chips)")
                        best_result = all_chip_results[0]
                        best_gw_totals = best_result.get("gw_totals", {})

                        # Find baseline result (empty schedule)
                        baseline_result = next((r for r in all_chip_results if not r["schedule"]), None)
                        baseline_gw_totals = baseline_result.get("gw_totals", {}) if baseline_result else {}

                        reverse_map_display = {v: k for k, v in chip_name_map.items()}
                        for gw in sorted(best_gw_totals.keys()):
                            best_pts = best_gw_totals.get(gw, 0)
                            base_pts = baseline_gw_totals.get(gw, 0)
                            diff = best_pts - base_pts

                            chip_on = best_result["schedule"].get(gw)
                            if chip_on:
                                chip_key = reverse_map_display.get(chip_on, chip_on)
                                chip_str = f" — {chip_labels_chip.get(chip_key, chip_on)}"
                                diff_color = "#34d399" if diff > 0.5 else "#fbbf24" if diff > 0 else "#8892a8"
                            else:
                                chip_str = ""
                                diff_color = "#8892a8"

                            st.markdown(
                                f'<span style="color:#8892a8;font-size:0.8rem;">'
                                f'GW{gw}: <b style="color:#e2e8f0;">{best_pts:.1f}</b> xPts '
                                f'(baseline: {base_pts:.1f}, '
                                f'<span style="color:{diff_color};">{"+" if diff >= 0 else ""}{diff:.1f}</span>)'
                                f'<span style="color:#f02d6e;font-weight:600;">{chip_str}</span></span>',
                                unsafe_allow_html=True,
                            )

    # === Footer ===
    st.markdown("---")
    st.markdown(
        f"<div style='text-align:center; color:#5a6580; font-size:0.7rem;'>"
        f"Datumly · Data-driven FPL intelligence · "
        f"FPL API + football-data.co.uk + PuLP MILP · "
        f"{datetime.now().strftime('%d %b %Y, %H:%M')}"
        f"</div>",
        unsafe_allow_html=True,
    )


if __name__ == "__main__":
    main()
